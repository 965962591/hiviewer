# -*- coding: utf-8 -*-
import re
import os
import sys
import xml.etree.ElementTree as ET
from pathlib import Path
from openpyxl import Workbook

"""设置本项目的入口路径,BASEPATH"""
# 方法一：手动找寻上级目录，获取项目入口路径
BASEPATH = Path(__file__).parent.parent.parent.as_posix()



def save_excel_data(images_path):
    """将从XML文件中提取的数据保存到Excel表格中"""

    # 配置保存的Excel文件路径
    excel_path = os.path.join(images_path, "extracted_data.xlsx")
    if os.path.exists(excel_path):
        # 若存在excel文件则不需要保存新的excel文件
        return

    # 初始化二维列表，第一行的表头数据
    get_excel_list = [
        [
            "文件名",
            "Lux",
            "BL",
            "DRCgain",
            "Safe_gain",
            "Short_gain",
            "Long_gain",
            "CCT",
            "R_gain",
            "B_gain",
            "Awb_sa",
            "Triangle_index",
            "AE",   # AE 问题点
            "AWB",  # AWB 问题点    
            "ISP",  # ISP 问题点
            "AF",   # AF 问题点
        ]
    ]

    # 遍历文件夹，列出所有满足条件的xml文件
    xml_files = [f for f in os.listdir(images_path) if f.endswith('_new.xml')]

    # 遍历xml文件
    for xml_file in xml_files:
        xml_file = str(Path(images_path) / xml_file)
        if os.path.exists(xml_file):
            # 使用函数load_xml_data加载并解析xml文件
            result_list = load_xml_data(xml_file)
            if result_list:
                get_excel_list.append(result_list)

    # 创建一个新的工作簿
    wb = Workbook()
    ws = wb.active

    # 设置边框样式
    from openpyxl.styles import Border, Side
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 写入数据并设置列宽和边框
    for row in get_excel_list:
        ws.append(row)
        for cell in row:
            # 设置边框
            ws.cell(row=ws.max_row, column=row.index(cell) + 1).border = border

    # 自适应列宽
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter  # 获取列字母
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)  # 加2以增加一些额外的空间
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(excel_path)
    print(f"[save_excel_data]-->数据成功写入到 {excel_path}")


def save_excel_data_by_unisoc(images_path):
    """将从XML文件中提取的数据保存到Excel表格中"""
    # 配置保存的Excel文件路径,若存在则返回
    if (excel_path := Path(images_path, "extracted_data.xlsx")).exists():
        return

    # 初始化二维列表，第一行的表头数据
    get_excel_list = [
        [
            "文件名",
            "BV",
            "EVD",
            "Backlight",
            "LCGgain",
            "Stable",
            "Mulaes_thd",
            "HM_thd",
            "Face_thd",
            "AE",   # AE 问题点
            "AWB",  # AWB 问题点    
            "ISP",  # ISP 问题点
            "AF",   # AF 问题点
        ]
    ]

    # 遍历文件夹，列出所有满足条件的xml文件
    txt_files = [f for f in os.listdir(images_path) if f.endswith('.txt')]

    # 遍历xml文件，使用函数load_xml_data加载并解析xml文件
    for txt_file in txt_files:
        if (txt_file := Path(images_path, txt_file)).exists():
            if result_list := load_txt_data_by_unisoc(txt_file.as_posix()):
                get_excel_list.append(result_list)

    # 创建一个新的工作簿
    wb = Workbook()
    ws = wb.active

    # 设置边框样式
    from openpyxl.styles import Border, Side
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 写入数据并设置列宽和边框
    for row in get_excel_list:
        ws.append(row)
        for cell in row:
            # 设置边框
            ws.cell(row=ws.max_row, column=row.index(cell) + 1).border = border

    # 自适应列宽
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter  # 获取列字母
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)  # 加2以增加一些额外的空间
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(excel_path.as_posix())
    print(f"[save_excel_data_by_unisoc]-->数据成功写入到 {excel_path.as_posix()}")

def load_xml_data(xml_path):
    """加载XML文件并提取Lux值和DRCgain值等EXIF信息"""
    try:
        # 加载xml文件
        tree = ET.parse(xml_path)
        root = tree.getroot()

        """提取背光值"""
        # 获取FrameSA的luma值
        luma_frame = root.find('.//SA/FrameSA/luma').text if root.find('.//SA/FrameSA/luma') is not None else None
        luma_frame_ev = root.find('.//SA/EVFrameSA/luma').text if root.find('.//SA/EVFrameSA/luma') is not None else None
        frame_luma = luma_frame if luma_frame else luma_frame_ev if luma_frame_ev else 0.0001
        
        # 获取FaceSA的luma值
        luma_face = root.find('.//SA/FaceSA/luma').text if root.find('.//SA/FaceSA/luma') is not None else None
        face_luma = luma_face if luma_face else 0.0001

        # 计算背光值
        backlight = float(face_luma)/float(frame_luma) if frame_luma and face_luma else 0.0

        # 提取值并转换为列表
        result_list = [
            # 文件名
            str(os.path.basename(xml_path).split('_new.xml')[0]+".jpg"),
            # Lux
            float(root.find('lux_index').text) if root.find('lux_index') is not None else None, 
            # BL
            backlight,
            # DRCgain
            root.find('DRCgain').text if root.find('DRCgain') is not None else None,
            # Safe_gain
            float(root.find('safe_gain').text) if root.find('safe_gain') is not None else None,
            # Short_gain
            float(root.find('short_gain').text) if root.find('short_gain') is not None else None,
            # Long_gain
            float(root.find('long_gain').text) if root.find('long_gain') is not None else None,
            # CCT
            float(root.find('CCT').text) if root.find('CCT') is not None else None,
            # R_gain
            float(root.find('r_gain').text) if root.find('r_gain') is not None else None,
            # B_gain
            float(root.find('b_gain').text) if root.find('b_gain') is not None else None,
            # Awb_sa
            root.find('awb_sa').text if root.find('awb_sa') is not None else None,
            # Triangle_index
            float(root.find('triangle_index').text) if root.find('triangle_index') is not None else None,
        ]

        return result_list
        
    except Exception as e:
        print(f"解析XML失败{xml_path}:\n {e}")
        return None


def load_txt_data_by_unisoc(txt_path):
    """加载XML文件并提取Lux值和DRCgain值等EXIF信息"""
    try:
        # 读取txt文件
        text_norm = ""
        with open(txt_path, "r", encoding="utf-8", errors="ignore") as f:
            text_norm = f.read()


        # 关键字段查找
        patterns = {
            "Lux": "AE-cur_bv|AE-hm-hm_evd|AE-face-calc_fd_param-calc_face_luma-face_backlight|AE-ae_stable",
            "Mulaes":"AE-mulae_target|AE-mulae-mulae_thd|AE-mulae-mulae_y|AE-mulae-cur_lum",
            "HM":("AE-short_hm_target|AE-safe_hm_target|"
                  "AE-hm-short_hm-hm_final_target_min|AE-hm-short_hm-hm_final_target_max|"
                  "AE-hm-safe_hm-hm_final_target_min|AE-hm-safe_hm-hm_final_target_max|"
                  "AE-hm-short_hm-hm_bt_target|AE-hm-short_hm-hm_aftaoe_target|AE-hm-short_hm-hm_aftcoe_target|AE-hm-short_hm-hm_dt_target|"
                  "AE-hm-short_hm-hm_dt_target_min|AE-hm-short_hm-hm_dt_target_max|"
                  "AE-hm-safe_hm-hm_bt_target|AE-hm-safe_hm-hm_aftaoe_target|AE-hm-safe_hm-hm_aftcoe_target|AE-hm-safe_hm-hm_dt_target|"
                  "AE-hm-safe_hm-hm_dt_target_min|AE-hm-safe_hm-hm_dt_target_max"
                 ),
            "Face":("AE-face-face_num|AE-face-calc_fd_param-face_target-short_face_thd|AE-face-calc_fd_param-face_target-safe_face_thd|"
                    "AE-face-calc_fd_param-face_target-short_final_face_luma|AE-face-calc_fd_param-face_target-safe_final_face_luma|"
                    "AE-face-cur_lum|AE-short_face_target|AE-safe_face_target|AE-face-calc_fd_param-face_target-min_facelum_protection_target|"
                    "AE-face-calc_fd_param-face_target-short_down_limit|AE-face-calc_fd_param-face_target-safe_down_limit|"
                    "AE-face-calc_fd_param-face_target-short_up_limit|AE-face-calc_fd_param-face_target-safe_up_limit|"
                    "AE-face-calc_fd_param-face_target-short_face_target_before_mflumtype1|AE-face-calc_fd_param-face_target-safe_face_target_before_mflumtype2"
                 ),
            "LCG": "AE-safe_final_target_lum|AE-short_final_target_lum|AE-ae_lcg|AE-ae_lcg_down_limit|AE-ae_lcg_up_limit",
        }

        # 定义提取关键字函数
        def extract_value_pat(pat, text_norm):
            try:
                value = 0
                format = rf'{re.escape(pat)}\s*:\s*([-+]?\d+(?:\.\d+)?)'
                match = re.search(format, text_norm)
                if match:
                    value = float(match.group(1)) if '.' in match.group(1) else int(match.group(1))
                return value if value else 0
            except Exception as e:
                print(f"extract_value_pat解析TXT失败{txt_path}:\n报错信息: {e}")
                return 0

        # 查找关键字数值并拼接, 先定义基础变量
        bv = extract_value_pat("AE-cur_bv", text_norm)
        evd = extract_value_pat("AE-hm-hm_evd", text_norm)
        bl = extract_value_pat("AE-face-calc_fd_param-calc_face_luma-face_backlight", text_norm)
        stb = extract_value_pat("AE-ae_stable", text_norm)
        lcg = extract_value_pat("AE-ae_lcg", text_norm)
        Mulaes = extract_value_pat("AE-mulae-mulae_thd", text_norm)
        Hm_safe_thd = extract_value_pat("AE-hm-safe_hm-hm_bt_thd", text_norm)
        Hm_short_thd = extract_value_pat("AE-hm-short_hm-hm_bt_thd", text_norm)
        Face_safe_thd = extract_value_pat("AE-face-calc_fd_param-face_target-safe_face_thd", text_norm)
        Face_short_thd = extract_value_pat("AE-face-calc_fd_param-face_target-short_face_thd", text_norm)

        # 提取值并转换为列表
        result_list = [
            # 文件名
            Path(txt_path).stem,
            # bv
            float(bv) if bv is not None else None, 
            # evd
            float(evd) if evd is not None else None, 
            # bl
            float(bl) if bl is not None else None, 
            # lcg
            float(lcg) if lcg is not None else None, 
            # stable
            float(stb) if stb is not None else None, 
            # Mulaes
            float(Mulaes) if Mulaes is not None else None, 
            # HM
            f"[short:{Hm_short_thd}, safe:{Hm_safe_thd}]", 
            # Face
            f"[short:{Face_short_thd}, safe:{Face_safe_thd}]", 
        ]

        return result_list
        
    except Exception as e:
        print(f"解析展锐txt数据失败{txt_path}:\n {e}")
        return None
