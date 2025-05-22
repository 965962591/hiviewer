"""导入python内置模块"""
import re
import os
import io
import gc
import sys
import time
import json
import pathlib
import threading
from multiprocessing import cpu_count
from concurrent.futures import ThreadPoolExecutor

"""导入python第三方模块"""
import cv2
import piexif
import openpyxl
import numpy as np
import win32com.client as win32
import matplotlib.pyplot as plt
from lxml import etree as ETT
from PIL import Image, ImageCms
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtGui import QIcon, QColor, QPixmap, QKeySequence, QPainter, QCursor, QTransform, QImage, QPen
from PyQt5.QtCore import Qt, QEvent, pyqtSignal, QTimer, QThreadPool, QRunnable
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QLabel, QHeaderView, QShortcut, QGraphicsView, 
    QGraphicsScene, QGraphicsPixmapItem, QMessageBox, QProgressBar, QGraphicsRectItem, 
    QGraphicsItem, QDialogButtonBox, QDialog, QVBoxLayout, QHBoxLayout, QPushButton, QLineEdit, QCheckBox, QComboBox, QFileDialog)

"""导入自定义模块"""
from src.components.UiSub import Ui_MainWindow                              # 看图子界面，导入界面UI
from src.components.QMessageBox import show_message_box                     # 导入消息框类
from src.common.SettingInit import load_exif_settings,load_color_settings   # 导入json配置模块
from src.common.FontManager import SingleFontManager                        # 看图子界面，导入字体管理器
from src.utils.aitips import CustomLLM_Siliconflow                          # 看图子界面，AI提示看图复选框功能模块
from src.utils.hisnot import WScreenshot                                    # 看图子界面，导入自定义截图的类
from src.utils.aeboxlink import check_process_running,get_api_data          # 导入与AEBOX通信的模块函数
from src.utils.heic import extract_jpg_from_heic                            # 导入heic图片转换为jpg图片的模块

"""设置本项目的入口路径,全局变量BasePath"""
# 方法一：手动找寻上级目录，获取项目入口路径，支持单独运行该模块
if True:
    BasePath = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
# 方法二：直接读取主函数的路径，获取项目入口目录,只适用于hiviewer.py同级目录下的py文件调用
if False: # 暂时禁用，不支持单独运行该模块
    BasePath = os.path.dirname(os.path.abspath(sys.argv[0]))    

"""
设置全局函数区域开始线
------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
"""

def convert_to_dict(exif_string):
    """将字符串转换为字典
    Args:
        exif_string: str,
            '曝光时间: 1/100_(10000.0ms)\nISO值: 75\n图片名称: IMG_20250325_053138.jpg\n图片尺寸: 2448 x 3264\n
            图片张数: 3/9\nHDR: off\nZoom: 1\nLux: 174.8547\nCCT: 5110.869\nDRCgain: 1.411515 / 0.9743042 = 1.45\n
            Awb_sa: indoor yellow0319,FACE Assist\nTriangle_index: 5\nR_gain: 1.792648, B_gain: 1.619696\n
            Safe_gain: 2.268417, Short_gain: 1.499691, Long_gain: 2.268417'

        return: dict ,
            '曝光时间' ='1/100_(10000.0ms)'
            'ISO值' ='75'
            '图片名称' ='IMG_20250325_053138.jpg'
            '图片尺寸' ='2448 x 3264'
            '图片张数' ='3/9'
            'HDR' ='off'
            'Zoom' ='1'
            'Lux' ='174.8547'
            'CCT' ='5110.869'
            'DRCgain' ='1.411515 / 0.9743042 = 1.45'
            'Awb_sa' ='indoor yellow0319,FACE Assist'
            'Triangle_index' ='5'
            'R_gain' ='1.792648, B_gain: 1.619696'
            'Safe_gain' ='2.268417, Short_gain: 1.499691, Long_gain: 2.268417'
    """
    # 使用正则表达式匹配键值对
    pattern = r'([^:]+): ([^\n]+)'
    matches = re.findall(pattern, exif_string)
    return {key.strip(): value.strip() for key, value in matches}


def rgb_str_to_qcolor(rgb_str):
    """将 'rgb(r,g,b)' 格式的字符串转换为 QColor"""
    # 提取RGB值
    rgb = rgb_str.strip('rgb()')  # 移除 'rgb()' 
    r, g, b = map(int, rgb.split(','))  # 分割并转换为整数
    return QColor(r, g, b)


def qcolor_to_rgb_str(qcolor):
    """将 QColor 转换为 'rgb(r,g,b)' 格式的字符串"""
    return f"rgb({qcolor.red()}, {qcolor.green()}, {qcolor.blue()})"


def imread_chinese(path):
    """支持中文路径的图片读取函数"""
    # 使用二进制读取+解码，跨平台支持
    try:
        with open(path, 'rb') as f:
            data = np.frombuffer(f.read(), dtype=np.uint8)
            img = cv2.imdecode(data, cv2.IMREAD_COLOR)
            # 自动转换颜色通道（可选）
            # img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB) 
            return img
    except Exception as e:
        print(f"读取图片失败: {e}")
        return None


def calculate_image_stats(image_input, resize_factor=1):
    """使用OpenCV计算图片的亮度、RGB、LAB和对比度"""
    # modify by diamond_cz 20250409 移除LAB计算，添加R/G和B/G计算
    try:
        # 类型判断分支处理，支持传入文件路径和PIL图像
        if isinstance(image_input, str):  # 处理文件路径
            with open(image_input, 'rb') as f:
                image_data = np.frombuffer(f.read(), dtype=np.uint8)
                img = cv2.imdecode(image_data, cv2.IMREAD_COLOR)
        elif isinstance(image_input, Image.Image):  # 处理PIL图像对象
            # 转换PIL图像到OpenCV格式
            img = np.array(image_input.convert('RGB'))[:, :, ::-1].copy()
        elif isinstance(image_input, np.ndarray):  # 处理PIL图像对象
            # 传入的是opencv格式图
            # print("image_input is ok!!!")
            img = image_input
        else:
            print(f"calculate_image_stats无法加载图像, 当前图像格式:{type(image_input)}")
            return None

        # 缩小图片
        height, width = img.shape[:2]
        new_size = (int(width * resize_factor), int(height * resize_factor))
        img = cv2.resize(img, new_size, interpolation=cv2.INTER_LANCZOS4)

        # 将BGR转换为RGB
        img_rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        
        # 将RGB转换为LAB
        img_lab = cv2.cvtColor(img_rgb, cv2.COLOR_RGB2LAB)
        
        # 计算RGB和LAB平均值
        avg_rgb = np.mean(img_rgb, axis=(0, 1))
        avg_lab = np.mean(img_lab, axis=(0, 1))

        # 将LAB值调整到标准范围
        avg_lab = (
            avg_lab[0] * (100 / 255),   # L: [0, 255] -> [0, 100]
            avg_lab[1] - 128,           # A: [0, 255] -> [-127, 127]
            avg_lab[2] - 128            # B: [0, 255] -> [-127, 127]
        )

        # 计算R/G和B/G
        R_G = avg_rgb[0] / avg_rgb[1] if avg_rgb[1] != 0 else float('inf')  # 避免除以零
        B_G = avg_rgb[2] / avg_rgb[1] if avg_rgb[1] != 0 else float('inf')  # 避免除以零

        # 计算亮度（直接使用RGB值）
        avg_brightness = 0.299 * avg_rgb[0] + 0.587 * avg_rgb[1] + 0.114 * avg_rgb[2]
        
        # 计算全局对比度（使用LAB的L通道标准差）
        l_channel = img_lab[:, :, 0].astype(np.float32) * (100 / 255)  # 转换为标准L范围[0,100]
        contrast = np.std(l_channel)  # 标准差作为对比度指标

        # 格式化输出
        return {
            'width': new_size[0],   # 新增区域宽度
            'height': new_size[1],  # 新增区域高度
            'avg_brightness': round(float(avg_brightness), 1),
            'contrast': round(float(contrast), 1),  # 新增对比度指标
            'avg_rgb': tuple(round(float(x), 1) for x in avg_rgb),
            'avg_lab': tuple(round(float(x), 1) for x in avg_lab),
            'R_G': round(float(R_G), 5),  # 新增R/G计算结果
            'B_G': round(float(B_G), 5)   # 新增B/G计算结果
        }
        
    except Exception as e:
        print(f"calculate_image_stats计算图片统计信息失败, 错误: {e}")
        return None


def convert_to_dci_p3(cv_img, source_img, original_pixmap):
    """将QPixmap转换为DCI-P3色域
    Args:
        cv_img: Opencv图像
        source_img: PIL图像
        original_pixmap: pixmap原图
    Returns:
        转换后的QPixmap对象
    """
    try:

        if source_img:
            color_space, _ = get_color_profile(source_img)
            print(f"设备识别: {color_space}")
            
            # 根据设备选择转换矩阵
            if "Display P3" in color_space:
                # Apple Display P3到DCI-P3的转换矩阵
                matrix = np.array([
                    [1.2249, -0.2247, 0.0000],
                    [-0.0420, 1.0419, 0.0000],
                    [-0.0197, -0.0786, 1.0973]
                ])
            elif "Adobe RGB" in color_space:
                matrix = np.array([
                    [0.7152, 0.2848, 0.0000],
                    [0.0000, 1.0000, 0.0000],
                    [0.0000, 0.0000, 0.9999]
                ])
            else:  # 默认sRGB
                matrix = np.array([
                    [0.4865, 0.2657, 0.1982],
                    [0.2289, 0.6917, 0.0793],
                    [0.0000, 0.0451, 1.0439]
                ])

        if True:  # 识别设备进行转换
            # DCI-P3转换矩阵
            dci_p3_matrix = matrix

            # 确保cv_img是3通道的RGB图像
            if cv_img.shape[2] == 4:  # 如果是4通道（RGBA）
                rgb_img = cv2.cvtColor(cv_img, cv2.COLOR_RGBA2RGB)
            elif cv_img.shape[2] == 1:  # 如果是单通道（灰度）
                rgb_img = cv2.cvtColor(cv_img, cv2.COLOR_GRAY2RGB)
            else:  # 假设已经是3通道（RGB）
                rgb_img = cv2.cvtColor(cv_img, cv2.COLOR_BGR2RGB)
            
            # 应用矩阵变换并限制数值范围
            dci_p3_img = cv2.transform(rgb_img, dci_p3_matrix)
            dci_p3_img = np.clip(dci_p3_img, 0, 255).astype(np.uint8)

        if False and cv_img is not None:  # 强制转换,效果不佳，待研究
 
            # 定义 XYZ 到 P3 的转换矩阵
            xyz_to_p3 = np.array([
                [1.2249044, -0.2246222, 0.0158858],
                [-0.0420464, 1.1472752, -0.1052288],
                [-0.0134924, -0.0346351, 1.0481275]
            ])

            # 将 sRGB 转换为 XYZ
            image_xyz = cv2.cvtColor(cv_img, cv2.COLOR_BGR2XYZ)

            # 将 XYZ 转换为 P3
            dci_p3_img = np.dot(image_xyz, xyz_to_p3.T)

        
        # 转换回QImage
        height, width, _ = dci_p3_img.shape
        bytes_per_line = 3 * width
        converted_qimage = QImage(
            dci_p3_img.data, 
            width, 
            height,
            bytes_per_line,
            QImage.Format_RGB888
        )

        pixmap = QPixmap.fromImage(converted_qimage)

        return pixmap
    
    except Exception as e:
        print(f"convert_to_dci_p3()-DCI-P3转换失败: {str(e)}")
        return original_pixmap
    
def get_color_profile(img):
    """增强版色彩空间检测，适配移动设备"""
    try:
        # 获取EXIF制造商信息
        exif = img.getexif()
        make = exif.get(271, "").strip().lower()  # Make标签
        model = exif.get(272, "").strip().lower() # Model标签

        # 方法1：优先处理已知设备的特殊逻辑
        def check_device_space():
            if 'apple' in make or 'iphone' in model:
                return "Display P3 (Apple)", True
            if 'xiaomi' in make or 'mi' in make or '22122' in model:  # 小米14U型号包含22122
                # 检查专业模式标签（小米特有EXIF）
                xiaomi_mode = exif.get(0xB001, 0)
                if xiaomi_mode == 2:  # 专业模式通常使用Adobe RGB
                    return "Adobe RGB (Xiaomi Pro)", True
                return "sRGB (Xiaomi)", False
            return None

        # 方法2：解析ICC配置文件
        def check_icc_profile():
            if "icc_profile" in img.info:
                icc_data = img.info["icc_profile"]
                icc_file = io.BytesIO(icc_data)
                try:
                    profile = ImageCms.getProfile(icc_file)
                    desc = ImageCms.getProfileDescription(profile).lower()
                    if 'display p3' in desc:
                        return "Display P3", True
                    if 'adobe' in desc:
                        return "Adobe RGB", True
                    if 'dci-p3' in desc:
                        return "DCI-P3", True
                except Exception as e:
                    print(f"ICC解析异常: {str(e)}")
            return None

        # 方法3：EXIF色彩空间标签
        def check_exif_space():
            color_space = exif.get(40961, 1)
            if color_space == 1:
                return "sRGB", False
            elif color_space == 2:
                return "Adobe RGB", False
            elif color_space == 65535:  # 未校准
                if 'apple' in make:
                    return "Display P3 (Uncalibrated)", False
                return "Uncalibrated", False
            return None

        # 检测优先级：设备特征 > ICC > EXIF > 默认
        for checker in [check_device_space, check_icc_profile, check_exif_space]:
            result = checker()
            if result: 
                return result
                
        return "sRGB (assumed)", False

    except Exception as e:
        print(f"get_color_profile()-色彩检测异常: {str(e)}")
        return "Unknown", False
    
def close_excel():
    "强制关闭一个EXCEL表格"
    try:
        # 获取 Excel 应用程序的实例
        excel_app = win32.gencache.EnsureDispatch('Excel.Application')
        
        # 检查是否有打开的工作簿
        if excel_app.Workbooks.Count > 0:
            # 强制关闭所有工作簿，不保存更改
            for workbook in excel_app.Workbooks:
                workbook.Close(SaveChanges=False)
        
        # 退出 Excel 应用程序
        excel_app.Quit()
        print("Excel 已成功关闭。")
        
    except Exception as e:
        print(f"关闭 Excel 时发生错误: {e}")


def load_xml_data(xml_path):
    """加载XML文件并提取Lux值和DRCgain值等EXIF信息"""
    try:
        # 加载xml文件 
        tree = ETT.parse(xml_path)
        root = tree.getroot()

        # 定义需要提取的标签, tag:name
        XPATHS = {
            'Lux': ETT.XPath('lux_index'),
            'DRCgain': ETT.XPath('DRCgain'),
            'Safe_gain': ETT.XPath('safe_gain'),
            'Short_gain': ETT.XPath('short_gain'),
            'Long_gain': ETT.XPath('long_gain'),
            'CCT': ETT.XPath('CCT'),
            'R_gain': ETT.XPath('r_gain'),
            'B_gain': ETT.XPath('b_gain'),
            'Awb_sa': ETT.XPath('awb_sa'),
            'Triangle_index': ETT.XPath('triangle_index'),
            'FaceSA': ETT.XPath('.//SA/FaceSA')
        }

        # ETT.XPath('lux_index')(root)[0].text
        # 提取值并拼接
        extracted_values = []
        for name, tag  in XPATHS.items():
            value = tag(root)
            if name  != 'FaceSA':
                if value and value[0].text:
                    extracted_values.append(f"\n{name}: {value[0].text}")
            else: # 解析人脸SA的相关value
                if value:
                    # 获取FrameSA的luma值
                    luma_frame = ETT.XPath('.//SA/FrameSA/luma')(root)
                    luma_frame_ev = ETT.XPath('.//SA/EVFrameSA/luma')(root)
                    frame_luma = luma_frame[0].text if luma_frame and luma_frame[0].text else luma_frame_ev[0].text if luma_frame_ev and luma_frame_ev[0].text else 0.0001
                    
                    # 获取FaceSA的luma值
                    luma_face = ETT.XPath('.//SA/FaceSA/luma')(root)
                    face_luma = luma_face[0].text if luma_face and luma_face[0].text else 0.0001

                    # 计算背光值
                    backlight = float(face_luma)/float(frame_luma) if frame_luma and face_luma else 0.0
                    
                    # 获取FaceSA的target值
                    target = ETT.XPath('.//SA/FaceSA/target/start')(root)
                    if target and target[0].text:
                        extracted_values.append(f"\n{name}: {target[0].text}(target) & {backlight:.4f}(backlight)")
                else:
                    extracted_values.append(f"\n{name}: 未识别人脸")

        # 汇总字符串    
        qualcom_exif_info = ''.join(extracted_values)
        return qualcom_exif_info, bool(luma_frame_ev)

    except Exception as e:
        print(f"解析XML失败{xml_path}:\n {e}")
        return None
    

def get_aebox_host():
    """读取aebox连接配置"""
    config_path = os.path.join(BasePath, "cache", "aebox_link_host.json")
    default_host = "http://127.0.0.1:8000"
    
    try:
        if os.path.exists(config_path):
            with open(config_path, 'r', encoding='utf-8') as f:
                config = json.load(f)
                return config.get('host', default_host)
        else:
            # 创建默认配置
            os.makedirs(os.path.dirname(config_path), exist_ok=True)
            with open(config_path, 'w', encoding='utf-8') as f:
                json.dump({"host": default_host}, f, indent=4)
            return default_host
    except Exception as e:
        print(f"读取aebox配置失败，使用默认值: {str(e)}")
        return default_host  


"""
设置全局函数区域结束线
------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
"""


"""
设置独立封装类区域开始线
------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
"""

class CameraTestDialog(QDialog):
    """自定义对话框类, 用于输入信息"""
    def __init__(self, images_path_list, parent=None):
        super().__init__(parent)

        # 初始化对话框UI
        self.init_ui()

        # 根据传入的图片路径列表设置关联图片下拉框；
        self.images_path_list = images_path_list

        # 设置是否在同一文件夹下 self.is_same_folder
        if self.images_path_list:
            self.is_same_folder = len(set([os.path.dirname(path) for path in self.images_path_list])) == 1
        
        # 先判断是否存在变量self.is_same_folder
        if hasattr(self, "is_same_folder"): 
            # 判断self.is_same_folder是否为True
            if self.is_same_folder:
                # 将 上一级文件夹名+图片名称 设置到 combo_box0 中
                image_names_all = [os.path.relpath(path, start=os.path.dirname(os.path.dirname(path))) for path in self.images_path_list]
                self.combo_box0.addItems(image_names_all)
            else:
                parent_folders = [os.path.basename(os.path.dirname(path)) for path in self.images_path_list]
                image_names = [os.path.basename(path) for path in self.images_path_list]
                # 将 上一级文件夹名 设置到 combo_box0 中
                self.combo_box0.addItems(parent_folders)
                # 设置字典将上一级文件夹名和图片名称对应起来
                self.parent_folder_dict = dict(zip(parent_folders, image_names))
        else:
            print(f"类CameraTestDialog: 传入的图片路径列表为空, 请检查传入的图片路径列表")


        # 设置是否加载设置
        self.load_settings()

        # 连接按钮信号
        self.button_box.accepted.connect(self.accept)
        self.button_box.rejected.connect(self.reject)
        self.combo_box0.activated.connect(self.update_input0_text)  # 连接 combo_box0 的信号
        self.load_button.clicked.connect(self.load_data)            # 连接加载按钮信号
        self.write_button.clicked.connect(self.write_data)          # 连接写入按钮信号
        self.refresh_button.clicked.connect(self.refresh_data)      # 连接汇总按钮信号
        self.save_button.clicked.connect(self.save_data)            # 连接保存按钮信号
        self.finished.connect(self.save_settings)                   # 连接关闭信号

    def init_ui(self):
        """初始化对话框UI"""

        # 设置窗口标题
        self.setWindowTitle("Camera Test 问题点记录")
        # 设置窗口大小
        self.setFixedSize(1200, 600)  # 设置对话框大小
        
        # 初始化字体管理器，标签组件使用;设置全局变量，定义项目基础路径
        # BasePath = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))) # 使用全局变量 BasePath
        font_path_jetbrains = os.path.join(BasePath, "resource", "fonts", "JetBrainsMapleMono_Regular.ttf")
        self.font_manager_jetbrains_big = SingleFontManager.get_font(size=12, font_path=font_path_jetbrains) 
        self.font_manager_jetbrains_small = SingleFontManager.get_font(size=10, font_path=font_path_jetbrains)

        # 创建主布局
        self.layout = QVBoxLayout(self)

        # 统一的下拉框高度
        combo_box_height = 35

        # 第零行：标签 + 下拉框 + 输入框
        layout_zero = QHBoxLayout()
        self.label0 = QLabel("关联图片项:", self)
        self.label0.setFont(self.font_manager_jetbrains_big)
        self.combo_box0 = QComboBox(self)
        self.combo_box0.setFont(self.font_manager_jetbrains_small)
        self.combo_box0.setFixedHeight(combo_box_height)  # 设置下拉框高度
        self.text_input0 = QLineEdit(self)
        self.text_input0.setFont(self.font_manager_jetbrains_small)
        self.text_input0.setFixedHeight(combo_box_height)  # 设置输入框高度
        layout_zero.addWidget(self.label0)
        layout_zero.addWidget(self.combo_box0)
        layout_zero.addWidget(self.text_input0)
        # 设置比例
        layout_zero.setStretch(0, 1)  # label0 的比例
        layout_zero.setStretch(1, 4)  # combo_box0 的比例
        layout_zero.setStretch(2, 6)  # text_input0 的比例
        self.layout.addLayout(layout_zero)
        
        # 第一行：标签 + 输入框 + 加载按钮 + 写入按钮
        layout_one = QHBoxLayout()
        self.label1 = QLabel("问题点路径:", self)
        self.label1.setFont(self.font_manager_jetbrains_big)
        self.text_input1 = QLineEdit(self)
        self.text_input1.setFixedHeight(combo_box_height)  # 设置下拉框高度
        self.text_input1.setFont(self.font_manager_jetbrains_small)
        self.text_input1.setPlaceholderText("输入保存问题点的EXCEL路径...")  # 设置提示文本
        self.load_button = QPushButton("加载", self)
        self.load_button.setFont(self.font_manager_jetbrains_big)
        self.load_button.setFixedHeight(combo_box_height)  # 设置下拉框高度
        self.write_button = QPushButton("写入", self)
        self.write_button.setFont(self.font_manager_jetbrains_big)
        self.write_button.setFixedHeight(combo_box_height)  # 设置下拉框高度
        layout_one.addWidget(self.label1)
        layout_one.addWidget(self.text_input1)
        layout_one.addWidget(self.load_button)
        layout_one.addWidget(self.write_button)
        # 设置比例
        layout_one.setStretch(0, 1)   # label1 的比例
        layout_one.setStretch(1, 10)  # combo_box1 的比例
        layout_one.setStretch(2, 1)   # load_button 的比例
        layout_one.setStretch(3, 1)   # write_button 的比例
        self.layout.addLayout(layout_one)

        # 第二行：标签 + 输入框 + 汇总按钮 + 保存按钮
        layout_two = QHBoxLayout()
        self.label2 = QLabel("问题点汇总:", self)
        self.label2.setFont(self.font_manager_jetbrains_big)
        self.text_input2 = QLineEdit(self)
        self.text_input2.setFixedHeight(combo_box_height)  # 设置下拉框高度
        self.text_input2.setFont(self.font_manager_jetbrains_small)
        self.text_input2.setPlaceholderText("汇总显示各模块问题点...")  # 设置提示文本
        self.refresh_button = QPushButton("汇总", self)
        self.refresh_button.setFont(self.font_manager_jetbrains_big)
        self.refresh_button.setFixedHeight(combo_box_height)  # 设置下拉框高度
        self.save_button = QPushButton("保存", self)
        self.save_button.setFont(self.font_manager_jetbrains_big)
        self.save_button.setFixedHeight(combo_box_height)  # 设置下拉框高度
        layout_two.addWidget(self.label2)
        layout_two.addWidget(self.text_input2)
        layout_two.addWidget(self.refresh_button)
        layout_two.addWidget(self.save_button)
        # 设置比例
        layout_two.setStretch(0, 1)   # label2 的比例
        layout_two.setStretch(1, 10)  # combo_box2 的比例
        layout_two.setStretch(2, 1)   # refresh_button 的比例
        layout_two.setStretch(3, 1)   # save_button 的比例
        self.layout.addLayout(layout_two)

        # 第三行：复选框 + 输入框
        layout_three = QHBoxLayout()
        self.checkbox1 = QCheckBox("AE", self)
        self.checkbox1.setFont(self.font_manager_jetbrains_small)
        # 设置复选框的初始状态
        self.checkbox1.setChecked(True)
        self.combo_box3 = QComboBox(self)
        self.combo_box3.setFixedHeight(combo_box_height)  # 设置下拉框高度
        self.combo_box3.setFont(self.font_manager_jetbrains_small)
        self.combo_box3.setEditable(True)  
        layout_three.addWidget(self.checkbox1)
        layout_three.addWidget(self.combo_box3)
        # 设置比例
        layout_three.setStretch(0, 1)
        layout_three.setStretch(1, 10) 
        self.layout.addLayout(layout_three)

        # 第四行：复选框 + 输入框
        layout_four = QHBoxLayout()
        self.checkbox2 = QCheckBox("AWB", self)
        self.checkbox2.setFont(self.font_manager_jetbrains_small)
        # 设置复选框的初始状态
        self.checkbox2.setChecked(True)
        self.combo_box4 = QComboBox(self)
        self.combo_box4.setFixedHeight(combo_box_height)  # 设置下拉框高度
        self.combo_box4.setFont(self.font_manager_jetbrains_small)
        self.combo_box4.setEditable(True)  
        layout_four.addWidget(self.checkbox2)
        layout_four.addWidget(self.combo_box4)
        # 设置比例
        layout_four.setStretch(0, 1)
        layout_four.setStretch(1, 10)  
        self.layout.addLayout(layout_four)

        # 第五行：复选框 + 输入框
        layout_five = QHBoxLayout()
        self.checkbox3 = QCheckBox("AF", self)
        self.checkbox3.setFont(self.font_manager_jetbrains_small)
        # 设置复选框的初始状态
        self.checkbox3.setChecked(True)
        self.combo_box5 = QComboBox(self)
        self.combo_box5.setFixedHeight(combo_box_height)  # 设置下拉框高度
        self.combo_box5.setFont(self.font_manager_jetbrains_small)
        self.combo_box5.setEditable(True) 
        layout_five.addWidget(self.checkbox3)
        layout_five.addWidget(self.combo_box5)
        # 设置比例
        layout_five.setStretch(0, 1)
        layout_five.setStretch(1, 10)  
        self.layout.addLayout(layout_five)

        # 第六行：复选框 + 输入框
        layout_six = QHBoxLayout()
        self.checkbox4 = QCheckBox("ISP", self)
        self.checkbox4.setFont(self.font_manager_jetbrains_small)
        # 设置复选框的初始状态
        self.checkbox4.setChecked(True)
        self.combo_box6 = QComboBox(self)
        self.combo_box6.setEditable(True)  
        self.combo_box6.setFixedHeight(combo_box_height)  # 设置下拉框高度
        self.combo_box6.setFont(self.font_manager_jetbrains_small)
        layout_six.addWidget(self.checkbox4)
        layout_six.addWidget(self.combo_box6)
        # 设置比例
        layout_six.setStretch(0, 1)
        layout_six.setStretch(1, 10)  
        self.layout.addLayout(layout_six)

        # 添加确认和取消按钮
        self.button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel, self)
        self.button_box.setFont(self.font_manager_jetbrains_big)
        self.layout.addWidget(self.button_box)



    def update_input0_text(self):   
        """combo_box0 的信号"""
        # 获取 combo_box0 的当前文本
        current_text = self.combo_box0.currentText()
        # 判断当前文本是否为空
        if current_text and hasattr(self, "is_same_folder"):
            if self.is_same_folder: # 同一文件夹下 current_text 为 上一级文件夹名+图片名称
                # 将解析后的图片名称直接设置到 text_input0 中
                self.text_input0.setText(current_text.split("\\")[-1])
            else: # 不同文件夹下 current_text 为 上一级文件夹名
                # 将对应上一级文件夹内的图片名称设置到 text_input0 中
                if hasattr(self, "parent_folder_dict"):
                    self.text_input0.setText(self.parent_folder_dict[current_text])
                else:
                    print(f"类CameraTestDialog: 父文件夹字典self.parent_folder_dict不存在, 请检查self.parent_folder_dict")


    def refresh_data(self):
        """汇总数据"""
        # 汇总3A+ISP问题点
        items = []
        if self.checkbox1.isChecked() and self.combo_box3.currentText():
            items.append(f"AE{self.combo_box3.currentText()}")
        if self.checkbox2.isChecked() and self.combo_box4.currentText():
            items.append(f"AWB{self.combo_box4.currentText()}")
        if self.checkbox3.isChecked() and self.combo_box5.currentText():
            items.append(f"AF{self.combo_box5.currentText()}")
        if self.checkbox4.isChecked() and self.combo_box6.currentText():
            items.append(f"ISP{self.combo_box6.currentText()}")
        # 拼接内容并设置到 combo_box2
        self.text_input2.clear()  # 清空现有内容
        self.text_input2.setText("_".join(items))  # 添加拼接后的内容

    def save_data(self):
        """保存数据"""
        pass

    
    def write_data_backup(self):
        """写入数据,不支持实时写入,初始版本"""  
        excel_path = self.text_input1.text()  # 获取Excel文件路径
        search_value = self.text_input0.text()  # 获取要查找的值

        # 判断excel_path是否为空
        if not excel_path:
            print("Excel文件路径为空，请输入Excel文件路径。")
            return

        # 判断search_value是否为空
        if not search_value:
            print("要查找的值为空，请输入要查找的值。")
            return

        #尝试打开文件判断是否被占用,若被占用了则强制关闭excel
        try:
            with open(excel_path, 'a') as f:
                pass
        except PermissionError:
            print(f"Error: Excel file is locked or in use: {excel_path}")
            # 强制关闭excel
            close_excel()

        try:
            # 对excel表格进行读写操作
            workbook = openpyxl.load_workbook(excel_path) 
            sheet = workbook['Sheet1'] # workbook.active

            # 获取第一行的值
            first_row_values = [cell.value for cell in sheet[1]]  # sheet[1]表示第一行
            target_column_ae = first_row_values.index("AE") + 1
            problem_ae = self.combo_box3.currentText()
            target_column_awb = first_row_values.index("AWB") + 1
            problem_awb = self.combo_box4.currentText()
            target_column_af = first_row_values.index("AF") + 1
            problem_af = self.combo_box5.currentText()
            target_column_isp = first_row_values.index("ISP") + 1
            problem_isp = self.combo_box6.currentText()

            # 获取第一列的值
            first_column_values = [cell.value for cell in sheet['A']]  # 'A'表示第一列

            # search_value匹配，获取目标行索引
            target_row = 0
            for index, value in enumerate(first_column_values):
                if value == search_value:
                    target_row  = index + 1
            if not target_row: # 如果找不到对应的图片行，目标索引就是没有数据的最后一行
                target_row = len(sheet['A']) + 1
                # 写入对应的图片名称
                sheet.cell(target_row, 1).value = search_value
                
            # 向表格写入数据
            sheet.cell(target_row, target_column_ae).value = problem_ae
            sheet.cell(target_row, target_column_awb).value = problem_awb
            sheet.cell(target_row, target_column_af).value = problem_af
            sheet.cell(target_row, target_column_isp).value = problem_isp
        
            # 保存修改后的文件
            workbook.save(excel_path)
            print("数据已成功写入Excel文件。")

        except Exception as e:
            print(f"写入数据时发生错误: {e}")


    def write_data(self):
        """写入数据，支持实时写入"""  
        excel_path = self.text_input1.text()  # 获取Excel文件路径
        search_value = self.text_input0.text()  # 获取要查找的值

        # 判断excel_path是否为空
        if not os.path.exists(excel_path):
            show_message_box("Excel文件路径为空或不存在, 请加载或手动输入Excel文件路径", "提示", 1500)
            return

        if not search_value:
            show_message_box("没有要关联的图片项, 请点击下拉框加载", "提示", 1500)
            return

        try:
            # 启动 Excel 应用程序
            excel_app = win32.gencache.EnsureDispatch('Excel.Application')
            excel_app.Visible = True  # 设置为可见

            # 打开工作簿
            workbook = excel_app.Workbooks.Open(excel_path)
            # 选择第一个工作表
            sheet = workbook.Worksheets(1)  # 索引从 1 开始

            # 查找目标行
            target_row = 0
            for row in range(1, sheet.UsedRange.Rows.Count + 1):
                if sheet.Cells(row, 1).Value == search_value:  # 假设在第一列查找
                    target_row = row
                    break

            if target_row == 0:  # 如果找不到对应的行
                target_row = sheet.UsedRange.Rows.Count + 1      # 写入到最后一行
                sheet.Cells(target_row, 1).Value = search_value  # 写入文件名


            # 找到目标行后；按列，向表格写入数据
            for col in range(1, sheet.UsedRange.Columns.Count + 1):
                if sheet.Cells(1, col).Value  == "AE":
                    sheet.Cells(target_row, col).Value = self.combo_box3.currentText()
                if sheet.Cells(1, col).Value  == "AWB":
                    sheet.Cells(target_row, col).Value = self.combo_box4.currentText()
                if sheet.Cells(1, col).Value  == "AF":
                    sheet.Cells(target_row, col).Value = self.combo_box5.currentText()
                if sheet.Cells(1, col).Value  == "ISP":
                    sheet.Cells(target_row, col).Value = self.combo_box6.currentText()     

            # 保存工作簿
            workbook.Save()

            print("数据已成功写入Excel文件。")

        except Exception as e:
            print(f"写入数据时发生错误: {e}")
        finally:
            pass
            # 关闭工作簿（如果需要）
            # workbook.Close(SaveChanges=True)  # 如果需要保存更改
            # excel_app.Quit()  # 关闭 Excel 应用程序


    def load_data(self):
        """加载EXCEL表格"""
        try:
            options = QFileDialog.Options()
            file_path, _ = QFileDialog.getOpenFileName(self, "选择EXCEL文件", "", "Excel Files (*.xls *.xlsx);;All Files (*)", options=options)
            if file_path:
                self.text_input1.setText(file_path)  # 显示选定的文件路径
        except Exception as e:
            print(f"加载EXCEL表格时发生错误: {e}")

    def get_data(self):
        """获取用户输入的数据"""
        return {
            "关联图片项": self.combo_box0.currentText() + "/" + self.text_input0.text(),
            "问题点": self.text_input2.text(),
            "AE": self.combo_box3.currentText(),
            "AWB": self.combo_box4.currentText(),
            "AF": self.combo_box5.currentText(),
            "ISP": self.combo_box6.currentText(),
        }

    # 新增方法：保存设置
    def save_settings(self):
        """保存当前设置"""
        settings = {
            "关联图片项": self.combo_box0.currentText(),
            # "输入框0": self.text_input0.text(),
            "问题点路径": self.text_input1.text(),
            # "问题点汇总": self.text_input2.text(),
            "AE": [self.combo_box3.itemText(i) for i in range(self.combo_box3.count())],
            "AWB": [self.combo_box4.itemText(i) for i in range(self.combo_box4.count())],
            "AF": [self.combo_box5.itemText(i) for i in range(self.combo_box5.count())],
            "ISP": [self.combo_box6.itemText(i) for i in range(self.combo_box6.count())],
        }
        save_path = os.path.join(BasePath, "cache", "test_settings.json")
        with open(save_path, 'w', encoding='utf-8') as f:
            json.dump(settings, f, ensure_ascii=False, indent=4)

        print("设置已保存。")


    # 新增方法：加载设置
    def load_settings(self):
        """加载上次保存的设置"""
        try:
            with open('./cache/test_settings.json', 'r', encoding='utf-8') as f:
                settings = json.load(f)
                combo_box0_text = settings.get("关联图片项", "")
                combo_box0_items = [self.combo_box0.itemText(i) for i in range(self.combo_box0.count())]
                if combo_box0_text and combo_box0_text in combo_box0_items:
                    # 如果 combo_box0 的当前文本不为空，则更新 text_input0 的文本
                    self.combo_box0.setCurrentText(combo_box0_text)
                    self.update_input0_text()
                
                self.text_input1.setText(settings.get("问题点路径", ""))
                # self.text_input2.setText(settings.get("问题点汇总", ""))

                # 恢复AE、AWB、AF、ISP的选项
                for item in settings.get("AE", []):
                    self.combo_box3.addItem(item)
                for item in settings.get("AWB", []):
                    self.combo_box4.addItem(item)
                for item in settings.get("AF", []):
                    self.combo_box5.addItem(item)
                for item in settings.get("ISP", []):
                    self.combo_box6.addItem(item)
        except FileNotFoundError:
            print("未找到设置文件，使用默认值。")
        except json.JSONDecodeError:
            print("设置文件格式错误，使用默认值。")

    def keyPressEvent(self, event):
        """重写键盘按下事件，防止在输入框或下拉框中按下回车时关闭对话框"""
        if event.key() == Qt.Key_Return or event.key() == Qt.Key_Enter:
            # 如果当前焦点在输入框或下拉框中，阻止默认行为
            if self.focusWidget() in [self.text_input0, self.text_input2, self.combo_box0, self.combo_box3, self.combo_box4, self.combo_box5, self.combo_box6]:
                event.ignore()  # 忽略事件
            # 如果当前焦点在问题点输入框中，打开Excel文件
            elif self.focusWidget() == self.text_input1:
                excel_path = self.text_input1.text()
                if os.path.exists(excel_path):
                    os.startfile(excel_path)  # 在Windows上打开文件
                else:
                    print("指定的Excel文件不存在。")
            else:
                super().keyPressEvent(event)  # 处理其他情况
        else:
            super().keyPressEvent(event)  # 处理其他按键事件


class ImageTransform:
    """图片旋转exif信息调整类"""
    # 定义EXIF方向值对应的QTransform变换
    _ORIENTATION_TRANSFORMS = {
        1: QTransform(),  # 0度 - 正常
        2: QTransform().scale(-1, 1),  # 水平翻转
        3: QTransform().rotate(180),  # 180度
        4: QTransform().scale(1, -1),  # 垂直翻转
        5: QTransform().rotate(90).scale(-1, 1),  # 顺时针90度+水平翻转
        6: QTransform().rotate(90),  # 顺时针90度
        7: QTransform().rotate(-90).scale(-1, 1),  # 逆时针90度+水平翻转
        8: QTransform().rotate(-90)  # 逆时针90度
    }
    @classmethod
    def pic_size(cls, path, pixmap, index):
        """获取图片名称、尺寸、大小等基础信息"""
        pic_name = os.path.basename(path)
        # pixmap是旋转后的图像，尺寸会更准确
        width = pixmap.width()
        height = pixmap.height()
        file_size = os.path.getsize(path)  # 文件大小（字节）
        if file_size < 1024:
            size_str = f"{file_size} B"
        elif file_size < 1024 ** 2:
            size_str = f"{file_size / 1024:.2f} KB"
        else:
            size_str = f"{file_size / (1024 ** 2):.2f} MB"
        exif_size_info = f"图片名称: {pic_name}\n图片大小: {size_str}\n图片尺寸: {width} x {height}\n图片张数: {index}"
        return exif_size_info

    @classmethod
    def get_orientation(cls, image):
        """获取图片的EXIF方向信息（优化版）"""
        try:
            # 检查是否是支持EXIF的格式
            if image.format not in ('JPEG', 'TIFF', 'MPO'):
                return 1
            
            # 获取EXIF数据（使用更可靠的获取方式）
            exif_data = image.info.get('exif')
            if not exif_data:
                return 1
            
            # 使用piexif的字节加载方式
            exif_dict = piexif.load(exif_data)
            return exif_dict['0th'].get(piexif.ImageIFD.Orientation, 1)
                
        except (KeyError, AttributeError, ValueError):
            # 当EXIF数据不包含方向信息时
            return 1
        except Exception as e:
            print(f"get_orientation函数读取EXIF方向信息失败: {str(e)}")
            return 1

    @classmethod
    def auto_rotate_image(cls, icon_path: str, index: str , img):
        """函数功能： icon_path 文件路径, index 文件索引信息
        1. 获取图片文件的旋转方向信息并旋转图片到正常方向, 返回 pixmap
        2. 使用PIL(函数pic_size)解析图片文件的基础信息, 返回 basic_info
        3. 返回图片文件的格式，返回 image.format
        """
        try:

            if True: # 低效率方案，为复现TIFF格式QPixmap直接加载失败的问题, 移除该逻辑   
                # 优先使用PIL打开图片（兼容更多格式）
                # image = Image.open(icon_path)
                if isinstance(img, Image.Image):  # 处理PIL图像对象
                    # 转换PIL图像到OpenCV格式
                    pass
                else:
                    print(f"auto_rotate_image无法加载图像")
                    return None, None

                # 获取图片文件的格式信息
                image_format = img.format if img.format else 'None'

                if image_format not in ["TIFF"]: # "JPEG",, "MPO"
                    # 使用QPixmap(icon_path)直接创建
                    pixmap = QPixmap(icon_path)

                else:
                    # 转换为QPixmap（解决TIFF格式QPixmap直接加载失败的问题）
                    if img.mode == 'RGBA':
                        # 处理带透明通道的图片
                        image_ = img.convert("RGBA")
                        data = image_.tobytes("raw", "RGBA")
                        qimage = QImage(data, image_.size[0], image_.size[1], QImage.Format_RGBA8888)
                    else:
                        # 转换为RGB模式保证兼容性
                        image_ = img.convert("RGB")
                        data = image_.tobytes("raw", "RGB")
                        qimage = QImage(data, image_.size[0], image_.size[1], QImage.Format_RGB888)
                    pixmap = QPixmap.fromImage(qimage)

            if False:
                # 使用QPixmap(icon_path)直接创建，TIFF格式QPixmap直接加载会失败, 移除该逻辑
                pixmap = QPixmap(icon_path)                    

            # 获取EXIF方向信息
            orientation = cls.get_orientation(img)

            # 应用方向变换
            transform = cls._ORIENTATION_TRANSFORMS.get(orientation, QTransform())
            if not transform.isIdentity():  # 只在需要变换时执行
                pixmap = pixmap.transformed(transform, Qt.SmoothTransformation)
            
            # 解析img的hdr相关信息，
            # {"mirror":false,"sensorType":"rear","Hdr":"auto","OpMode":36869,"smallPicture":false,"AIScene":0,"filterId":66048,"zoomMultiple":1}'
            ultra_info = None
            ultra_info_ = None
            exif_dict = img._getexif()
            if exif_dict:
                ultra_info_ = exif_dict.get(39321,None)
            if ultra_info_ and isinstance(ultra_info_,str):
                # 将字符串解析为字典
                data = json.loads(ultra_info_)
                hdr_value = data.get("Hdr", None)  # 使用get方法获取值
                # mirror_value = data.get("mirror", None)  # 使用get方法获取值
                zoom_value = data.get("zoomMultiple", None)  # 使用get方法获取值
                # 拼接HDR等信息            
                ultra_info = f"\nHDR: {hdr_value}\nZoom: {zoom_value}"
            else:
                ultra_info = f"\nHDR: null\nZoom: null"

            # 获取基本exif信息,少使用一次Image.open
            basic_info = cls.pic_size(icon_path, pixmap, index)
            # 如果HDRZOOM信息存在，就更新到basic_info中
            if ultra_info:
                basic_info = basic_info + ultra_info
            
            return pixmap, basic_info
            
        except Exception as e:
            print(f"处理图片方向变换失败 {icon_path}: {str(e)}")
            return pixmap

""""继承 QGraphicsRectItem 并重写 itemChange 方法来实现对矩形框变化的监听"""
class CustomGraphicsRectItem(QGraphicsRectItem):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.change_callback = None

    def set_change_callback(self, callback):
        """设置变化回调函数"""
        self.change_callback = callback

    def itemChange(self, change, value):
        """重写itemChange方法，监听矩形框变化"""
        if self.change_callback and change in [QGraphicsItem.ItemPositionHasChanged, 
                                             QGraphicsItem.ItemTransformHasChanged]:
            # 使用QTimer.singleShot延迟回调，确保矩形框位置更新完成
            QtCore.QTimer.singleShot(0, self.change_callback)
        return super().itemChange(change, value)
    


""""后台线程计算矩形框亮度等统计信息, 使用线程池管理"""
class StatsTask(QRunnable):
    def __init__(self, roi, callback):
        super().__init__()
        self.roi = roi
        self.callback = callback

    def run(self):
        """线程执行函数，计算统计信息"""
        try:
            stats = calculate_image_stats(self.roi, resize_factor=1)
            self.callback(stats)
        except Exception as e:
            print(f"计算统计信息时出错: {e}")
            self.callback({})


"""图片视图类"""
class MyGraphicsView(QGraphicsView):
    def __init__(self, scene, exif_text=None, stats_text=None, *args, **kwargs):
        super(MyGraphicsView, self).__init__(scene, *args, **kwargs)
        self.setRenderHint(QPainter.Antialiasing)
        self.setRenderHint(QPainter.SmoothPixmapTransform)
        self.setDragMode(QGraphicsView.ScrollHandDrag)
        self.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.setTransformationAnchor(QGraphicsView.AnchorUnderMouse)
        self.setResizeAnchor(QGraphicsView.AnchorUnderMouse)
        self.setViewportUpdateMode(QGraphicsView.FullViewportUpdate)  # 重要！确保及时更新
        
        # 设置顶层窗口，方便调用类SubMainWindow中的函数与变量
        self.parent_SubMainWindow = self.window()


        # 初始化字体管理器
        self.font_manager_view = SingleFontManager.get_font(10) 

        # 初始化基本信息
        self.exif_text = exif_text  # 存储 EXIF 信息
        self.stats_text = stats_text  # 存储 stats 信息
        self.show_exif = True if exif_text else False  # 控制 EXIF 显示
        self.stats_visible = True if stats_text else False  # 控制 stats 显示
        self.histogram = None  # 存储直方图数据
        self.show_histogram = False  # 控制直方图显示

        # 添加 QLabel 显示 EXIF 信息
        self.exif_label = QLabel(self)
        self.exif_label.setText(self.exif_text if self.exif_text else "解析不出exif信息!")
        self.exif_label.setStyleSheet("color: white; background-color: transparent;")
        self.exif_label.setFont(self.font_manager_view)
        self.exif_label.setVisible(self.show_exif)
        self.exif_label.setAttribute(Qt.WA_TransparentForMouseEvents)

        # 添加新的QLabel用于显示直方图信息
        self.histogram_label = QLabel(self)
        self.histogram_label.setStyleSheet("border: none;")
        self.histogram_label.setFixedSize(150, 100)
        self.histogram_label.setVisible(self.show_histogram)
        self.histogram_label.setAttribute(Qt.WA_TransparentForMouseEvents)

        # 添加新的QLabel用于显示亮度、RGB、LAB的统计信息
        self.stats_label = QLabel(self)
        self.stats_label.setText(self.stats_text if self.stats_text else "不存在亮度统计信息!")
        self.stats_label.setStyleSheet("color: white; background-color: rgba(0, 0, 0, 0.7); padding: 2px;")
        self.stats_label.setFont(self.font_manager_view)
        self.stats_label.setVisible(self.stats_visible)
        self.stats_label.setAttribute(Qt.WA_TransparentForMouseEvents)

        # 添加ROI矩形框相关属性
        self.selection_rect = None
        self.original_image = None  # 存储原始OpenCV图像数据
        self.selection_visible = False
        self.last_pos = None  # 记录鼠标右键拖动的起始位置
        self.move_step = 1.0  # 动态设置矩形框跟随鼠标移动步长

        # 初始化线程池，用于计算ROI矩形框的统计信息
        self.thread_pool = QThreadPool.globalInstance()

        # 初始更新标签位置
        self.update_labels_position()
        

    def set_original_image(self, cv_img):
        """设置原始OpenCV图像用于统计计算"""
        self.original_image = cv_img

    def toggle_selection_rect(self, visible):
        """切换选择框的显示状态"""
        self.selection_visible = visible
        if visible: 
            if not self.selection_rect:
                
                self.selection_rect = CustomGraphicsRectItem()
                self.selection_rect.setPen(QPen(QColor(0, 255, 0, 255), 15))
                self.selection_rect.setFlag(QGraphicsItem.ItemIsMovable)
                self.selection_rect.setFlag(QGraphicsItem.ItemIsSelectable)
                self.selection_rect.setFlag(QGraphicsItem.ItemSendsGeometryChanges)

                """ 设置变化回调函数_统计矩形框内的亮度等信息"""
                self.selection_rect.set_change_callback(self.update_roi_stats)
                
                # 将矩形项添加到场景中
                self.scene().addItem(self.selection_rect)
                # 将矩形项添加到视图的viewport中
                
                
                # 设置初始大小和位置
                scene_rect = self.scene().sceneRect()
                size = min(scene_rect.width(), scene_rect.height()) / 10
                # size = 100
                self.selection_rect.setRect(
                    scene_rect.center().x() - size/2,
                    scene_rect.center().y() - size/2,
                    size, size
                )
                
            # 显示ROI矩形框
            self.selection_rect.show()

            # 初始更新统计信息
            self.update_roi_stats()
        else:
            if self.selection_rect:
                self.selection_rect.hide()
                # 清除ROI统计信息
                if hasattr(self, 'stats_label'):
                    self.set_stats_data(self.stats_text if self.stats_text else "")


    def update_roi_stats(self):
        """更新ROI区域的统计信息"""
        if not self.selection_rect or self.original_image is None:
            print("update_roi_stats error!")
            return
        # 使用 QTimer 延迟调用，避免频繁计算
        QTimer.singleShot(100, self._calculate_roi_stats)


    def _calculate_roi_stats(self):
        """提取 ROI 区域并启动线程计算统计信息"""
        try:
            if not self.selection_rect or self.original_image is None:
                return

            # 获取选择框在场景中的位置和大小
            scene_rect = self.selection_rect.sceneBoundingRect()
            
            # 获取原始图像尺寸
            img_h, img_w = self.original_image.shape[:2]
            
            # 转换场景坐标到图像坐标
            x1 = max(0, min(img_w-1, int(scene_rect.left())))
            y1 = max(0, min(img_h-1, int(scene_rect.top())))
            x2 = max(0, min(img_w, int(scene_rect.right())))
            y2 = max(0, min(img_h, int(scene_rect.bottom())))
            
            # 确保有效的 ROI 区域
            if x2 > x1 and y2 > y1:
                # 提取 ROI 区域
                roi = self.original_image[y1:y2, x1:x2]
                
                # 创建并启动新的任务
                task = StatsTask(roi, self._update_stats_display)

                self.thread_pool.start(task)
        except Exception as e:
            print(f"提取 ROI 区域时出错: {e}")

    def _update_stats_display(self, stats):
        """更新统计信息显示"""
        if not stats:
            return

        # 格式化 ROI 统计信息
        roi_stats = (
            f"--- ROI统计信息 ---\n"
            # f"ROI: {stats['width']}x{stats['height']}\n"
            f"亮度: {stats['avg_brightness']:.1f}\n"
            f"对比度: {stats['contrast']:.1f}\n"
            # f"LAB均值: {stats['avg_lab']}\n"
            f"RGB均值: {stats['avg_rgb']}\n"
            f"(R/G:{stats['R_G']} B/G: {stats['B_G']})\n"
            f"(roi:{stats['width']}x{stats['height']})"
            # f"区域大小: {stats['width']}x{stats['height']}"
        )
        self.set_stats_data(roi_stats)


    def adjust_roi_size(self, delta, anchor_point=None):
        """调整ROI选择框大小（基于变换矩阵）"""
        if not self.selection_rect:
            return

        try:
            # 设置调整步长
            scale_factor = 1.15 if delta > 0 else 0.955
            
            # 获取当前矩形的位置和尺寸
            current_rect = self.selection_rect.rect()
            scene_rect = self.scene().sceneRect()
            
            # 计算新的尺寸
            new_width = max(50, min(current_rect.width() * scale_factor, scene_rect.width()))
            new_height = max(50, min(current_rect.height() * scale_factor, scene_rect.height()))
            
            # 计算缩放比例
            width_scale = new_width / current_rect.width()
            height_scale = new_height / current_rect.height()
            
            # 创建变换矩阵
            transform = QTransform()
            
            # 如果有锚点，则以锚点为中心进行缩放
            if anchor_point:
                # 将锚点转换为局部坐标系
                local_anchor = self.selection_rect.mapFromScene(anchor_point)
                # 以锚点为中心进行缩放
                transform.translate(local_anchor.x(), local_anchor.y())
                transform.scale(width_scale, height_scale)
                transform.translate(-local_anchor.x(), -local_anchor.y())
            else:
                # 以中心点进行缩放
                center = current_rect.center()
                transform.translate(center.x(), center.y())
                transform.scale(width_scale, height_scale)
                transform.translate(-center.x(), -center.y())
            
            # 应用变换
            self.selection_rect.setTransform(transform, True)
            
            # 确保ROI不会超出场景边界
            new_rect = self.selection_rect.sceneBoundingRect()
            if new_rect.left() < scene_rect.left():
                self.selection_rect.setX(scene_rect.left())
            if new_rect.top() < scene_rect.top():
                self.selection_rect.setY(scene_rect.top())
            if new_rect.right() > scene_rect.right():
                self.selection_rect.setX(scene_rect.right() - new_rect.width())
            if new_rect.bottom() > scene_rect.bottom():
                self.selection_rect.setY(scene_rect.bottom() - new_rect.height())
            
            # 同步其他视图的ROI大小和位置
            main_window = self.window()
            if isinstance(main_window, SubMainWindow):
                for view in main_window.graphics_views:
                    if view and view != self and view.selection_rect:
                        view.selection_rect.setTransform(transform, True)
                        view.update_roi_stats()
            
            # 更新当前视图的ROI统计信息
            self.update_roi_stats()
            
        except Exception as e:
            print(f"adjust_roi_size error: {str(e)}")


    def update_labels_position(self):
        """更新标签位置"""
        padding = 5  # 边距
        
        # 更新统计信息标签位置（左下角）
        if self.stats_label:
            self.stats_label.move(padding, self.height() - self.stats_label.height() - padding)
            self.stats_label.adjustSize()  # 根据内容调整大小
        

        if self.show_histogram and self.show_exif:
            # 两个标签都显示时，直方图在上，EXIF在下
            self.histogram_label.move(padding, padding)
            self.exif_label.move(padding, padding + self.histogram_label.height() + padding)
        else:
            # 只显示其中一个时，都显示在左上角
            if self.show_histogram:
                self.histogram_label.move(padding, padding)
            if self.show_exif:
                self.exif_label.move(padding, padding)


    def set_histogram_data(self, histogram):
        """设置直方图数据"""
        if histogram is None:
            self.histogram_label.setText("无直方图数据")
            return
        # 使用 matplotlib 生成直方图图像
        try:
            plt.figure(figsize=(3, 2), dpi=100, facecolor='none', edgecolor='none')  # 设置背景透明
            ax = plt.gca()
            # 计算相对频率
            total_pixels = sum(histogram)
            relative_frequency = [count / total_pixels for count in histogram]
            ax.plot(range(len(relative_frequency)), relative_frequency, color='skyblue', linewidth=1)
            ax.fill_between(range(len(relative_frequency)), relative_frequency, color='skyblue', alpha=0.7)            
            ax.set_xlim(0, 255)
            ax.set_ylim(0, max(relative_frequency)*1.1)
            ax.yaxis.set_visible(False)  # 隐藏 Y 轴
            ax.xaxis.set_tick_params(labelsize=8)
            plt.tight_layout()

            buf = io.BytesIO()
            plt.savefig(buf, format='PNG', transparent=True, bbox_inches='tight', pad_inches=0)
            buf.seek(0)
            plt.close()

            histogram_pixmap = QPixmap()
            histogram_pixmap.loadFromData(buf.getvalue(), 'PNG')
            buf.close()

            # 缩放直方图图像以适应 QLabel
            self.histogram_label.setPixmap(histogram_pixmap.scaled(
                self.histogram_label.size(), Qt.KeepAspectRatio, Qt.SmoothTransformation))
        except Exception as e:
            print(f"生成直方图图像失败: {e}")
            self.histogram_label.setText("无法生成直方图")


    def set_histogram_visibility(self, visible: bool):
        """设置直方图可见性"""
        self.show_histogram = visible
        self.histogram_label.setVisible(visible)

        self.update_labels_position()


    def set_exif_visibility(self, visible: bool, font_color: str):
        """设置EXIF信息可见性"""
        self.show_exif = visible
        self.exif_label.setVisible(visible)
        self.exif_label.setStyleSheet(f"color: {font_color}; background-color: transparent; font-weight: 400;")

        self.update_labels_position()


    def set_stats_visibility(self, visible: bool):
        """设置亮度统计信息的可见性"""
        self.stats_visible = visible
        self.stats_label.adjustSize()  # 根据内容调整大小
        self.stats_label.setVisible(visible)

        self.update_labels_position()


    def set_stats_data(self, text: str = ""):
        """设置亮度统计信息的数据"""
        self.stats_label.setText(text)
        self.stats_label.adjustSize()  # 根据内容调整大小
    
        self.update_labels_position()

    
    # 窗口尺寸改变事件函数
    def resizeEvent(self, event):
        """重写调整大小事件"""
        super(MyGraphicsView, self).resizeEvent(event)
        self.update_labels_position()
    

    # 鼠标滚轮事件函数
    def wheelEvent(self, event):
        """重写鼠标滚轮事件"""
        try:
            # 控制ROI信息显示的变量 self.parent_SubMainWindow.stats_visible 
            # 控制矩形框显示的变量 self.parent_SubMainWindow.roi_selection_active
            # if event.modifiers() & Qt.AltModifier:
            if self.parent_SubMainWindow.roi_selection_active:
                # Alt + 滚轮调整ROI大小 
                if self.selection_rect:
                    # 获取鼠标在场景中的位置 视图位置 event.pos()
                    mouse_scene_pos = self.mapToScene(event.pos())
                            
                    # 调整ROI大小，并保持鼠标位置相对ROI的位置不变
                    self.adjust_roi_size(event.angleDelta().y(), mouse_scene_pos)

                    # 添加调试信息
                    # print(f"MyGraphicsView - 当前滚轮数值: {event.angleDelta()}")

                    # 事件已处理，阻止进一步传递
                    event.accept()
                    return  
                
            # 如果没有Alt键，调用父类的wheelEvent（确保父类是SubMainWindow）
            elif isinstance(self.parent_SubMainWindow, SubMainWindow):
                self.parent_SubMainWindow.wheelEvent(event)
                
            else:
                # 如果父类不是SubMainWindow，直接调用QGraphicsView的默认实现
                super().wheelEvent(event)

        except Exception as e:
            print(f"MyGraphicsView - wheelEvent error: {str(e)}")


    # 鼠标按压事件函数
    def mousePressEvent(self, event):
        try:

            # 动态计算self.move_step的值,确保移动ROI矩形框的时候不卡顿
            base_scales = self.parent_SubMainWindow.base_scales
            if base_scales:
               self.move_step = 1.0 / max(base_scales)
               # 确保结果在 [1, 4] 之间
               self.move_step = max(1, min(4, self.move_step))

            if event.button() == Qt.LeftButton:
                # 左键按下，记录起始位置
                self.last_pos = event.pos()
                # event.accept()

                # 如果在矩形框控制模式下，直接将矩形框移动到鼠标位置
                if self.parent_SubMainWindow.roi_selection_active and self.selection_rect:
                    # 获取鼠标在场景中的位置
                    mouse_scene_pos = self.mapToScene(event.pos())
                    
                    # 计算矩形框的新位置
                    rect = self.selection_rect.rect()
                    new_pos = mouse_scene_pos - rect.center()
                    
                    # 移动当前视图的矩形框
                    self.selection_rect.setPos(new_pos)
                    
                    # 同步其他视图的矩形框位置
                    main_window = self.window()
                    if isinstance(main_window, SubMainWindow):
                        for view in main_window.graphics_views:
                            if view and view != self and view.selection_rect:
                                view.selection_rect.setPos(new_pos)
                                view.update_roi_stats()
                    
                    event.accept()
                    return


            elif event.button() == Qt.RightButton:
                # 右键按下，记录起始位置
                self.last_pos = event.pos()
                
                print(f"当前鼠标所在视图：视图尺寸：{self.width()}x{self.height()}--场景尺寸：{self.scene().sceneRect().width()}x{self.scene().sceneRect().height()}")

                event.accept()
            else:
                super().mousePressEvent(event)
        except Exception as e:
            print(f"鼠标事件处理错误: {e}")
            event.ignore()


    # 鼠标移动事件函数
    def mouseMoveEvent(self, event):
        try:
            
            if self.parent_SubMainWindow.roi_selection_active: # 矩形框控制模式
                # print(self.move_step)
                if event.buttons() & Qt.LeftButton and self.last_pos is not None:
                    # Alt+左键移动，同步所有视图的ROI矩形框
                    delta = (event.pos() - self.last_pos)*self.move_step
                    self.last_pos = event.pos()
                    main_window = self.window()
                    if isinstance(main_window, SubMainWindow):
                        for view in main_window.graphics_views:
                            if view and view.selection_rect:
                                view.selection_rect.moveBy(delta.x(), delta.y())
                    event.accept()
                elif event.buttons() & Qt.RightButton and self.last_pos is not None:
                    # Alt+右键移动，只移动当前视图的ROI矩形框
                    delta = (event.pos() - self.last_pos)*self.move_step
                    self.last_pos = event.pos()
                    if self.selection_rect:
                        self.selection_rect.moveBy(delta.x(), delta.y())
                    event.accept()
                else:
                    super().mouseMoveEvent(event)
            else:  # 正常模式
                if event.buttons() & Qt.LeftButton and self.last_pos is not None:
                    delta = event.pos() - self.last_pos
                    self.last_pos = event.pos()
                    main_window = self.window()
                    if isinstance(main_window, SubMainWindow):
                        for view in main_window.graphics_views:
                            view.horizontalScrollBar().setValue(
                                view.horizontalScrollBar().value() - delta.x())
                            view.verticalScrollBar().setValue(
                                view.verticalScrollBar().value() - delta.y())
                    event.accept()
                elif event.buttons() & Qt.RightButton and self.last_pos is not None:
                    delta = event.pos() - self.last_pos
                    self.last_pos = event.pos()
                    self.horizontalScrollBar().setValue(
                        self.horizontalScrollBar().value() - delta.x())
                    self.verticalScrollBar().setValue(
                        self.verticalScrollBar().value() - delta.y())
                    event.accept()
                else:
                    super().mouseMoveEvent(event)
        except Exception as e:
            print(f"在 mouseMoveEvent 中发生错误: {e}")


    # 鼠标释放事件函数
    def mouseReleaseEvent(self, event):
        if event.button() in (Qt.LeftButton, Qt.RightButton):
            # 左键或右键释放，重置位置
            self.last_pos = None
            event.accept()
        else:
            super().mouseReleaseEvent(event)


"""
设置独立封装类区域结束线
------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
"""



"""
设置看图界面类区域开始线
------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
"""


class SubMainWindow(QMainWindow, Ui_MainWindow):
    """看图主界面类"""
    # 添加一个信号用于通知主窗口子窗口已关闭
    closed = pyqtSignal()
    # 在类级别定义信号
    ai_response_signal = pyqtSignal(str)
    # 新增进度条更新信号
    progress_updated = pyqtSignal(int)

    def __init__(self, images_path_list, index_list=None, parent=None):
        super(SubMainWindow, self).__init__(parent)
        # 初始化UI
        self.setupUi(self) 
        
        # 初始化传入的images_path_list(图像路径列表)、index_list(当前索引)以及parent(主窗口的self)
        self.parent_window = parent
        self.images_path_list = images_path_list
        self.index_list = index_list

        # 初始化变量
        self.init_variables()

        # 加载之前的配置
        self.load_settings()

        # 设置窗口标题组件和样式表
        self.set_stylesheet()

        # 初始化图片视图，集成了看图子界面的主要功能
        self.set_images(self.images_path_list, self.index_list)

        # 设置快捷键和槽函数
        self.set_shortcut()

        # 显示窗口
        self.showMaximized()
        # self.showMinimized()
        # self.show()

        # 更新颜色样式表，放到最后，确保生效
        self.update_ui_styles()


    def init_variables(self):
        """初始化变量"""
        # 初始化SubMainWindow类中的一些列表属性
        self.exif_texts = []
        self.histograms = []
        self.original_rotation = []
        self.graphics_views = []
        self.original_pixmaps = []
        self.gray_pixmaps = []
        self.cv_imgs = []
        self.pil_imgs = []
        self.base_scales = []
        self._scales_min = []

        # 设置表格的宽高初始大小
        self.table_width_heigth_default = [2534,1376]

        # 初始化一些显示相关的标志位
        self.roi_selection_active = False  # 初始化roi亮度等信息统计框的显示标志位
        self.is_fullscreen = False         # 初始化全屏标志位
        self.is_updating = False           # 设置更新状态标志位
        
        # 导入主界面的一些设置:字体设置，颜色设置等
        if self.parent_window:
            # 导入字体设置
            self.custom_font = self.parent_window.custom_font if self.parent_window.custom_font else SingleFontManager.get_font(12) 
            self.font_manager_jetbrains = (self.parent_window.custom_font_jetbrains_medium if self.parent_window.custom_font_jetbrains_medium 
                                           else SingleFontManager.get_font(size=11, font_path=os.path.join(BasePath, "resource", "fonts", "JetBrainsMapleMono_Regular.ttf")))

            # 导入颜色设置, 背景色，表格背景色，字体颜色，exif字体颜色; 从load_settings()中读取
            # self.background_color_default = self.parent_window.background_color_default
            # self.background_color_table = self.parent_window.background_color_table
            # self.font_color_default = self.parent_window.font_color_default
            # self.font_color_exif = self.parent_window.font_color_exif

        # 设置rgb颜色值
        # self.color_rgb_settings = {}         
        # 初始化exif信息可见性字典,支持用户在json配置文件中调整顺序以及是否显示该项
        # self.dict_exif_info_visibility = {} 
        # 初始化图像显示色彩空间变量,默认设置srgb显示空间,在load_settings()中初始化
        # self.dict_label_info_visibility = {}
        # self.srgb_color_space = True  
        # self.p3_color_space = False   
        # self.gray_color_space = False

    def set_shortcut(self):
        """设置快捷键和槽函数"""

        """1. 设置快捷键"""
        # 创建快捷键，按住Esc键退出整个界面
        self.shortcut_esc = QShortcut(QKeySequence(Qt.Key_Escape), self)
        self.shortcut_esc.activated.connect(self.Escape_close)

        # 创建快捷键，F11 全屏
        fullscreen_shortcut = QShortcut(QKeySequence('F11'), self)
        fullscreen_shortcut.activated.connect(self.toggle_fullscreen)
        self.is_fullscreen = False
        
        # 添加Ctrl+A和Ctrl+D快捷键
        self.shortcut_rotate_left = QShortcut(QKeySequence("Ctrl+A"), self)
        self.shortcut_rotate_left.activated.connect(self.rotate_left)
        self.shortcut_rotate_right = QShortcut(QKeySequence("Ctrl+D"), self)
        self.shortcut_rotate_right.activated.connect(self.rotate_right)

        # 添加空格键和B键的快捷键
        self.shortcut_space = QShortcut(QKeySequence(Qt.Key_Space), self)
        self.shortcut_space.activated.connect(self.on_space_pressed)
        self.shortcut_b = QShortcut(QKeySequence('b'), self)
        self.shortcut_b.activated.connect(self.on_b_pressed)

        # 添加P键的快捷键
        self.shortcut_p = QShortcut(QKeySequence('p'), self)
        self.shortcut_p.activated.connect(self.on_p_pressed)


        # 添加V键的快捷键
        self.shortcut_v = QShortcut(QKeySequence('v'), self)
        self.shortcut_v.activated.connect(self.on_v_pressed)
        # 添加N键的快捷键
        self.shortcut_v = QShortcut(QKeySequence('n'), self)
        self.shortcut_v.activated.connect(self.on_n_pressed)

        # 添加T键的快捷键,实现截图功能
        self.shortcut_t = QShortcut(QKeySequence('t'), self)
        self.shortcut_t.activated.connect(self.on_t_pressed)
        # 添加Ctrl+T键的快捷键,实现评测问题点描述
        self.shortcut_rotate_left = QShortcut(QKeySequence("Ctrl+T"), self)
        self.shortcut_rotate_left.activated.connect(self.on_ctrl_t_pressed)

        """2. 连接复选框信号到槽函数"""
        # 连接复选框信号到槽函数
        self.checkBox_1.stateChanged.connect(self.toggle_histogram_info)  # 新增直方图显示
        self.checkBox_2.stateChanged.connect(self.toggle_exif_info)       # 新增EXIF信息显示
        self.checkBox_3.stateChanged.connect(self.roi_stats_checkbox)     # 新增ROI信息
        self.checkBox_4.stateChanged.connect(self.ai_tips_info)           # 新增AI提示看图
        
        # 连接下拉列表信号到槽函数
        self.comboBox_1.activated.connect(self.show_menu_combox1) # 连接 QComboBox 的点击事件到显示菜单，self.on_comboBox_1_changed
        # self.comboBox_2.currentIndexChanged.connect(self.on_comboBox_2_changed)  # 当用户选择不同选项的时候触发
        self.comboBox_2.activated.connect(self.on_comboBox_2_changed)            # 当用户选择任何选项的时候都会触发 

        # 连接底部状态栏按钮信号到槽函数
        self.statusbar_left_button.clicked.connect(self.open_settings_window)
        self.statusbar_button1.clicked.connect(self.on_b_pressed)
        self.statusbar_button2.clicked.connect(self.on_space_pressed)


        # 连接AI响应信号到槽函数
        self.ai_response_signal.connect(self.update_ai_response)
        # 连接进度条更新信号到槽函数
        self.progress_updated.connect(self.update_progress)

        
    def set_stylesheet(self):
        """设置窗口标题组件和样式表"""
        """窗口组件概览
        第一排, self.label_0, self.comboBox_1, self.comboBox_2, self.checkBox_1, self.checkBox_2, self.checkBox_3
        第二排, self.tableWidget_medium
        第三排, self.label_bottom
        """
        # 设置主界面图标以及标题
        icon_path = os.path.join(BasePath, "resource", "icons", "viewer.ico")
        self.setWindowIcon(QIcon(icon_path))
        self.setWindowTitle("图片对比界面")

        # 获取鼠标所在屏幕，并根据当前屏幕计算界面大小与居中位置，调整大小并移动到该位置
        screen = QtWidgets.QApplication.desktop().screenNumber(QtWidgets.QApplication.desktop().cursor().pos())
        screen_geometry = QtWidgets.QApplication.desktop().screenGeometry(screen)
        width = int(screen_geometry.width() * 0.8)
        height = int(screen_geometry.height() * 0.65)
        self.resize(width, height)
        x = screen_geometry.x() + (screen_geometry.width() - self.width()) // 2
        y = screen_geometry.y() + (screen_geometry.height() - self.height()) // 2
        self.move(x, y)

        # 设置第一排标签
        self.label_0.setText("提示:鼠标左键拖动所有图像,滚轮控制放大/缩小;按住Ctrl+滚轮或者鼠标右键操作单独图像")
        # self.label_0.setFont(self.custom_font)，移除之前的设置，使用新的字体管理器
        self.label_0.setFont(self.font_manager_jetbrains)

        # 设置下拉框选项,会自动进入槽函数self.show_menu_combox1-->on_comboBox_1_changed
        self.comboBox_1.clear()  # 清除已有项
        self.comboBox_1.addItems(["✅颜色设置", "⭕一键重置", "🔽背景颜色>>", "🔽表格填充颜色>>", "🔽字体颜色>>", "🔽exif字体颜色>>"])  # 添加主选项
        self.comboBox_1.setEditable(False)  # 设置 QComboBox 不可编辑
        self.comboBox_1.setCurrentIndex(0)  # 设置默认显示索引为0
        self.comboBox_1.setFont(self.custom_font)

        """设置下拉框self.comboBox_2"""
        if False: # 旧版本方案
            # 设置下拉框选项,会自动进入槽函数on_comboBox_2_changed
            color_space_list = [self.srgb_color_space, self.gray_color_space, self.p3_color_space]  # 列表中存放三个颜色空间显示标志位
            _front = {
                True:"✅",
                False:"" }
            _text = [f"{_front.get(color_space_list[0])}sRGB色域", 
                f"{_front.get(color_space_list[1])}灰度图空间色域", 
                f"{_front.get(color_space_list[2])}p3色域"]
            self.comboBox_2.clear()  # 清除已有项
            self.comboBox_2.addItems([_text[0], _text[1], _text[2]])
            # 设置默认显示索引为三个颜色空间中为TRUE的那个
            self.comboBox_2.setCurrentIndex(color_space_list.index(True))  
            self.comboBox_2.setFont(self.custom_font)
        
        # 设置下拉框选项（优化版）
        color_space_list = [self.srgb_color_space, self.gray_color_space, self.p3_color_space]  # 列表中存放三个颜色空间显示标志位
        # 使用列表推导生成选项文本
        options = [f"{'✅' if state else ''}{name}" for state, name in zip(color_space_list, ["sRGB色域", "灰度图空间色域", "p3色域"])]
        # 清除并添加选项
        self.comboBox_2.clear(); self.comboBox_2.addItems(options)
        # 设置默认显示索引为当前激活的颜色空间
        self.ComBox2Curindex = next(i for i, state in enumerate(color_space_list) if state)
        self.comboBox_2.setCurrentIndex(self.ComBox2Curindex)
        self.comboBox_2.setFont(self.custom_font)

        # 设置复选框
        for checkbox in [self.checkBox_1, self.checkBox_2, self.checkBox_3, self.checkBox_4]:
            checkbox.setFont(self.custom_font)
        self.checkBox_1.setText("直方图")
        self.checkBox_2.setText("EXIF信息")
        self.checkBox_3.setText("ROI信息")
        self.checkBox_4.setText("AI提示看图")   

        # 根据self.dict_label_info_visibility设置复选框状态--> 配置在函数load_settings()
        self.checkBox_1.setChecked(self.dict_label_info_visibility.get("histogram_info", False))
        self.checkBox_2.setChecked(self.dict_label_info_visibility.get("exif_info", False))
        self.checkBox_3.setChecked(self.dict_label_info_visibility.get("roi_info", False))
        self.checkBox_4.setChecked(self.dict_label_info_visibility.get("ai_tips", False))
        

        # 设置表格列和行自动调整
        header = self.tableWidget_medium.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.Stretch)
        self.tableWidget_medium.verticalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.tableWidget_medium.verticalHeader().setVisible(False)
        self.tableWidget_medium.verticalHeader().setDefaultSectionSize(0)
        # self.tableWidget_medium.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch) # 设置表格列宽自适应
        header.setFont(self.custom_font)

        # 设置底部状态栏组件文本显示
        self.label_bottom.setText("📢:选中ROI信息复选框选后, 按下P键即可调出矩形框(矩形框移动逻辑同图片移动逻辑); 选中AI提示看图复选框选后, 按下P键即可发起请求(仅支持两张图); ")
        self.statusbar_button1.setText("(prev)🔼")
        self.statusbar_button2.setText("🔽(next)")


    def update_ui_styles(self):
        """更新所有UI组件的样式"""
        # 更新底部状态栏样式表
        statusbar_style = f"""
            QStatusBar {{
                background-color: {self.background_color_default};
                color: {self.font_color_default};
            }}
        """
        self.statusbar.setStyleSheet(statusbar_style)
        
        # 更新标签样式
        label_style = f"background-color: {self.background_color_default}; color: {self.font_color_default}; text-align: center; border-radius:10px;"
        self.label_0.setStyleSheet(label_style)
        statusbar_label_style = f"""
            color: {self.font_color_default}; 
            text-align: center;
            font-family: "{self.font_manager_jetbrains.family()}";
            font-size: {self.font_manager_jetbrains.pointSize()}pt;
        """
        self.label_bottom.setStyleSheet(statusbar_label_style)

        # 更新按钮样式
        statusbar_button_style = f"""
            QPushButton {{
                color: {self.font_color_default};
                text-align: center;
                font-family: "{self.font_manager_jetbrains.family()}";
                font-size: {self.font_manager_jetbrains.pointSize()}pt;
            }}
            QPushButton:hover {{
                background-color: {self.background_color_table};
                color: {self.font_color_default};
            }}
        """
        statusbar_left_button_style = f"""
            QPushButton {{
                border: none;
                color: {self.font_color_default};
                text-align: center;
                font-family: "{self.font_manager_jetbrains.family()}";
                font-size: {self.font_manager_jetbrains.pointSize()}pt;
            }}
            QPushButton:hover {{
                background-color: {self.background_color_table};
                color: {self.font_color_default};
            }}
        """
        self.statusbar_button1.setStyleSheet(statusbar_button_style)
        self.statusbar_button2.setStyleSheet(statusbar_button_style)
        self.statusbar_left_button.setStyleSheet(statusbar_left_button_style)

        # 更新复选框样式
        checkbox_style = f"""
        QCheckBox {{
            color: {self.font_color_default}; 
            font-weight: bold;
        }}"""
        for checkbox in [self.checkBox_1, self.checkBox_2, self.checkBox_3, self.checkBox_4]:
            checkbox.setStyleSheet(checkbox_style)

        # 更新下拉列表样式
        combobox_style = f"""
            QComboBox {{
                /* 下拉框本体样式*/
                background-color: {"rgb(240,240,240)"};                       /* 背景色 */
                color: {self.font_color_default};                             /* 字体颜色 */
                selection-background-color: {self.background_color_default};  /* 选中时背景色 */
                selection-color: {self.font_color_default};                   /* 选中时字体颜色 */
                min-height: 30px;                                             /* 最小高度 */
            }}
            /* 下拉框本体悬停样式*/
            QComboBox::hover {{
                background-color: {self.background_color_default};
                color: {self.font_color_default};
            }}   
            /* 下拉列表项样式*/
            QComboBox::item {{
                background-color: {self.background_color_default};
                color: {self.font_color_default};
            }}  
            /* 下拉列表样式*/
            QComboBox QAbstractItemView {{
                color: {self.font_color_default};              /* 字体颜色 */
                background-color: white;                       /* 背景色 */
                selection-color: {self.font_color_default};    /* 选中时字体颜色 */
                selection-background-color: {self.background_color_default}; /* 选中时背景色 */
            }}
            /* 下拉框列表项悬停样式*/
            QComboBox QAbstractItemView::item:hover {{
                background-color: {self.background_color_default};
                color: {self.font_color_default};
            }}
        """
       

        self.comboBox_1.setStyleSheet(combobox_style)
        self.comboBox_2.setStyleSheet(combobox_style)

        # 更新表格样式
        table_style = f"""
            QTableWidget {{
                background-color: {self.background_color_table};
                border: 1px solid black;
            }}
            QHeaderView::section {{
                background-color: {self.background_color_default};
                color: {self.font_color_default};
                text-align: center; 
                border-radius:10px;
            }}
        """
        self.tableWidget_medium.setStyleSheet(table_style)
        self.tableWidget_medium.horizontalHeader().setStyleSheet(table_style)

        # 更新所有图形视图的场景背景色和EXIF标签
        for view in self.graphics_views:
            if view and view.scene():
                # 更新场景背景色
                qcolor = rgb_str_to_qcolor(self.background_color_table)
                view.scene().setBackgroundBrush(QtGui.QBrush(qcolor))
                
                # 更新EXIF标签
                if hasattr(view, 'exif_label') and hasattr(view, 'exif_text'):
                    exif_info = self.process_exif_info(self.dict_exif_info_visibility, view.exif_text, False)
                    view.exif_label.setText(exif_info if exif_info else "解析不出exif信息!")
                    view.exif_label.setStyleSheet(f"color: {self.font_color_exif}; background-color: transparent; font-weight: 400;")



    def set_progress_bar(self):
        """设置进度条"""
        # 添加进度条并设置样式
        self.progress_bar = QProgressBar(self)
        # 初始化时设置进度条位置
        self.update_progress_bar_position()

        # 设置进度条样式
        self.progress_bar.setStyleSheet("""
            QProgressBar {
                text-align: center;
            }
            QProgressBar::chunk {
                background-color: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                stop:0 #36D1DC, stop:1 #5B86E5);
            }
        """)

        self.progress_bar.setAlignment(Qt.AlignCenter)  # 设置文字居中

        # 设置默认不显示
        self.progress_bar.setVisible(False) 
        

    def update_progress_bar_position(self):
        """更新进度条位置，确保其始终在窗口中心"""
        self.progress_bar.setGeometry(
            (self.width() - self.progress_bar.width()) // 2,
            (self.height() - self.progress_bar.height()) // 2,
            400, 40
        )

    def open_settings_window(self):
        """打开设置窗口"""
        # self.settings_window.show()
        print("打开设置窗口,还在开发中...")


    def update_progress(self, value):
        """更新进度条数值""" 
        self.progress_bar.setValue(value)
        self.progress_bar.repaint()
        # QApplication.processEvents()


    def resizeEvent(self, event):
        # 窗口大小改变时更新进度条位置
        self.update_progress_bar_position()

        # 获取表格的尺寸信息 print("table_width_heigth_default:", self.table_width_heigth_default)
        table_width = self.tableWidget_medium.width()
        table_height = self.tableWidget_medium.height()
        self.table_width_heigth_default = [table_width, table_height]

        super(SubMainWindow, self).resizeEvent(event)


    """怎么优化函数，加快处理速度"""
    def set_images(self, image_paths, index_list):
        """更新图片显示"""
        # 记录开始时间
        start_time1 = time.time()
        self.is_updating = True

        print("开始更新图片...")
        if not image_paths:
            print("没有有效的图片路径")
            return False

        # 更新当前显示的图片路径列表
        self.images_path_list = image_paths
        self.index_list = index_list

        # 调用封装后的函数,将看图界面图片索引发送到aebox中
        self.sync_image_index_with_aebox(self.images_path_list, self.index_list)

        # 设置进度条初始化
        if not hasattr(self, 'progress_bar'):
            self.set_progress_bar()
        # 设置进度条总数
        num_all = len(image_paths) + 5
        # 启动进度条显示
        self.progress_bar.setMaximum(num_all)
        self.progress_bar.setValue(0)
        self.progress_bar.setVisible(True)
        # 强制立即重绘界面
        # self.progress_bar.repaint()   # 重绘进度条
        # QApplication.processEvents()  # 处理所有挂起的事件
        
        try:
            # 确保表格可见
            self.tableWidget_medium.setUpdatesEnabled(False) # 禁用表格自动刷新
            self.tableWidget_medium.show()
            
            # 1. 预先分配数据结构
            self.progress_updated.emit(1)  # 发送进度条更新信号
            if self.parent_window:         # 主界面标签进度更新
                self.parent_window.statusbar_label1.setText(f"🔉: 正在更新图片...10%")
                self.parent_window.statusbar_label1.repaint()  # 刷新标签文本 
            self.cleanup()
            num_images = len(image_paths)
            self.exif_texts = [None] * num_images
            self.histograms = [None] * num_images
            self.original_rotation = [None] * num_images
            self.graphics_views = [None] * num_images
            self.original_pixmaps = [None] * num_images  
            self.gray_pixmaps = [None] * num_images  
            self.cv_imgs = [None] * num_images 
            self.pil_imgs = [None] * num_images 
            self.base_scales = [None] * num_images
            self._scales_min = [None] * num_images

            # 定义图片处理函数
            def process_image(args):
                """
                图片基础信息处理:
                (1) 图片旋转处理 (ImageTransform.auto_rotate_image)
                (2) EXIF信息获取 (self.get_exif_info)
                (3) 直方图计算 (self.calculate_brightness_histogram)
                (4) 图片亮度统计信息计算 (calculate_image_stats)
                
                """
                # 记录开始时间
                start_time = time.time()  
                index, path = args
                try:
                    # 如果图片是heic格式，则转换为jpg格式
                    if path.endswith(".heic"):
                        if new_path:= extract_jpg_from_heic(path):
                            path = new_path

                    # 如果图片不存在，则抛出异常
                    if not os.path.exists(path):
                        raise FileNotFoundError(f"❌ 图片不存在: {path}")

                    # 获取isinstance(image_input, Image.Image)格式图像
                    pil_image = Image.open(path)
                    iamge_format = pil_image.format

                    # 获取cv图像，方案一：转换PIL图像到OpenCV格式
                    if False:
                        cv_img = np.array(image.convert('RGB'))[:, :, ::-1].copy()

                    # 获取cv图像，方案二：使用 open 函数以二进制模式读取图片数据，使用 OpenCV 的 imdecode 函数解码图片数据
                    if True:   # 效率更高的方案
                        with open(path, "rb") as f:
                            image_data = np.asarray(bytearray(f.read()), dtype=np.uint8)
                        # 读取彩色图像
                        cv_img = cv2.imdecode(image_data, cv2.IMREAD_COLOR)
                        # 转换为灰度图
                        gray_img = cv2.cvtColor(cv_img, cv2.COLOR_BGR2GRAY)
                        # 将灰度图转换为QImage
                        height, width = gray_img.shape
                        gray_qimage = QImage(gray_img.data, width, height, width, QImage.Format_Grayscale8)
                        # 将QImage转换为QPixmap
                        gray_pixmap = QPixmap.fromImage(gray_qimage)

                    # 1. 处理图片旋转、基础exif信息获取（PIL）、图片文件格式信息获取
                    pixmap, basic_info = ImageTransform.auto_rotate_image(path, index_list[index], pil_image)

                    # 2. piexf解析曝光时间光圈值ISO等复杂的EXIF信息
                    exif_info_temp = self.get_exif_info(path, iamge_format)

                    if exif_info_temp != '\n':
                        exif_info = exif_info_temp + basic_info
                    else:
                        exif_info = basic_info

                    # 检测是否存在同图片路径的xml文件  将lux_index、DRCgain写入到exif信息中去
                    xml_path = os.path.join(os.path.dirname(path), os.path.basename(path).split('.')[0] + "_new.xml")
                    hdr_flag = False
                    if os.path.exists(xml_path):
                        # 提取xml中lux_index、cct、drcgain等关键信息，拼接到exif_info
                        exif_info_qpm, hdr_flag = load_xml_data(xml_path)
                        exif_info = exif_info + exif_info_qpm
                        

                    # 处理EXIF信息，根据可见性字典更新
                    exif_info = self.process_exif_info(self.dict_exif_info_visibility, exif_info, hdr_flag)

                    # 3. 解析直方图信息
                    histogram = self.calculate_brightness_histogram(pil_image) 

                    # 4. 计算亮度等统计信息
                    stats = calculate_image_stats(path, resize_factor=0.1)
                    # 移除LAB显示，替换为R/G和B/G
                    stats_text = f"亮度: {stats['avg_brightness']}\n对比度(L值标准差): {stats['contrast']}\nLAB: {stats['avg_lab']}\nRGB: {stats['avg_rgb']}\nR/G: {stats['R_G']}  B/G: {stats['B_G']}"

                    # 记录结束时间并计算耗时
                    end_time = time.time()
                    elapsed_time = end_time - start_time
                    print(f"处理图片{index}_{os.path.basename(path)} 耗时: {elapsed_time:.2f} 秒")
                
                    return index, {
                        'pixmap': pixmap,            # 原始pixmap格式图
                        'gray_pixmap':gray_pixmap,   # pixmap格式灰度图
                        'exif_info': exif_info,      # exif信息
                        'histogram': histogram,      # 直方图信息
                        'stats': stats_text,         # 添加亮度/RGB/LAB等信息
                        'pil_image':pil_image,       # PIL图像
                        'cv_img':cv_img              # OpenCV图像
                    }
                except Exception as e:
                    print(f"处理图片失败 {path}: {e}")
                    return index, None
                    

            # 2. 使用线程池并行处理图片
            self.progress_updated.emit(2)
            if self.parent_window:
                self.parent_window.statusbar_label1.setText(f"🔉: 正在更新图片...20%")
                self.parent_window.statusbar_label1.repaint()  # 刷新标签文本 
            max_workers = min(len(image_paths), cpu_count() - 2)
            with ThreadPoolExecutor(max_workers=max_workers) as executor:
                futures = list(executor.map(process_image, enumerate(image_paths)))

            # 4. 计算目标尺寸
            self.progress_updated.emit(3)
            if self.parent_window:
                self.parent_window.statusbar_label1.setText(f"🔉: 正在更新图片...40%")
                self.parent_window.statusbar_label1.repaint()  # 刷新标签文本 
            # 使用生成器表达式提高效率
            valid_sizes = ((result[1]['pixmap'].width(), result[1]['pixmap'].height()) for result in futures if result and result[1])
            # 计算多张图片中的最大宽（max_width）和高（max_height）
            widths, heights = zip(*valid_sizes)
            max_width, max_height = max(widths), max(heights)


            # 计算加权平均宽高比（根据图片面积加权）
            total_area = sum(w * h for w, h in zip(widths, heights))      
            avg_aspect_ratio = sum((w/h) * (w*h)/total_area for w, h in zip(widths, heights))
        
            # 根据动态阈值判断方向，确定多张图片的统一宽（target_width）和高（target_height）
            aspect_threshold = 1.2  # 可调整的阈值参数
            if avg_aspect_ratio > aspect_threshold:  # 明显横向
                target_width = max_width
                target_height = int(target_width / avg_aspect_ratio)
            elif avg_aspect_ratio < 1/aspect_threshold:  # 明显纵向
                target_height = max_height
                target_width = int(target_height * avg_aspect_ratio)
            else:  # 接近方形
                target_width = int((max_width + max_height * avg_aspect_ratio) / 2)
                target_height = int((max_height + max_width / avg_aspect_ratio) / 2)
            

            # 4. 更新表格设置
            self.progress_updated.emit(4)
            if self.parent_window:
                self.parent_window.statusbar_label1.setText(f"🔉: 正在更新图片...80%")
                self.parent_window.statusbar_label1.repaint()  # 刷新标签文本 
            self.tableWidget_medium.setUpdatesEnabled(True) # 表格自动刷新
            self.tableWidget_medium.clearContents()
            self.tableWidget_medium.setColumnCount(num_images)
            self.tableWidget_medium.setRowCount(1)
            folder_names = [os.path.basename(os.path.dirname(path)) for path in image_paths]
            if len(set(folder_names)) == 1: # 如果所有图片都在同一个文件夹中，则显示文件名
                folder_names = [os.path.basename(path) for path in image_paths]
            self.tableWidget_medium.setHorizontalHeaderLabels(folder_names) # 更新表头


            # 5. 批量更新UI并计算基准缩放比例
            self.progress_updated.emit(5)
            if self.parent_window:
                self.parent_window.statusbar_label1.setText(f"🔉: 正在更新图片...90%")
                self.parent_window.statusbar_label1.repaint()  # 刷新标签文本 
            for index, result in enumerate(futures):
                if result and result[1]:
                    data = result[1]

                    # 这是原始的彩色图pixmap
                    pixmap = data['pixmap']
                    # 跟根据下拉框2判断是否获取灰度图gray_pixmap
                    if self.comboBox_2.currentIndex() == 1:
                        pixmap = data['gray_pixmap']
                    
                    # 创建和设置场景
                    qcolor = rgb_str_to_qcolor(self.background_color_table) # 将背景色转换为QColor
                    scene = QGraphicsScene(self)
                    scene.setBackgroundBrush(QtGui.QBrush(qcolor)) # 设置场景背景色

                    
                    # 创建图片项
                    pixmap_item = QGraphicsPixmapItem(pixmap)
                    pixmap_item.setTransformOriginPoint(pixmap.rect().center())
                    scene.addItem(pixmap_item)
                    
                    # 创建和设置视图
                    view = MyGraphicsView(scene, data['exif_info'], data['stats'], self)
                    view.pixmap_items = [pixmap_item]
                    
                    # 计算缩放比例
                    current_width, current_height = (pixmap.width(), pixmap.height())
                    scale_x = target_width / current_width
                    scale_y = target_height / current_height

                    # 保持宽高比,得到当前图片尺寸到统一尺寸的缩放比例
                    base_scale = min(scale_x, scale_y)  

                    # 计算图片缩放比例以适应当前表格大小显示 
                    zoom_scale = self.set_zoom_scale(num_images,avg_aspect_ratio,target_width,target_height)
                    base_scale = base_scale*zoom_scale

                    # 设置视图的缩放
                    view.scale(base_scale, base_scale)
                    
                    # 设置直方图、EXIF、亮度统计信息
                    view.set_histogram_visibility(self.checkBox_1.isChecked())
                    if data['histogram']:
                        view.set_histogram_data(data['histogram'])  
                    view.set_exif_visibility(self.checkBox_2.isChecked(), self.font_color_exif)
                    # view.set_stats_visibility(self.checkBox_3.isChecked())
                    view.set_stats_visibility(self.stats_visible) 

                    # 设置原始OpenCV图像
                    view.set_original_image(data['cv_img'])

                    # 保存数据
                    self.graphics_views[index] = view
                    self.original_pixmaps[index] = data['pixmap']
                    self.gray_pixmaps[index] = data['gray_pixmap']
                    self.original_rotation[index] = pixmap_item.rotation()
                    self.exif_texts[index] = data['exif_info']
                    self.histograms[index] = data['histogram']
                    self.cv_imgs[index] = data['cv_img']
                    self.pil_imgs[index] = data['pil_image']
                    # 保存基准缩放比例
                    self.base_scales[index] = base_scale
                    # 设置缩放的最大最小基准
                    self._scales_min[index] = base_scale

                        
                    # 更新表格
                    self.tableWidget_medium.setCellWidget(0, index, view)

                # 更新进度条
                self.progress_updated.emit(index  + 7)
                if self.parent_window:
                    self.parent_window.statusbar_label1.setText(f"🔉: 正在更新图片...100%")
                    self.parent_window.statusbar_label1.repaint()  # 刷新标签文本    

            return True
        except Exception as e:
            print(f"更新图片时发生错误: {e}")
            return False
        finally:
            # 完成后恢复确定模式
            self.progress_bar.setVisible(False)  # 隐藏进度条
            self.is_updating = False

            # 记录结束时间并计算耗时
            end_time1 = time.time()
            elapsed_time = end_time1 - start_time1
            print(f"处理图片总耗时: {elapsed_time:.2f} 秒")


    def sync_image_index_with_aebox(self, images_path_list, index_list):
        """同步当前图片索引到aebox应用"""
        try:
            # 预检查程序aebox是否启动
            if not check_process_running("aebox.exe"):
                print("❌ sync_image_index_with_aebox()--同步当前图片索引到aebox应用失败--aebox应用未启动")
                return False

            # 新增配置文件读取
            host = get_aebox_host()
            origin_image_names = [os.path.basename(path) for path in images_path_list]

            # 发送初始索引
            select_url = f"{host}/select_image/{index_list[0].split('/')[0]}"
            if not get_api_data(url=select_url, timeout=3):
                print("❌ sync_image_index_with_aebox()--初始索引发送失败")
                return False

            # 获取aebox当前图片信息
            current_data = json.loads(get_api_data(
                url=f"{host}/current_image", 
                timeout=3) or '{}'
            )
            current_name = current_data.get('filename', '')

            if current_name and current_name in origin_image_names:
                print(f"✅ sync_image_index_with_aebox()--初始索引发送成功匹配: {current_name}")
                return True

            # 执行图片列表匹配
            list_data = json.loads(get_api_data(
                url=f"{host}/image_list",
                timeout=3) or '{}'
            )
            aebox_images = list_data.get('filenames', [])

            # 使用集合提高查找效率
            origin_set = set(origin_image_names)
            matching_indices = [i for i, name in enumerate(aebox_images) if name in origin_set]

            if len(matching_indices) == 1:
                new_index = matching_indices[0] + 1
                if get_api_data(f"{host}/select_image/{new_index}", timeout=3):
                    print(f"✅ sync_image_index_with_aebox()--成功同步图片到aebox: {aebox_images[matching_indices[0]]}")
                    return True

            print("❌ sync_image_index_with_aebox()--未找到唯一匹配的图片")
            return False

        except Exception as e:
            print(f"❌ sync_image_index_with_aebox()--同步索引异常: {str(e)}")
            return False


    def process_exif_info(self, visibility_dict, exif_info, hdr_flag):
        """处理EXIF信息，将其转换为字典并根据可见性字典更新"""
        try:
            # 将 exif_info 转换为字典
            exif_dict = convert_to_dict(exif_info)

            # HDR标签为auto时，如果hdr_flag为True，则设置为auto-on，否则设置为auto-off
            # hdr_flag 是读取xml文件时，是否存在EVFrameSA标签判断
            if exif_dict.get("HDR", "") == 'auto' and  exif_dict.get("Lux", ""):
                exif_dict['HDR'] = 'auto_(ON)' if hdr_flag else 'auto_(OFF)'
            
            # 根据字典中的键值对，更新 exif_dict 中的可见性值
            for key, value in visibility_dict.items():
                if not value and key in exif_dict:  # 仅在键存在时才删除
                    exif_dict.pop(key)

            # 按照 visibility_dict 的顺序生成字符串
            result = []
            for key in visibility_dict.keys():
                if key in exif_dict:  # 仅添加存在于 exif_dict 中的键
                    result.append(f"{key}: {exif_dict[key]}")

            return '\n'.join(result)

        except Exception as e:
            print(f"处理EXIF信息时发生错误: {e}")
            return ""  # 返回空字符串或其他适当的默认值


    def set_zoom_scale(self, num_images,avg_aspect_ratio,target_width,target_height):
        """根据图片数量设置缩放因子"""
        # 计算表格中单个单元格的宽度和高度
        sigle_table_w = (self.table_width_heigth_default[0]-18)/num_images
        sigle_table_h = self.table_width_heigth_default[1]-55

        if avg_aspect_ratio > 1: #横向图片
            # 以当前的最大宽度为基准
            zoom_scale = sigle_table_w/target_width
            while ((zoom_scale*target_height) >= sigle_table_h):
                zoom_scale *= 0.995
    
        else: #纵向图片
            zoom_scale = sigle_table_h/target_height
            while ((zoom_scale*target_width) >= sigle_table_w):
                zoom_scale *= 0.995

        # 更新当前缩放因子, zoom_scale = 0 时 直接设置为 1.0不缩放
        zoom_scale = zoom_scale if zoom_scale != 0 else 1.0

        return zoom_scale
    

    def cleanup(self):
        """清理所有资源"""
        try:
            # 清理表格
            self.tableWidget_medium.clearContents()
            self.tableWidget_medium.setRowCount(0)
            self.tableWidget_medium.setColumnCount(0)
            
            # 清理场景和视图
            for view in self.graphics_views:
                if view:
                    if view.scene():
                        view.scene().clear()
                    view.setScene(None)

            if self.roi_selection_active: # 切换图片自动清除ROI信息框
                self.roi_selection_active = False
            
            # 清理所有列表
            self.graphics_views.clear()
            self.exif_texts.clear()
            self.histograms.clear()
            self.original_pixmaps.clear()
            self.original_rotation.clear()
            self.base_scales.clear()
            self._scales_min.clear()

            # 强制垃圾回收
            gc.collect()
            
        except Exception as e:
            print(f"清理资源时发生错误: {e}")


    def show_menu_combox1(self, index):
        """下拉框self.comboBox_1中显示多级菜单项
        下拉框1的主选项如下:
            ["📌颜色设置", "🔁一键重置", "🔽背景颜色", "🔽表格填充颜色", "🔽字体颜色", "🔽exif字体颜色"]
        """
        try:
            if not index:     # index == 0 颜色设置，不做任何操作
                print("show_menu_combox1()-看图子界面--点击了颜色配置选项")
                # 从json文件加载配置
                self.load_settings()
                # 更新样式表
                self.update_ui_styles()
            elif index == 1:  # index == 1 一键重置
                self.background_color_default = "rgb(173,216,230)" # 背景默认色_好蓝
                self.background_color_table = "rgb(127,127,127)"   # 表格填充背景色_18度灰
                self.font_color_default = "rgb(0, 0, 0)"           # 字体默认颜色_纯黑色
                self.font_color_exif = "rgb(255,255,255)"          # exif字体默认颜色_纯白色
                self.comboBox_1.setCurrentIndex(0)                 # 设置默认显示索引为0
                # 更新样式表
                self.update_ui_styles()
            else: 
                # 创建菜单
                self.menu_1 = QtWidgets.QMenu(self)

                # 设置菜单项悬停样式
                hover_bg = self.background_color_default  # 背景颜色
                hover_text = self.font_color_default      # 字体颜色
                self.menu_1.setStyleSheet(f"""
                    QMenu::item:selected {{
                        background-color: {hover_bg};
                        color: {hover_text};
                    }}
                    QMenu::item:hover {{
                        background-color: {hover_bg};
                        color: {hover_text};
                    }}
                """)

                # 定义颜色选项 从self.color_rgb_settings中获取
                # color_options = ['18度灰', '石榴红', '乌漆嘛黑', '铅白', '水色', '石青', '茶色', '天际', '晴空', '苍穹', 
                # '湖光', '曜石', '天际黑', '晴空黑', '苍穹黑', '湖光黑', '曜石黑']
                color_options = list(self.color_rgb_settings.keys())

                # 添加颜色选项到菜单
                for color in color_options:
                    action = QtWidgets.QAction(color, self)
                    # 传递 color 和 index
                    action.triggered.connect(lambda checked, color=color, index=index: self.on_comboBox_1_changed(color, index))  
                    self.menu_1.addAction(action)
                self.menu_1.setFont(self.custom_font)

                # 获取 QComboBox 顶部的矩形区域
                rect = self.comboBox_1.rect()
                global_pos = self.comboBox_1.mapToGlobal(rect.bottomLeft())

                # 弹出 QMenu
                self.menu_1.exec_(global_pos)
        except Exception as e:
            print(f"self.comboBox_1()--处理下拉框选项时发生未知错误: {e}")
        

    def on_comboBox_1_changed(self, color, index):
        """颜色设置二级菜单触发事件"""
        """优化方案说明：
        1. 使用字典映射替代多个if-elif分支，提高可维护性
        2. 使用海象运算符(walrus operator)合并条件判断使用setattr动态设置属性，避免重复代码
        4. 将关联操作集中到单个条件判断中，逻辑更紧凑保持原有功能不变，但代码行数减少50%
        6. 更易于扩展新的颜色配置项，只需更新index_map字典即可
        """
        if True: # 优化方案
            # 使用字典映射索引与属性名的关系
            index_map = {
                2: 'background_color_default',
                3: 'background_color_table', 
                4: 'font_color_default',
                5: 'font_color_exif'
            }
            
            if color_rgb := self.color_rgb_settings.get(color):
                if prop_name := index_map.get(index):
                    setattr(self, prop_name, color_rgb)
                    self.update_ui_styles()
                    self.comboBox_1.setCurrentIndex(0)
        if False:  # 原始方案
            # 根据选择的color颜色从color_rgb_settings中获取rgb
            color_rgb = self.color_rgb_settings.get(color, "")
            if color_rgb:
                if index==2:
                    self.background_color_default = color_rgb
                elif index ==3:
                    self.background_color_table = color_rgb
                elif index ==4:
                    self.font_color_default = color_rgb
                elif index ==5:
                    self.font_color_exif = color_rgb
            # 更新样式表
            self.update_ui_styles()
            # 设置默认显示索引为0
            self.comboBox_1.setCurrentIndex(0)                 
            # print(f"Selected color: {color}, Index: {index}")
        

    def update_comboBox2(self):
        """更新下拉框self.comboBox_2的显示"""
        # 定义颜色空间状态列表
        color_spaces = [
            (self.srgb_color_space, "sRGB色域"),
            (self.gray_color_space, "灰度图空间色域"), 
            (self.p3_color_space, "p3色域")
        ]
        
        # 使用列表推导生成选项文本
        options = [f"{'✅' if state else ''}{label}" for state, label in color_spaces]
        
        # 批量更新下拉框选项
        for i, text in enumerate(options):
            self.comboBox_2.setItemText(i, text)
    

    def clean_color_space(self,):
        """清除颜色空间的显示标志位"""
        self.srgb_color_space = False
        self.gray_color_space = False
        self.p3_color_space = False


    def on_comboBox_2_changed(self, index):
        """图像色彩显示空间下拉框self.comboBox_2内容改变时触发事件
        ["✅sRGB色域", "✅灰度图色域", "✅p3色域"]
        """
        # 更新所有图形视图的场景视图
        for i, view in enumerate(self.graphics_views):
            if view and view.scene() :
                try:
                    original_pixmap = self.original_pixmaps[i]
                    current_rotation = view.pixmap_items[0].rotation() if view.pixmap_items else 0
                    
                    # 根据选择的色彩空间转换图像
                    if index == 0 :  # sRGB色域
                        # 设置当前启用的图像色彩显示空间
                        self.clean_color_space()
                        self.srgb_color_space = True
                        self.update_comboBox2()
                        converted_pixmap = original_pixmap
                    elif index == 1 and self.gray_pixmaps[i] is not None:  # 灰度图色域
                        # 设置当前启用的图像色彩显示空间
                        self.clean_color_space()
                        self.gray_color_space = True
                        self.update_comboBox2()
                        # 调用列表self.gray_pixmaps[i]中存储的灰度图pixmap
                        converted_pixmap = self.gray_pixmaps[i]
                    elif index == 2 and self.pil_imgs[i] is not None and self.cv_imgs[i] is not None:  # p3色域
                        # 设置当前启用的图像色彩显示空间
                        self.clean_color_space()
                        self.p3_color_space = True
                        self.update_comboBox2()
                        # 调用PIL图像转P3色域图像的方法
                        converted_pixmap = convert_to_dci_p3(self.cv_imgs[i], self.pil_imgs[i],self.original_pixmaps[i])

                    # 更新视图显示
                    view.pixmap_items[0].setPixmap(converted_pixmap)
                    view.pixmap_items[0].setRotation(current_rotation)
                    view.centerOn(view.mapToScene(view.viewport().rect().center()))
                    
                    # 更新场景背景色
                    qcolor = rgb_str_to_qcolor(self.background_color_table)
                    view.scene().setBackgroundBrush(QtGui.QBrush(qcolor))
                    
                except Exception as e:
                    print(f"on_comboBox_2_changed()-色彩空间转换失败: {str(e)}")
        # 更新UI
        self.update()
        QApplication.processEvents() 
                
                
    def toggle_exif_info(self, state):
        print(f"切换 EXIF 信息: {'显示' if state == Qt.Checked else '隐藏'}")
        try:
            for view, exif_text in zip(self.graphics_views, self.exif_texts):
                if exif_text:
                    # 传入字体颜色参数
                    view.set_exif_visibility(state == Qt.Checked, self.font_color_exif)
        except Exception as e:
            print(f"处理toggle_exif_info函数时发生错误: {e}")


    def toggle_histogram_info(self, state):
        print(f"切换直方图信息: {'显示' if state == Qt.Checked else '隐藏'}")
        try:
            for view, histogram in zip(self.graphics_views, self.histograms):
                if histogram:
                    view.set_histogram_visibility(state == Qt.Checked)
        except Exception as e:
            print(f"处理toggle_histogram_info函数时发生错误: {e}")

    def ai_tips_info(self, state):
        try:    
            if state == Qt.Checked:
                self.ai_tips_flag = True
                self.is_updating = False
                self.label_bottom.setText(f"📢:开启AI提示看图复选框提示, 按下快捷键P发起请求(仅支持两张图). 另: 关闭AI提示看图复选框, 打开ROI信息复选框的状态下, 按P键才会调出矩形框")
            else:
                self.ai_tips_flag = False
                self.is_updating = False
                self.label_bottom.setText(f"📢:选中ROI信息复选框选后, 按下P键即可调出矩形框(矩形框移动逻辑同图片移动逻辑); 选中AI提示看图复选框选后, 按下P键即可发起请求(仅支持两张图);")
        except Exception as e:
            print(f"处理ai_tips_info函数时发生错误: {e}")

    """modify by diamond_cz 20250218 移除该函数，使用更加高效的opencv方法实现
    def calculate_brightness_histogram(self, path):
        try:
            image = Image.open(path).convert('L')  # 转换为灰度图
            histogram = image.histogram()
            # 只保留0-255的灰度值
            histogram = histogram[:256]
            return histogram
        except Exception as e:
            print(f"计算直方图失败: {path}\n错误: {e}")
            return None
    """ 
    
    def calculate_brightness_histogram(self, img):
        """传入PIL图像img,将其转换为灰度图, 输出直方图"""
        try:
            # 处理PIL图像对象
            if isinstance(img, Image.Image):  
                # 转换PIL图像到OpenCV格式
                pass
            else:
                print(f"calculate_image_stats无法加载图像")
                return None

            # 转换为灰度图
            gray_img = img.convert('L')
            
            # 使用numpy计算直方图
            histogram = np.array(gray_img).flatten()
            hist_counts = np.bincount(histogram, minlength=256)

            return hist_counts.tolist()
        except Exception as e:
            print(f"calculate_brightness_histogram计算直方图失败\n错误: {e}")
            return None
        
    """modify by diamond_cz 20250218 移除该函数， 在类ImageTransform中实现图片基础信息的获取， 减少使用Image.open
    def pic_size(self, path, index):
        # 获取图片名称、尺寸、大小等基础信息
        pic_name = os.path.basename(path)
        image = Image.open(path)
        width, height = image.size
        file_size = os.path.getsize(path)  # 文件大小（字节）
        if file_size < 1024:
            size_str = f"{file_size} B"
        elif file_size < 1024 ** 2:
            size_str = f"{file_size / 1024:.2f} KB"
        else:
            size_str = f"{file_size / (1024 ** 2):.2f} MB"
        exif_size_info = f"图片名称: {pic_name}\n图片大小: {size_str}\n图片尺寸: {width} x {height}\n图片张数: {index}"
        return exif_size_info, image
    """
    
    
    def get_exif_info(self, path, image_format):
        """
        函数功能： 使用piexif解析特定格式（"JPEG", "TIFF", "MPO"）图片的曝光时间、光圈、ISO等详细信息
        输入： path 图片文件路径, image_format图片文件的PIL_image 格式
        输出： exif_info 解析出来的详细信息（exif_tags_id）
        """
        try:
            # 使用PIL解析基本信息 modify by diamond_cz 20250218 移除该函数， 在类ImageTransform中实现图片基础信息的获取， 减少使用Image.open
            # exif_size_info, image = self.pic_size(path, index)

            # 检查文件格式，PIL 或 piexif 只能处理 JPEG 或 TIFF 文件
            if image_format not in ["JPEG", "TIFF", "MPO"]:
                exif_info = "" 
                return exif_info

            # 使用piexif读取EXIF信息
            exif_dict = piexif.load(path)
            
            # 如果存在EXIF信息
            if exif_dict and "0th" in exif_dict:

                exif_tags_id = {
                    "271": "品牌",  # exif_dict["0th"]
                    "272": "型号",  # exif_dict["0th"]
                    "33434": "曝光时间",
                    "33437": "光圈值",
                    "34855": "ISO值",
                    "36867": "原始时间",
                    # "306": "文件修改时间",   # exif_dict["0th"][306] 移除该项
                    "37383": "测光模式", # 需要单独处理
                }
                
                # 测光模式映射
                metering_mode_mapping = {
                    0: "未知",
                    1: "平均测光",
                    2: "中央重点测光",
                    3: "点测光",
                    4: "多点测光",
                    5: "多区域测光",
                    6: "部分测光",
                    255: "其他"
                }
                # 存储解析exif_dict["Exif"]，exif_dict["0th"]得到的数据
                # 图像旋转信息 exif_dict["0th"][274]
                exif_info_list = []
                for tag_id, tag_cn in exif_tags_id.items():
                    # 根据标签id获取相应数据并解析成需要的数据形式
                    # 将字符串类型转换为整型
                    tag_id = int(tag_id) 
                    # 解析Exif
                    if tag_id in exif_dict["Exif"]:
                        value = exif_dict["Exif"][tag_id]   #.decode('utf-8')
                        if value:
                            if isinstance(value, bytes): # 字节类型处理
                                value = value.decode('utf-8')
                            if tag_id == 33434: # 曝光时间处理
                                exp_s = (value[0]/value[1])*1000000
                                # 设置保留小数点后两位
                                exp_s = round(exp_s, 2)
                                exp_s = f"{exp_s}ms"
                                if value[0] == 1:
                                    value = f"{value[0]}/{value[1]}_({exp_s})"
                                else: # 处理曝光时间分子不为1的情况
                                    value = int(value[1]/value[0])
                                    value = f"1/{value}_({exp_s})"
                            if tag_id == 33437: # 光圈值处理
                                value = value[0]/value[1]
                                value = round(value, 2)
                            if tag_id == 37383: # 测光模式处理
                                value = metering_mode_mapping.get(value, "其他")
                            exif_info_list.append(f"{tag_cn}: {value}")
                    # 解析0th
                    elif tag_id in exif_dict["0th"]:
                        value = exif_dict["0th"][tag_id]
                        if value:
                            if isinstance(value, bytes):
                                value = value.decode('utf-8')
                            exif_info_list.append(f"{tag_cn}: {value}")
                        
                exif_info = "\n".join(exif_info_list)
            else:
                exif_info = ""
                return exif_info
            
            exif_info = exif_info + '\n'
            return exif_info
        except Exception as e:
            return f"无法读取EXIF信息:{os.path.basename(path)}\n错误: {e}"


    def wheelEvent(self, event: QEvent):
        """鼠标滚轮事件"""
        try:
            # 图片还在更新中，不触发鼠标滚轮事件
            if self.is_updating:
                print("wheelEvent事件: 图片还在更新中,请稍等...") 
                return

            # 确保视图中有值&存在基准缩放比例
            if not self.graphics_views or not self.base_scales:
                print("wheelEvent事件: 无效的视图或基准缩放比例")
                return

            # 确保 self._scales_min 中没有 None 值
            if None in self._scales_min:
                print("wheelEvent事件: 缩放最小值包含None，重新初始化")
                # 使用默认值替换None
                self._scales_min = [0.2 if x is None else x for x in self._scales_min]

            # 计算新的缩放因子
            zoom_step = 1.15 if event.angleDelta().y() > 0 else 0.9
            
            # 更新基准尺寸信息,并限制缩放范围（0.08<0.5*0.20>~130<800*0.20>）
            self.base_scales = [
                max(0.5 * self._scales_min[i], min(scale * zoom_step, 800 * self._scales_min[i])) 
                for i, scale in enumerate(self.base_scales) if scale is not None
            ]

            # 如果按下了Ctrl键，则仅缩放鼠标所在的视图,否则同步缩放所有视图
            if event.modifiers() & Qt.ControlModifier:
                # 仅缩放鼠标所在的视图
                pos = self.mapFromGlobal(event.globalPos())
                for i, view in enumerate(self.graphics_views):
                    if view and view.rect().contains(view.mapFromParent(pos)) and i < len(self.base_scales):
                        view_new = self._apply_scale_to_view(view, self.base_scales[i])
                        self.graphics_views[i] = view_new
                        break    
            else:
                # 同步缩放所有视图
                for i, view in enumerate(self.graphics_views):
                    if view and i < len(self.base_scales):
                        view_new = self._apply_scale_to_view(view, self.base_scales[i])
                        self.graphics_views[i] = view_new

        except Exception as e:
            print(f"处理滚轮事件时发生错误: {e}")
      

    def _apply_scale_to_view(self, view, zoom_step):
        """应用缩放到指定视图"""
        try:
            if not view.pixmap_items:
                return

            # 获取当前视图中心
            center = view.mapToScene(view.viewport().rect().center())
            
            # 计算并应用新的变换
            new_transform = QtGui.QTransform()
            
            # 设置新的变换矩阵
            new_transform.scale(zoom_step, zoom_step)
            
            # 应用变换
            view.setTransform(new_transform)
            
            # 保持视图中心
            view.centerOn(center)
            
        except Exception as e:
            print(f"应用缩放时发生错误: {e}")
        finally:
            return view
        

    def toggle_fullscreen(self):
        """F11全屏快键键, 切换全屏"""
        self.is_fullscreen = not self.is_fullscreen
        try:
            if self.is_fullscreen:
                self.showFullScreen()  # 全屏 hl_top
                self.label_bottom.setVisible(False)
                # 隐藏
                for i in range(self.hl_top.count()):
                    item = self.hl_top.itemAt(i)
                    if item.widget():
                        item.widget().setVisible(False)
                
            else:
                self.label_bottom.setVisible(True)
                # 显示
                for i in range(self.hl_top.count()):
                    item = self.hl_top.itemAt(i)
                    if item.widget():
                        item.widget().setVisible(True)
                self.showMaximized()   # 最大化
        except Exception as e:
            print(f"应用F11切换全屏时发生错误: {e}")


    def rotate_left(self):
        try:
            self.rotate_image(-90)
        except Exception as e:
            print(f"旋转图片时发生错误: {e}")

    def rotate_right(self):
        try:
            self.rotate_image(90)
        except Exception as e:
            print(f"旋转图片时发生错误: {e}")

    def rotate_image(self, angle):
        """旋转图片函数，接受角度进行旋转图片"""
        # 获取鼠标的全局位置
        cursor_pos = QCursor.pos()
        # 将全局位置转换为窗口内的位置
        pos = self.mapFromGlobal(cursor_pos)
        
        for i, view in enumerate(self.graphics_views):
            if view is None:
                continue
            
            # 将全局坐标转换为view的本地坐标
            local_pos = view.mapFromParent(pos)
            if view.rect().contains(local_pos):
                # 获取当前视图中的图片项
                pixmap_item = view.pixmap_items[0] if view.pixmap_items else None
                if pixmap_item:
                    # 获取当前图片的原始尺寸
                    original_rect = pixmap_item.boundingRect()
                    
                    # 保存当前变换状态
                    current_transform = pixmap_item.transform()
                    current_scale = current_transform.m11()
                    current_rotation = pixmap_item.rotation()
                    
                    # 设置新的旋转角度
                    new_rotation = current_rotation + angle
                    
                    # 重置变换
                    pixmap_item.setTransform(QTransform())
                    
                    # 设置旋转中心点
                    pixmap_item.setTransformOriginPoint(original_rect.center())
                    
                    # 应用新的旋转角度
                    pixmap_item.setRotation(new_rotation)
                    
                    # 重新应用缩放
                    pixmap_item.setScale(current_scale)

                    # 计算旋转后的边界
                    rotated_rect = pixmap_item.mapRectToScene(pixmap_item.boundingRect())
                    
                    # 更新场景边界
                    view.scene().setSceneRect(rotated_rect)
                    
                    # 确保图片保持在视图中心
                    view.centerOn(pixmap_item)

                    # 保存当前旋转角度
                    self.original_rotation[i] = pixmap_item.rotation()

                break


    def keyPressEvent(self, event):
        if event.isAutoRepeat():  # 忽略按键重复
            event.accept()
            return

        if event.key() == Qt.Key_Q:
            self.handle_overlay('q')
        elif event.key() == Qt.Key_W:
            self.handle_overlay('w')


    def keyReleaseEvent(self, event):
        if event.isAutoRepeat():
            return  # 忽略自动重复事件

        if event.key() == Qt.Key_Q:
            self.restore_images('q')
        elif event.key() == Qt.Key_W:
            self.restore_images('w')


    def handle_overlay(self, key):
        """按压的动作函数"""
        if len(self.images_path_list) != 2:
            QMessageBox.warning(self, "警告", "只有两张图片时才能使用覆盖比较功能。")
            return
        
        try:    
            def create_unified_overlay(source_view, target_view):
                """创建统一尺寸的覆盖图像"""
                try:
                    source_pixmap_item = source_view.pixmap_items[0]

                    source_pixmap = source_pixmap_item.pixmap()

                    target_display_size = target_view.pixmap_items[0].boundingRect().size().toSize()

                    # 使用忽略宽高比的缩放方式
                    scaled_pixmap = source_pixmap.scaled(
                        target_display_size,
                        Qt.IgnoreAspectRatio,
                        Qt.SmoothTransformation
                    )

                    return scaled_pixmap
                except Exception as e:
                    print(f"创建统一覆盖图像失败: {str(e)}")
                    return None

            if key == 'q':
                source_view = self.graphics_views[1]
                target_view = self.graphics_views[0]
                if source_view and target_view and source_view.pixmap_items and target_view.pixmap_items:
                    scaled = create_unified_overlay(source_view, target_view)
                    source_rotation = self.original_rotation[1]
                    target_view.pixmap_items[0].setPixmap(scaled)
                    target_view.pixmap_items[0].setRotation(source_rotation)
                    target_view.centerOn(target_view.mapToScene(target_view.viewport().rect().center()))
            elif key == 'w':
                source_view = self.graphics_views[0]
                target_view = self.graphics_views[1]
                if source_view and target_view and source_view.pixmap_items and target_view.pixmap_items:
                    scaled = create_unified_overlay(source_view, target_view)     
                    source_rotation = self.original_rotation[0]
                    target_view.pixmap_items[0].setPixmap(scaled)
                    target_view.pixmap_items[0].setRotation(source_rotation)
                    target_view.centerOn(target_view.mapToScene(target_view.viewport().rect().center()))

        except Exception as e:
            print(f"覆盖操作失败: {e}")
            

    def restore_images(self, key):
        """释放动作的函数"""
        if len(self.images_path_list) != 2:
            return

        try:
            if key == 'q':
                index = 0
            elif key == 'w':
                index = 1
            else:
                return

            target_view = self.graphics_views[index]
            original_pixmap = self.original_pixmaps[index]
            original_rotation = self.original_rotation[index]


            if not target_view or not target_view.pixmap_items or not original_pixmap:
                return

            # 获取当前显示尺寸
            current_display_size = target_view.pixmap_items[0].boundingRect().size().toSize()

            # 重新缩放原始图像到当前显示尺寸
            scaled = original_pixmap.scaled(
                current_display_size,
                Qt.IgnoreAspectRatio,
                Qt.SmoothTransformation
            )

            # 直接设置缩放后的图像到目标视图
            target_view.pixmap_items[0].setPixmap(scaled)
            target_view.pixmap_items[0].setRotation(original_rotation)  
            target_view.centerOn(target_view.mapToScene(target_view.viewport().rect().center()))
        except Exception as e:
            print(f"恢复图片失败: {e}")

            
    def on_v_pressed(self):
        """处理V键事件"""
        print("按下了v键")

        if False:                            
            # 打印表格的实际尺寸
            table_width = self.tableWidget_medium.width()
            table_height = self.tableWidget_medium.height()
            table_columns = self.tableWidget_medium.columnCount()
            target_width = table_width/table_columns
            target_height = table_height/table_columns
            print(f"表格的实际尺寸: {table_width}x{table_height}，列数: {table_columns}")  # 添加调试信息

        try:
            # 设置缩放因子为0.995
            zoom_step = 0.995
            # 更新基准尺寸信息,并限制缩放范围（0.08<0.5*0.16>~130<800*0.16>） 使用self._scales_min限制是由于适应不同尺寸的图片
            self.base_scales = [max(0.5*self._scales_min[i], min(scale * zoom_step, 800*self._scales_min[i])) for i, scale in enumerate(self.base_scales)]

            # 应用缩放到所有视图
            for i, view in enumerate(self.graphics_views):
                if view:
                    self._apply_scale_to_view(view, self.base_scales[i])
            # print(f"新的缩放因子: {self.base_scales}")  # 添加调试信息

        except Exception as e:
            print(f"处理V键事件时发生错误: {e}")


    def on_n_pressed(self):
        """处理N键事件"""
        print("按下了n键")

        try:
            # 设置放大因子为1.005
            zoom_step = 1.005
            # 更新基准尺寸信息,并限制缩放范围（0.08<0.5*0.16>~130<800*0.16>） 使用self._scales_min限制是由于适应不同尺寸的图片
            self.base_scales = [max(0.5*self._scales_min[i], min(scale * zoom_step, 800*self._scales_min[i])) for i, scale in enumerate(self.base_scales)]

            # 应用缩放到所有视图
            for i, view in enumerate(self.graphics_views):
                if view:
                    self._apply_scale_to_view(view, self.base_scales[i])
            
            # print(f"新的缩放因子: {self.base_scales}")  # 添加调试信息
        except Exception as e:
            print(f"处理N键事件时发生错误: {e}")


    def on_t_pressed(self):
        """处理T键事件"""
        print("按下了t键")

        # 调用截屏工具
        WScreenshot.run()

        if False:
            # 创建并显示自定义对话框,传入图片列表
            dialog = CameraTestDialog(self.images_path_list)

            # 显示对话框  convert_to_dict()   self.exif_texts
            if dialog.exec_() == QDialog.Accepted:
                # 写入问题点到表格中
                dialog.write_data()
                # data = dialog.get_data()
                # print(f"用户输入的文字信息: {data}")
            
            # 无论对话框是接受还是取消，都手动销毁对话框
            dialog.deleteLater()
            dialog = None

    def on_ctrl_t_pressed(self):
        """处理Ctrl+T键事件"""
        print("按下了t键")

        try:
            # 创建并显示自定义对话框,传入图片列表
            dialog = CameraTestDialog(self.images_path_list)

            # 显示对话框  convert_to_dict()   self.exif_texts
            if dialog.exec_() == QDialog.Accepted:
                # 写入问题点到表格中
                dialog.write_data()
                # data = dialog.get_data()
                # print(f"用户输入的文字信息: {data}")
            
            # 无论对话框是接受还是取消，都手动销毁对话框
            dialog.deleteLater()
            dialog = None
        except Exception as e:
            print(f"处理Ctrl+T键事件失败:{e}")

    def roi_stats_checkbox(self):
        try:
            self.stats_visible = not self.stats_visible # 控制显示开关
            for view in self.graphics_views:
                if view:
                    view.set_stats_visibility(self.stats_visible)

        except Exception as e:
            print(f"显示图片统计信息时发生错误: {e}")
        finally:
            # 延时1秒后更新is_updating为False
            QTimer.singleShot(1000, lambda: setattr(self, 'is_updating', False))


    def on_p_pressed(self):  ## 会导致窗口闪退，当前策略是使用标志位self.is_updating锁定界面不退出
        """处理P键事件"""
        self.is_updating = True
        
        if self.ai_tips_flag:
            try:
                # 检查图片数量
                if len(self.images_path_list) != 2:
                    show_message_box("当前AI提示支持两张图比较，请选择两张图片进行比较！", "提示", 1000)
                    # 延时1秒后更新is_updatng为False
                    QTimer.singleShot(1000, lambda: setattr(self, 'is_updating', False))
                    return

                # 调用AI提示函数
                show_message_box("按下了p键,正在发起ai请求...", "提示", 500)
                # 更新底部信息提示栏
                self.label_bottom.setText(f"📢:按下了p键,正在发起ai请求...")

                def run_ai():
                    try:
                        llm = CustomLLM_Siliconflow()
                        tips = """假如你是一名专业的影像画质评测工程师,
                        请从亮度、对比度、清晰度、色调等专业角度比较两张图片差异, 用中文回复,
                        一句话总结概括内容, 不要换行，不要超过100字。"""
                        model = "Pro/OpenGVLab/InternVL2-8B"
                        response = llm(select_model=model, prompt=tips, image_path_list=self.images_path_list)
                        # 使用信号机制更新UI，避免跨线程直接操作UI
                        self.ai_response_signal.emit(response)
                    except Exception as e:
                        print(f"AI请求失败: {e}")
                        self.ai_response_signal.emit("AI请求失败，请检查网络连接或模型配置")
                        # 更新底部信息提示栏
                        self.label_bottom.setText(f"📢:AI请求失败，请检查网络连接或模型配置")

                # 创建并启动子线程
                tcp_thread = threading.Thread(target=run_ai)
                tcp_thread.daemon = True
                tcp_thread.start()

            except Exception as e:
                print(f"处理P键时发生错误: {e}")
                self.is_updating = False
                show_message_box("处理P键时发生错误，请重试", "错误", 1000)
        
        elif self.stats_visible:
            try:
                # 设置 P 键来打开ROI信息统计框
                self.roi_selection_active = not self.roi_selection_active
                for view in self.graphics_views:
                    if view:
                        view.toggle_selection_rect(self.roi_selection_active)

            except Exception as e:
                print(f"显示图片统计信息时发生错误: {e}")
            finally:
                # 延时0.01秒后更新is_updating为False
                QTimer.singleShot(10, lambda: setattr(self, 'is_updating', False))
        else:
            # 没有标志位的情况下按 P键不显示
            show_message_box("请勾选 ROI信息复选框 或者 AI提示看图复选框后, \n按P键发出相应请求!", "提示",1500)
            # 延时0.01秒后更新is_updating为False
            QTimer.singleShot(10, lambda: setattr(self, 'is_updating', False))
            pass

    def on_space_pressed(self):
        """处理看图子界面空格键事件"""
        try:
            # 预检查当前状态
            if self.is_updating:
                print("正在更新图片，请稍后...")
                return
            if not self.parent_window:
                print("未找到父窗口，无法获取下一组图片")
                return
            # 切换图片自动清除ROI信息框
            if self.roi_selection_active: 
                self.roi_selection_active = False

            # 设置更新标志
            self.is_updating = True
            # 开始获取下一组文件
            next_images, next_indexs = self.get_next_images()
            if not next_images:
                raise ValueError(f"无效获下一组文件")
            
            # 获取所有文件的扩展名并去重，判断这一组文件的格式，纯图片，纯视频，图片+视频
            is_video, is_image = False, False
            file_extensions = {os.path.splitext(path)[1].lower() for path in next_images}
            if not file_extensions:
                raise ValueError(f"无效的扩展名")
            for file_extension in list(file_extensions):
                if file_extension.endswith(self.parent_window.VIDEO_FORMATS):
                    is_video = True
                if file_extension.endswith(self.parent_window.IMAGE_FORMATS):
                    is_image = True

            # 根据当前组文件的格式选择调用子界面
            if is_image and not is_video:   # 调用图片显示
                self.set_images(next_images, next_indexs)
            elif is_video and not is_image: # 调用视频显示
                self.parent_window.create_video_player(next_images, next_indexs)   
                raise ValueError(f"看图子界面调用视频子界面，主动抛出异常关闭当前看图子界面")             
            elif is_image and is_video:
                # 提示信息框
                show_message_box("🔉: 这组文件同时包含图片和视频文件，无法调出子界面，返回主界面", "提示",1500)
                # 抛出异常，退出当前子界面
                raise ValueError(f"这组文件同时包含图片和视频文件，无法调出子界面，获取的文件如下：\n{next_images}")
            else:
                # 提示信息框
                show_message_box("🔉: 这组文件没有包含图片和视频文件，无法调出子界面，返回主界面", "提示",1500)
                # 抛出异常，退出当前子界面
                raise ValueError(f"这组文件没有包含图片和视频文件，无法调出子界面，获取的文件如下：\n{next_images}")

        except Exception as e:
            print(f"看图子界面-on_space_pressed()--处理空格键时发生错误: {e}")
            # 退出看图界面
            self.is_updating = False
            self.Escape_close()
            
    def on_b_pressed(self):
        """处理B键事件"""
        try:
            # 预检查当前状态
            if self.is_updating:
                print("正在更新图片，请稍后...")
                return
            if not self.parent_window:
                print("未找到父窗口，无法获取下一组图片")
                return
            # 切换图片自动清除ROI信息框
            if self.roi_selection_active: 
                self.roi_selection_active = False

            # 设置更新标志
            self.is_updating = True
            # 开始获取上一组文件
            prev_images, prev_indexs = self.get_prev_images()
            if not prev_images:
                raise ValueError(f"无法获取上一组图片")
            
            # 获取所有文件的扩展名并去重，判断这一组文件的格式，纯图片，纯视频，图片+视频
            is_image, is_video = False, False
            file_extensions = {os.path.splitext(path)[1].lower() for path in prev_images}
            if not file_extensions:
                raise ValueError(f"无效的扩展名")
            for file_extension in list(file_extensions):
                if file_extension.endswith(self.parent_window.VIDEO_FORMATS):
                    is_video = True
                if file_extension.endswith(self.parent_window.IMAGE_FORMATS):
                    is_image = True

            # 根据当前组文件的格式选择调用子界面
            if is_image and not is_video:   # 调用图片显示
                self.set_images(prev_images, prev_indexs)
            elif is_video and not is_image: # 调用视频显示
                self.parent_window.create_video_player(prev_images, prev_indexs)   
                raise ValueError(f"看图子界面调用视频子界面，主动抛出异常关闭当前看图子界面")             
            elif is_image and is_video:
                # 提示信息框
                show_message_box("🔉: 这组文件同时包含图片和视频文件，无法调出子界面，返回主界面", "提示",1500)
                # 抛出异常，退出当前子界面
                raise ValueError(f"这组文件同时包含图片和视频文件，无法调出子界面，获取的文件如下：\n{prev_images}")
            else:
                # 提示信息框
                show_message_box("🔉: 这组文件没有包含图片和视频文件，无法调出子界面，返回主界面", "提示",1500)
                # 抛出异常，退出当前子界面
                raise ValueError(f"这组文件没有包含图片和视频文件，无法调出子界面，获取的文件如下：\n{prev_images}")

        except Exception as e:
            print(f"处理B键时发生错误: {e}")
            self.is_updating = False
            self.Escape_close()
            
    
    def get_next_images(self):
        """获取下一组图片"""
        try:
            if self.parent_window:
                next_images, next_indexs = self.parent_window.press_space_and_b_get_selected_file_paths('space')
                if next_images and isinstance(next_images, list) and len(next_images) > 0:
                    # print(f"获取到下一组图片: {next_images}")
                    return next_images, next_indexs
                else:
                    print("获取到的下一组图片无效")
                    return None, None
        except Exception as e:
            print(f"获取下一组图片时发生错误: {e}")
            return None, None
        return None, None
    
    def get_prev_images(self):
        """获取上一组图片"""
        try:
            if self.parent_window:
                prev_images, prev_indexs = self.parent_window.press_space_and_b_get_selected_file_paths('b')
                if prev_images and isinstance(prev_images, list) and len(prev_images) > 0:
                    # print(f"获取到上一组图片: {prev_images}")
                    return prev_images, prev_indexs
                else:
                    print("获取到的上一组图片无效")
                    return None, None
        except Exception as e:
            print(f"获取上一组图片时发生错误: {e}")
            return None, None
        return None, None

    def closeEvent(self, event):
        """重写关闭事件处理"""
        # 检查是否应该忽略关闭事件
        print("closeEvent()-看图子界面--子窗口关闭事件:")
        if self.is_updating:
            print("看图子界面正在更新图片，忽略关闭事件")
            event.ignore()
            return

        try:
            self.save_settings()  # 保存颜色以及EXIF信息设置
            self.cleanup()        # 清理资源
            self.closed.emit()    # 发送关闭信号
            event.accept()
            print("接受看图子界面关闭事件, 并保存颜色以及EXIF信息设置")
        except Exception as e:
            print(f"closeEvent()-看图子界面--关闭窗口时发生错误: {e}")
            event.accept()

        """# 修复关闭事件, 可以正常关闭, modify by diamond_cz 2025-01-14
        if self.should_close:
            try:
                self.cleanup()  # 清理资源
                self.closed.emit()  # 发送关闭信号
                event.accept()
                print("子窗口关闭事件,接受")
            except Exception as e:
                print(f"关闭窗口时发生错误: {e}")
                event.accept()
        else:
            # 如果不是预期的关闭（比如快速按键导致的），则忽略关闭事件
            event.ignore()
            print("子窗口关闭事件,忽略")
        """

    def Escape_close(self):
        """统一处理窗口关闭逻辑"""
        if self.is_updating:
            print("正在更新图片，忽略关闭请求")
            return
            
        try:
            # 保存颜色设置
            self.save_settings()
            self.cleanup()      # 清理资源
            self.closed.emit()  # 发送关闭信号
            super().close()     # 调用父类的close方法
        except Exception as e:
            print(f"关闭窗口时发生错误: {e}")
            super().close()


    def load_settings(self):
        """加载颜色、exif等设置"""
        try:
            # 加载保存的颜色设置 basic_color_settings、basic_color_settings
            ColorSettings = load_color_settings()
            color_settings = ColorSettings.get("basic_color_settings",{})
            rgb_settings = ColorSettings.get("rgb_color_settings",{})
            # 加载保存的EXIF信息设置
            ExifSettings = load_exif_settings()
            label_visable = ExifSettings.get("label_visable_settings",{})   # label具体项显示的标志位
            exif_visable = ExifSettings.get("exif_visable_setting",{})      # exif具体项显示的标志位


            # 从颜色配置中读取基础颜色
            self.font_color_default = color_settings.get("font_color_default", "rgb(0, 0, 0)")                  # 默认字体颜色_纯黑色
            self.font_color_exif = color_settings.get("font_color_exif", "rgb(255, 255, 255)")                  # Exif字体颜色_纯白色
            self.background_color_default = color_settings.get("background_color_default", "rgb(173,216,230)")  # 深色背景色_好蓝
            self.background_color_table = color_settings.get("background_color_table", "rgb(127, 127, 127)")    # 表格背景色_18度灰

            # 读取rgb颜色配置
            self.color_rgb_settings = rgb_settings

            # 初始化exif信息可见性字典，支持用户在json配置文件中调整顺序以及是否显示该项
            self.dict_exif_info_visibility = exif_visable

            # 初始化label显示变量
            self.dict_label_info_visibility = label_visable
            # 读取图像颜色空间显示设置
            self.p3_color_space = self.dict_label_info_visibility.get("p3_color_space", False)     
            self.gray_color_space = self.dict_label_info_visibility.get("gray_color_space", False) 
            self.srgb_color_space = self.dict_label_info_visibility.get("srgb_color_space", True) 
            # 设置亮度统计信息的标志位；初始化ai提示标注位为False 
            self.stats_visible = self.dict_label_info_visibility.get("roi_info", False)         
            self.ai_tips_flag = self.dict_label_info_visibility.get("ai_tips", False)          

        except Exception as e:
            print(f"self.load_settings()--加载设置失败: {e}")
        

    def save_settings(self):
        """保存颜色设置"""
        try:
            # 确保config目录存在
            config_dir = pathlib.Path("./config")
            config_dir.mkdir(parents=True, exist_ok=True)

            # 1. 保存颜色配置文件
            settings_color_file = config_dir / "color_setting.json"
            basic_color_settings ={
                "background_color_default": self.background_color_default,
                "background_color_table": self.background_color_table,
                "font_color_default": self.font_color_default,
                "font_color_exif": self.font_color_exif,
            }
            # rgb_color_settings = {
            #     "18度灰": "rgb(127,127,127)",
            #     "石榴红": "rgb(242,12,0)",
            #     "乌漆嘛黑": "rgb(22, 24, 35)",
            #     "铅白": "rgb(240,240,244)", 
            #     "水色": "rgb(136,173,166)",   
            #     "石青": "rgb(123,207,166)",           
            #     "茶色": "rgb(242,12,0)",
            #     "天际": "rgb(236,237,236)",   
            #     "晴空": "rgb(234,243,244)",  
            #     "苍穹": "rgb(220,230,247)", 
            #     "湖光": "rgb(74,116,171)", 
            #     "曜石": "rgb(84, 99,125)", 
            #     "天际黑": "rgb(8,8,6)",   
            #     "晴空黑": "rgb(45,53,60)",  
            #     "苍穹黑": "rgb(47,51,68)", 
            #     "湖光黑": "rgb(49,69,96)", 
            #     "曜石黑": "rgb(57,63,78)", 
            # }
            setting = {
                "basic_color_settings": basic_color_settings,
                "rgb_color_settings": self.color_rgb_settings   # 默认使用配置中读取的配置
            }
            # 保存setting到配置文件config_dir / "color_setting.json"
            with open(settings_color_file, 'w', encoding='utf-8', errors='ignore') as f:
                json.dump(setting, f, indent=4, ensure_ascii=False)

            # 2. 保存exif配置文件
            settings_exif_file = config_dir / "exif_setting.json"
            label_visable_settings = {
                "histogram_info": self.checkBox_1.isChecked(),
                "exif_info": self.checkBox_2.isChecked(),
                "roi_info": self.checkBox_3.isChecked(),
                "ai_tips": self.checkBox_4.isChecked(),
                "srgb_color_space":self.srgb_color_space,
                "p3_color_space":self.p3_color_space,
                "gray_color_space":self.gray_color_space,
            }
            # exif_visable_setting = {
            #     '图片名称' : True,
            #     '品牌' : True,
            #     '型号' : True,
            #     '图片大小' : True,
            #     '图片尺寸' : True,
            #     '图片张数' : True,
            #     '曝光时间' : True,
            #     '光圈值' : True,
            #     'ISO值' : True,
            #     '原始时间' : True,
            #     '测光模式' : True,
            #     'HDR' : True,
            #     'Zoom' : True,
            #     'Lux' : True,
            #     'CCT' : True,
            #     'FaceSA' : True,
            #     'DRCgain' : True,
            #     'Awb_sa' : True,
            #     'Triangle_index' : True,
            #     'R_gain' : True,
            #     'B_gain' : True,
            #     'Safe_gain' : True,
            #     'Short_gain' : True,
            #     'Long_gain' : True
            # }
            setting = {
                "label_visable_settings": label_visable_settings,
                "exif_visable_setting": self.dict_exif_info_visibility  # 默认使用配置中读取的配置
            }
            # 保存setting到配置文件config_dir / "exif_setting.json"
            with open(settings_exif_file, 'w', encoding='utf-8', errors='ignore') as f:
                json.dump(setting, f, indent=4, ensure_ascii=False)

        except Exception as e:
            print(f"save_settings()-看图子界面--保存设置失败: {e}")



    def update_ai_response(self, response):
        """更新AI响应结果"""
        self.label_bottom.setText(f"📢:AI提示结果:{response}")
        # 延时1秒后更新is_updating为False
        QTimer.singleShot(1000, lambda: setattr(self, 'is_updating', False))



"""
设置看图界面类区域结束线
------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
"""


"""
设置看图界面功能测试区域开始线
------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
"""

one_pic = ['D:/Tuning/M5151/0_picture/20241209_FT/1209-C3N后置GL四供第三轮FT（小米园区）/photo\\四供\\四供_IMG_20241209_081113.jpg']
two_pic = ['D:/Tuning/M5151/0_picture/20241209_FT/1209-C3N后置GL四供第三轮FT（小米园区）/photo\\四供\\四供_IMG_20241209_081126.jpg', 'D:/Tuning/M5151/0_picture/20241209_FT/1209-C3N后置GL四供第三轮FT（小米园区）/photo\\四供\\四供_IMG_20241209_081300.jpg']
three_pic = ['D:/Tuning/M5151/0_picture/20241209_FT/1209-C3N后置GL四供第三轮FT（小米园区）/photo\\四供\\四供_IMG_20241209_081113.jpg', 'D:/Tuning/M5151/0_picture/20241209_FT/1209-C3N后置GL四供第三轮FT（小米园区）/photo\\四供\\四供_IMG_20241209_081124.jpg', 'D:/Tuning/M5151/0_picture/20241209_FT/1209-C3N后置GL四供第三轮FT（小米园区）/photo\\四供\\四供_IMG_20241209_081126.jpg']
two_pics = ['D:/Tuning/M5151/0_picture/20241224_FT/1224_第四轮FT问题截图（米家+小米园区）\\Photo\\10、小米之家 ISO 53，AWB 略偏黄绿.png', 'D:/Tuning/M5151/0_picture/20241224_FT/1224_第四轮FT问题截图（米家+小米园区）\\Photo\\111、小米园区 ISO 50,AWB 偏黄，ISP 涂抹重，细节清晰度不佳.png']
two_index = ['1/1', '1/1']



if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = SubMainWindow(two_pic, two_index)
    # 不需要额外的 show() 调用，因为在初始化时已经调用了 showMaximized()
    sys.exit(app.exec_())

"""
设置看图界面功能测试区域结束线
------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
"""