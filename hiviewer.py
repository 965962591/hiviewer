"""导入python内置模块"""
import re
import gc
import os
import sys
import time
import json
import shutil
import ctypes
import zipfile
import pathlib
import hashlib
import logging
import threading
import subprocess
from io import BytesIO
from queue import Queue
from pathlib import Path
from functools import lru_cache
from fractions import Fraction
from typing import Optional, Tuple
from itertools import zip_longest, chain
from logging.handlers import RotatingFileHandler

"""导入python第三方模块"""
# import av # 比原生 OpenCV 快 35%（实测 1000 个视频处理仅需 8.2 秒）  
import cv2
import piexif
from openpyxl import Workbook
import xml.etree.ElementTree as ET
from PIL import Image
from PyQt5 import QtWidgets, QtGui, QtCore
from PyQt5.QtGui import (
    QIcon, QKeySequence, QPixmap, QColor, QTransform, QFont, QPainter, QImageReader,QImage)
from PyQt5.QtWidgets import (
    QFileSystemModel, QAbstractItemView, QTableWidgetItem, QHeaderView, QShortcut, QSplashScreen, 
    QMessageBox, QStyledItemDelegate, QStyleOptionButton, QStyle, QApplication, QMenu, QProgressBar,
    QProgressDialog, QDialog, QVBoxLayout, QLabel, QHBoxLayout, QPushButton, QLineEdit, QCheckBox)
from PyQt5.QtCore import (
    Qt, QDir, QTimer, QSize, QTimer, QRunnable, QThreadPool, QObject, pyqtSignal, QAbstractListModel,
    QThread, QSize, QAbstractListModel, QModelIndex, QVariant, QItemSelection, QItemSelectionModel)

"""导入用户自定义的模块"""
from src.ui.main_ui import Ui_MainWindow                        # 假设你的主窗口类名为Ui_MainWindow
from src.modules.sub_compare_image_view import SubMainWindow    # 假设这是你的子窗口类名
from src.modules.sub_compare_video_view import VideoWall        # 假设这是你的子窗口类名 
from src.modules.sub_rename_view import FileOrganizer           # 添加这行以导入批量重名名类名
from src.modules.sub_image_process import SubCompare            # 确保导入 SubCompare 类
from src.modules.sub_bat_view import LogVerboseMaskApp          # 导入批量执行命令的类
from src.utils.about import AboutDialog                         # 导入关于对话框类,显示帮助信息
from src.utils.hisnot import WScreenshot                        # 导入截图工具类
from src.utils.raw2jpg import Mipi2RawConverterApp              # 导入MIPI RAW文件转换为JPG文件的类
from src.utils.dialog_class import Qualcom_Dialog               # 导入自定义对话框的类
from src.utils.font_class import SingleFontManager, MultiFontManager  # 字体管理器
from src.utils.update import check_update,pre_check_update            # 导入自动更新检查程序


"""python项目多文件夹路径说明

(1)获取当前py文件的路径: os.path.abspath(__file__)
(2)获取当前py文件的父文件夹路径: os.path.dirname(os.path.abspath(__file__))

(1)获取主函数py文件的路径: os.path.abspath(sys.argv[0])
(2)获取主函数py文件的父文件夹路径: os.path.dirname(os.path.abspath(sys.argv[0]))

"""


"""
设置全局变量以及全局函数区域开始线
------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
"""

# 预编译正则表达式，提高效率（针对实现类似widow的文件排名）
_natural_sort_re = re.compile('([0-9]+)')

def natural_sort_key(s):
    """将字符串转换为自然排序的键值（优化版）"""
    return [int(text) if text.isdigit() else text.lower() for text in _natural_sort_re.split(s)]

def show_message_box(text, title="提示", timeout=None):
    """显示消息框，宽度自适应文本内容
    
    Args:
        text: 显示的文本内容
        title: 窗口标题，默认为"提示" 
        timeout: 自动关闭的超时时间(毫秒)，默认为None不自动关闭
    """
    msg_box = QMessageBox()
    msg_box.setWindowTitle(title)
    msg_box.setText(text)
    msg_box.setStandardButtons(QMessageBox.Ok)
    
    if timeout is not None:
        # 设置定时器自动关闭
        QTimer.singleShot(timeout, msg_box.close)
    
    msg_box.exec_() # 使用 exec_ 显示模态对话框


def version_init(default_version_path, VERSION=str):
    """ 全局函数，版本号初始化 """
    # 检查文件是否存在，如果不存在则创建并写入默认版本号
    try:
        if not os.path.exists(default_version_path):
            with open(default_version_path, 'w') as f:
                f.write(VERSION) 
            return VERSION
        else:
            with open(default_version_path, 'r', encoding='utf-8') as f:
                return f.read().strip()
    except Exception as e:
        print(f"version_init()--版本号初始化失败: {e}")
        return VERSION

def load_color_settings():
    """加载颜色设置"""
    try:
        # 确保cache目录存在
        cache_dir = pathlib.Path("./cache")
        cache_dir.mkdir(parents=True, exist_ok=True)
        
        settings_file = cache_dir / "color_setting.json"
        if settings_file.exists():
            with open(settings_file, 'r', encoding='utf-8', errors='ignore') as f:
                return json.load(f)
        else: #设置默认颜色设置
            try:
                print(f"颜色设置文件不存在: {settings_file}, 设置默认颜色设置")
                
                settings = {
                    "background_color_default": "rgb(173,216,230)",  # 深色背景色_好蓝
                    "background_color_table": "rgb(127, 127, 127)",   # 表格背景色_18度灰
                    "font_color_default": "rgb(0, 0, 0)",         # 默认字体颜色_纯黑色
                    "font_color_exif": "rgb(255, 255, 255)"        # Exif字体颜色_纯白色
                }
                
                with open(settings_file, 'w', encoding='utf-8', errors='ignore') as f:
                    json.dump(settings, f, indent=4, ensure_ascii=False)
                
            except Exception as e:
                print(f"默认颜色设置失败: {e}")

    except Exception as e:
        print(f"加载颜色设置失败: {e}")
    return {}

def rgb_str_to_qcolor(rgb_str):
    """将 'rgb(r,g,b)' 格式的字符串转换为 QColor"""
    # 提取RGB值
    rgb = rgb_str.strip('rgb()')  # 移除 'rgb()' 
    r, g, b = map(int, rgb.split(','))  # 分割并转换为整数
    return QColor(r, g, b)

def force_delete_file(file_path):
    """强制删除文件"""
    try:
        os.remove(file_path)
    except PermissionError:
        # 如果文件被占用，尝试强制删除
        try:
            # 使用Windows API强制删除文件
            ctypes.windll.kernel32.DeleteFileW(file_path)
        except Exception as e:
            print(f"强制删除文件失败: {e}")

def force_delete_folder(folder_path, suffix='.zip'):
    """强制删除文件夹内指定后缀文件"""
    try:
        for file in os.listdir(folder_path):
            if file.endswith(suffix):
                force_delete_file(os.path.join(folder_path, file))
    except Exception as e:
        print(f"强制删除文件夹失败: {e}")   


def load_xml_data(xml_path):
    """加载XML文件并提取Lux值和DRCgain值等EXIF信息"""
    try:
        # 加载xml文件
        tree = ET.parse(xml_path)
        root = tree.getroot()
        
        """
        # 提取值并转换为字典
        result_dict = {
            "Lux": root.find('lux_index').text if root.find('lux_index') is not None else None,
            "DRCgain": root.find('DRCgain').text if root.find('DRCgain') is not None else None,
            "Safe_gain": root.find('safe_gain').text if root.find('safe_gain') is not None else None,
            "Short_gain": root.find('short_gain').text if root.find('short_gain') is not None else None,
            "Long_gain": root.find('long_gain').text if root.find('long_gain') is not None else None,
            "CCT": root.find('CCT').text if root.find('CCT') is not None else None,
            "R_gain": root.find('r_gain').text if root.find('r_gain') is not None else None,
            "B_gain": root.find('b_gain').text if root.find('b_gain') is not None else None,
            "Awb_sa": root.find('awb_sa').text if root.find('awb_sa') is not None else None,
            "Triangle_index": root.find('triangle_index').text if root.find('triangle_index') is not None else None,
        }

        result_one_list = [
            "文件名",
            "Lux",
            "DRCgain",
            "Safe_gain",
            "Short_gain",
            "Long_gain",
            "CCT",
            "R_gain",
            "B_gain",
            "Awb_sa",
            "Triangle_index",
        ]

        """

        # 提取值并转换为列表
        result_list = [
            # str(Path(xml_path).parent / (os.path.basename(xml_path).split('_new.xml')[0]+".jpg")),
            str(os.path.basename(xml_path).split('_new.xml')[0]+".jpg"),
            float(root.find('lux_index').text) if root.find('lux_index') is not None else None,
            root.find('DRCgain').text if root.find('DRCgain') is not None else None,
            float(root.find('safe_gain').text) if root.find('safe_gain') is not None else None,
            float(root.find('short_gain').text) if root.find('short_gain') is not None else None,
            float(root.find('long_gain').text) if root.find('long_gain') is not None else None,
            float(root.find('CCT').text) if root.find('CCT') is not None else None,
            float(root.find('r_gain').text) if root.find('r_gain') is not None else None,
            float(root.find('b_gain').text) if root.find('b_gain') is not None else None,
            root.find('awb_sa').text if root.find('awb_sa') is not None else None,
            float(root.find('triangle_index').text) if root.find('triangle_index') is not None else None,
        ]

        return result_list
        
    except Exception as e:
        print(f"解析XML失败{xml_path}:\n {e}")
        return None



def save_excel_data(images_path):
    """将从XML文件中提取的数据保存到Excel表格中"""

    # 配置保存的Excel文件路径
    excel_path = os.path.join(images_path, "extracted_data.xlsx")
    if os.path.exists(excel_path):
        # 若存在excel文件则不需要保存新的excel文件
        return

    # 初始化二维列表
    get_excel_list = [
        [
            "文件名",
            "Lux",
            "DRCgain",
            "Safe_gain",
            "Short_gain",
            "Long_gain",
            "CCT",
            "R_gain",
            "B_gain",
            "Awb_sa",
            "Triangle_index",
            "AE",
            "AWB",
            "ISP",
            "AF",
        ]
    ]

    # 遍历文件夹，列出所有满足条件的xml文件
    xml_files = [f for f in os.listdir(images_path) if f.endswith('_new.xml')]

    for xml_file in xml_files:
        xml_file = str(Path(images_path) / xml_file)
        if os.path.exists(xml_file):
            # 使用函数load_xml_data加载并解析xml文件
            result_list = load_xml_data(xml_file)
            if result_list:
                get_excel_list.append(result_list)

    # 创建一个新的工作簿
    wb = Workbook()
    ws = wb.active

    # 设置边框样式
    from openpyxl.styles import Border, Side
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 写入数据并设置列宽和边框
    for row in get_excel_list:
        ws.append(row)
        for cell in row:
            # 设置边框
            ws.cell(row=ws.max_row, column=row.index(cell) + 1).border = border

    # 自适应列宽
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter  # 获取列字母
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)  # 加2以增加一些额外的空间
        ws.column_dimensions[column_letter].width = adjusted_width


    wb.save(excel_path)  # 保存为Excel文件
    print(f"数据已保存到 {excel_path}")

"""
设置全局函数区域结束线
------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
"""



"""
设置独立封装类区域开始线
------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
"""
class WorkerSignals(QObject):
    """工作线程信号"""
    finished = pyqtSignal()  # 完成信号
    progress = pyqtSignal(int, int)  # 进度信号 (当前, 总数)
    error = pyqtSignal(str)  # 错误信号
    batch_loaded = pyqtSignal(list)  # 批量加载完成信号


class CommandThread(QThread):
    """执行高通图片解析工具独立线程类"""
    finished = pyqtSignal(bool, str, str)  # 添加 images_path 参数

    def __init__(self, command, images_path):
        super().__init__()
        self.command = command
        self.images_path = images_path

    def run(self):
        try:
            # result = subprocess.run(
            #     self.command, 
            #     check=True, 
            #     stdout=subprocess.PIPE, 
            #     stderr=subprocess.PIPE, 
            #     text=True, 
            #     encoding='utf-8')
            # self.finished.emit(result.returncode == 0, result.stderr, self.images_path)  # 发射信号，传递结果
            
            # 使用 /c 参数，命令执行完成后关闭窗口，直接独立线程
            result = subprocess.run(
                f'start /wait cmd /c {self.command}',  # /wait 等待新窗口关闭
                shell=True,
                stdout=subprocess.PIPE,  # 捕获标准输出
                stderr=subprocess.PIPE,  # 捕获标准错误
                text=True  # 将输出解码为字符串
            )
            
            # 发射信号，传递结果
            self.finished.emit(result.returncode == 0, result.stderr, self.images_path)
            
        except Exception as e:
            self.finished.emit(False, str(e), self.images_path)  # 发射信号，传递错误信息


class CompressWorker(QRunnable):
    """压缩工作线程类"""
    class Signals(QObject):
        """压缩工作线程信号"""
        progress = pyqtSignal(int, int)  # 当前进度,总数
        finished = pyqtSignal(str)  # 完成信号,返回压缩包路径
        error = pyqtSignal(str)  # 错误信号
        cancel = pyqtSignal()  # 取消信号
        
    def __init__(self, files_to_compress, zip_path):
        super().__init__()
        self.files = files_to_compress
        self.zip_path = zip_path
        self.signals = self.Signals()
        self._stop = False
        
    def run(self):
        try:
            with zipfile.ZipFile(self.zip_path, 'w') as zip_file:
                for i, (file_path, arcname) in enumerate(self.files):
                    if self._stop:
                        return
                    
                    try:
                        zip_file.write(file_path, arcname)
                        self.signals.progress.emit(i + 1, len(self.files))
                    except Exception as e:
                        self.signals.error.emit(f"压缩文件失败: {file_path}, 错误: {e}")
                        continue
                    
            self.signals.finished.emit(self.zip_path)
            
        except Exception as e:
            self.signals.error.emit(f"创建压缩包失败: {e}")
        
    def cancel(self):
        """取消压缩任务"""
        self._stop = True  # 设置停止标志


# 更新 ProgressDialog 类以添加取消按钮
class ProgressDialog(QtWidgets.QDialog):
    """压缩进度对话框"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("压缩进度")
        self.setModal(True)

        # 使用无边框窗口风格
        self.setWindowFlags(QtCore.Qt.Dialog | QtCore.Qt.WindowTitleHint | QtCore.Qt.WindowCloseButtonHint | QtCore.Qt.FramelessWindowHint)

        self.layout = QtWidgets.QVBoxLayout(self)
        self.progress_bar = QtWidgets.QProgressBar(self)
        self.message_label = QtWidgets.QLabel(self)  # 新增 QLabel 用于显示消息
        self.cancel_button = QtWidgets.QPushButton("取消", self)

        # 添加进度条、消息标签和取消按钮到布局
        self.layout.addWidget(self.progress_bar)
        self.layout.addWidget(self.message_label)  # 添加消息标签到布局
        self.layout.addWidget(self.cancel_button)

        # 设置窗口大小
        self.setFixedSize(450, 150)

        self.cancel_button.clicked.connect(self.cancel_compression)

        # 设置窗口位置为当前鼠标所在显示屏的中央
        self.center_on_current_screen()


    def center_on_current_screen(self):
        # 获取当前鼠标位置和显示屏
        cursor_pos = QtGui.QCursor.pos()  # 从QtCore导入QCursor
        screen = QtWidgets.QApplication.desktop().screenNumber(cursor_pos)

        # 获取该显示屏的矩形区域
        screen_geometry = QtWidgets.QApplication.desktop().screenGeometry(screen)

        # 计算中央位置
        center_x = screen_geometry.x() + (screen_geometry.width() - self.width()) // 2
        center_y = screen_geometry.y() + (screen_geometry.height() - self.height()) // 2

        # 设置窗口位置
        self.move(center_x, center_y)

    def update_progress(self, value):
        self.progress_bar.setValue(value)

    def set_message(self, message):
        self.message_label.setText(message)  # 更新 QLabel 内容

    def cancel_compression(self):
        # 发送取消信号
        self.parent().cancel_compression()
        self.close()

class ImagePreloader(QRunnable):
    """改进的图片预加载工作线程"""
    def __init__(self, file_paths):
        super().__init__()
        self.file_paths = file_paths
        self.signals = WorkerSignals()
        self._pause = False
        self._stop = False
        self._pause_condition = threading.Event()
        self._pause_condition.set()  # 初始状态为未暂停
        
    def pause(self):
        """暂停预加载"""
        self._pause = True
        self._pause_condition.clear()

    def resume(self):
        """恢复预加载"""
        self._pause = False
        self._pause_condition.set()
        
    def run(self):
        try:
            total = len(self.file_paths)
            batch = []
            batch_size = 10
            
            for i, file_path in enumerate(self.file_paths):
                if self._stop:
                    break
                    
                # 使用 Event 来实现暂停
                self._pause_condition.wait()
                    
                if file_path:
                    icon = IconCache.get_icon(file_path)  # 使用缓存系统获取图标
                    batch.append((file_path, icon))
                    
                    if len(batch) >= batch_size:
                        self.signals.batch_loaded.emit(batch)
                        batch = []
                        
                    self.signals.progress.emit(i + 1, total)
                    
            if batch:  # 发送最后的批次
                self.signals.batch_loaded.emit(batch)
                
            self.signals.finished.emit()
            
        except Exception as e:
            self.signals.error.emit(str(e))


class IconCache:
    """图标缓存类"""
    _cache = {}
    _cache_dir = os.path.join(os.path.dirname(__file__), "cache", "icons")
    _cache_index_file = os.path.join(_cache_dir, "icon_index.json")
    _max_cache_size = 1000  # 最大缓存数量，超过会删除最旧的缓存
    # 视频文件格式
    VIDEO_FORMATS = ('.mp4', '.avi', '.mov', '.wmv', '.mpeg', '.mpg', '.mkv')
    # 图片文件格式
    IMAGE_FORMATS = ('.png', '.jpg', '.jpeg', '.bmp', '.gif', '.tif', '.ico', '.webp')
    # 其他文件格式类型图标维护
    FILE_TYPE_ICONS = {
    '.txt': "text_icon.png",
    '.py': "default_py_icon.png",
    }

    @classmethod
    @lru_cache(maxsize=_max_cache_size) # 使用LRU缓存策略替代简单的时间戳排序,该策略如何手动清除缓存？使用cls.get_icon.cache_clear()
    def get_icon(cls, file_path):
        """获取图标，优先从缓存获取"""
        try:
            # 检查内存缓存
            if file_path in cls._cache:
                # print("获取图标, 进入读缓存")
                return cls._cache[file_path]

            # 检查文件缓存
            cache_path = cls._get_cache_path(file_path)
            if os.path.exists(cache_path):
                # print("获取图标, 进入文件缓存")
                icon = QIcon(cache_path)
                cls._cache[file_path] = icon
                return icon

            # 生成新图标 
            icon = cls._generate_icon(file_path)
            if icon:
                # print("获取图标, 进入生成新图标")
                cls._cache[file_path] = icon
                cls._save_to_cache(file_path, icon)
            return icon

        except Exception as e:
            print(f"获取图标失败: {e}")
            # show_message_box(f"get_icon获取图标失败: {e}", "提示", 1500)
            return QIcon()

    # 在_get_cache_path方法中添加文件修改时间校验
    @classmethod
    def _get_cache_path(cls, file_path):
        file_stat = os.stat(file_path)
        file_hash = hashlib.md5(f"{file_path}-{file_stat.st_mtime}".encode()).hexdigest()
        return os.path.join(cls._cache_dir, f"{file_hash}.png")

    @classmethod
    def _generate_icon(cls, file_path):
        """生成图标，确保正确处理图片旋转和视频缩略图
        
        Args:
            file_path: 文件路径
            
        Returns:
            QIcon: 处理后的图标对象
        """
        try:
            # 获取文件类型，提取后缀
            file_ext = os.path.splitext(file_path)[1].lower()
            
            # 视频文件处理
            if file_ext in cls.VIDEO_FORMATS:
                return cls.get_video_thumbnail(file_path)
            
            # 图片文件处理
            elif file_ext in cls.IMAGE_FORMATS:
                return cls._generate_image_icon(file_path)
    
            # 其它文件类型
            else:
                """特殊文件类型处理"""
                return cls.get_default_icon(cls.FILE_TYPE_ICONS.get(file_ext, "default_icon.png"), (48, 48))

        except Exception as e:
            print(f"生成图标失败: {e}")
            # show_message_box(f"_generate_icon生成图标失败: {e}", "提示", 1500)
            return cls.get_default_icon("default_icon.png", (48, 48))


    @classmethod
    def _generate_image_icon(cls, file_path):
        """优化后的图片图标生成"""
        try:

            if False:
                # 方案一：考虑到图片EXIF原始信息标记旋转的方案。现弃用，使用不考虑旋转信息的高效方案二
                with Image.open(file_path) as img:
                    image_format  = img.format

                # 单独加载TIFF格式图片, 处理该格式文件会增加程序耗时
                if image_format == "TIFF":
                    return cls._generate_fallback_icon(file_path)  
                            
                # 使用QImageReader进行高效加载
                reader = QImageReader(file_path)
                reader.setScaledSize(QSize(48, 48))
                reader.setAutoTransform(True)
                image = reader.read()
                pixmap = QPixmap.fromImage(image)

            if True:
                # 方案二：不考虑旋转信息,使用QImage直接加载图像
                image = QImage(file_path)
                if image.isNull():
                    raise ValueError("无法加载图像")

                # 缩放图像
                pixmap = QPixmap.fromImage(image.scaled(48, 48, Qt.KeepAspectRatio, Qt.SmoothTransformation))
           
            return QIcon(pixmap)
            
        except Exception as e:
            print(f"高效生成图标失败: {e}")
            # show_message_box(f"_generate_image_icon生成图标失败: {e}", "提示", 1500)
            return cls.get_default_icon("image_icon.png", (48, 48))

    @classmethod
    def _generate_fallback_icon(cls, file_path):
        """备用的高质量生成方式"""
        try:
            # 使用PIL进行渐进式加载
            with Image.open(file_path) as img:
                img.thumbnail((48, 48))
                buffer = BytesIO()
                img.save(buffer, format='PNG')
                
                # 转换为QPixmap并缩放
                pixmap = QPixmap()
                pixmap.loadFromData(buffer.getvalue())
                return QIcon(pixmap)
        except Exception as e:
            print(f"备用图标生成失败: {e}")
            # show_message_box(f"_generate_fallback_icon图标生成失败: {e}", "提示", 1500)
            return cls.get_default_icon("default_icon.png", (48, 48))


    @classmethod
    def get_default_icon(cls, icon_path: str, icon_size: Optional[Tuple[int, int]] = None) -> QIcon:
        """获取默认文件图标
        
        Args:
            icon_path: 图标文件路径
            icon_size: 可选的图标尺寸元组 (width, height)
            
        Returns:
            QIcon: 处理后的图标对象
        """
        try:
            # 构建默认图标路径
            default_icon_path = os.path.join(os.path.dirname(__file__), "icons", icon_path)
            
            # 检查图标文件是否存在
            if os.path.exists(default_icon_path):
                try:
                    cls._default_icon = QIcon(default_icon_path)
                    if cls._default_icon.isNull():
                        raise ValueError("无法加载图标文件")
                except Exception as e:
                    print(f"加载图标文件失败: {str(e)}")
                    # 创建备用图标
                    cls._default_icon = cls._create_fallback_icon()
            else:
                print(f"图标文件不存在: {default_icon_path}")
                cls._default_icon = cls._create_fallback_icon()
                
            # 处理图标尺寸
            if icon_size:
                try:
                    pixmap = cls._default_icon.pixmap(QSize(*icon_size))
                    if pixmap.isNull():
                        raise ValueError("调整图标尺寸失败")
                    return QIcon(pixmap)
                except Exception as e:
                    print(f"调整图标尺寸失败: {str(e)}")
                    return cls._default_icon
                    
            return cls._default_icon
            
        except Exception as e:
            print(f"获取默认图标时发生错误: {str(e)}")
            return cls._create_fallback_icon()

    @classmethod
    def _create_fallback_icon(cls) -> QIcon:
        """创建备用图标
        
        Returns:
            QIcon: 灰色背景的备用图标
        """
        try:
            pixmap = QPixmap(48, 48)
            pixmap.fill(Qt.gray)
            return QIcon(pixmap)
        except Exception as e:
            print(f"创建备用图标失败: {str(e)}")
            # 返回空图标作为最后的备选方案
            return QIcon()

    @classmethod
    def get_video_thumbnail(cls, video_path: str, size: Tuple[int, int] = (48, 48)):
        """获取视频的第一帧作为缩略图
        
        Args:
            video_path: 视频文件路径
            size: 缩略图大小
            
        Returns:
            QPixmap: 视频缩略图，失败返回None
        """
        cap = None
        try:
            # 尝试打开视频文件
            cap = cv2.VideoCapture(video_path)
            if not cap.isOpened():
                return None
                
            # 读取第一帧
            # 设置超时机制
            start_time = time.time()
            while time.time() - start_time < 2:  # 最多等待2秒
                ret, frame = cap.read()
                if ret:
                    break
            cap.release()
            
            if not ret:
                return None
                
            # 转换颜色空间从 BGR 到 RGB
            frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            
            # 创建 QImage
            height, width, channel = frame.shape
            bytes_per_line = channel * width
            q_img = QtGui.QImage(frame.data, width, height, bytes_per_line, QtGui.QImage.Format_RGB888)
            
            # 创建并缩放 QPixmap
            pixmap = QPixmap.fromImage(q_img)
            pixmap = pixmap.scaled(size[0], size[1], Qt.KeepAspectRatio, Qt.SmoothTransformation)
            
            return QIcon(pixmap)
            
        except Exception as e:
            print(f"获取视频缩略图失败 {video_path}: {str(e)}")
            # show_message_box(f"get_video_thumbnails获取视频缩略图失败{os.path.basename(video_path)}: {str(e)}", "提示", 1500)
            return cls.get_default_icon("video_icon.png", (48, 48))
        finally:
            # 确保资源释放
            if cap is not None:
                cap.release()  

    @classmethod
    def _save_to_cache(cls, file_path, icon):
        """保存图标到缓存"""
        try:
            # 确保缓存目录存在
            os.makedirs(cls._cache_dir, exist_ok=True)

            # 保存图标
            cache_path = cls._get_cache_path(file_path)
            icon.pixmap(48, 48).save(cache_path, "PNG")

            # 更新索引文件
            cls._update_cache_index(file_path)

        except Exception as e:
            print(f"保存图标缓存失败: {e}")
            # show_message_box(f"_save_to_cache保存图标缓存失败: {e}", "提示", 1500)
            
    @classmethod
    def _update_cache_index(cls, file_path):
        """更新缓存索引"""
        try:
            # 读取现有索引
            index = {}
            if os.path.exists(cls._cache_index_file):
                with open(cls._cache_index_file, 'r', encoding='utf-8', errors='ignore') as f:
                    index = json.load(f)

            # 更新索引
            cache_path = cls._get_cache_path(file_path)
            index[cache_path] = {
                'original_path': file_path,
                'timestamp': time.time()
            }

            # 检查缓存大小, 使用LRU缓存机制后移除该逻辑
            # if len(index) > cls._max_cache_size:
            #     # 按时间戳排序，删除最旧的缓存
            #     sorted_items = sorted(index.items(), key=lambda x: x[1]['timestamp'])
            #     items_to_remove = sorted_items[:len(index) - cls._max_cache_size]
                
            #     for cache_path, _ in items_to_remove:
            #         if os.path.exists(cache_path):
            #             os.remove(cache_path)
            #         del index[cache_path]

            # 保存索引
            with open(cls._cache_index_file, 'w', encoding='utf-8', errors='ignore') as f:
                json.dump(index, f, ensure_ascii=False, indent=4)

        except Exception as e:
            print(f"更新缓存索引失败: {e}")
            # show_message_box(f"_update_cache_index更新缓存索引失败: {e}", "提示", 1500)

    @classmethod
    def clear_cache(cls):
        """清理本地中的缓存"""
        try:

            # 清除lru_cache的缓存
            cls.get_icon.cache_clear()  

            # 清理内存缓存
            cls._cache.clear()

            # 清理文件缓存
            if os.path.exists(cls._cache_dir):
                shutil.rmtree(cls._cache_dir)
            if os.path.exists(cls._cache_index_file):
                os.remove(cls._cache_index_file)

        except Exception as e:
            print(f"清理缓存失败: {e}")
            # show_message_box(f"clear_cache清理缓存失败: {e}", "提示", 1500)
        

"""图片旋转exif信息调整类"""
class ImageTransform:
    # 定义EXIF方向值对应的QTransform变换
    _ORIENTATION_TRANSFORMS = {
        1: QTransform(),  # 0度 - 正常
        2: QTransform().scale(-1, 1),  # 水平翻转
        3: QTransform().rotate(180),  # 180度
        4: QTransform().scale(1, -1),  # 垂直翻转
        5: QTransform().rotate(90).scale(-1, 1),  # 顺时针90度+水平翻转
        6: QTransform().rotate(90),  # 顺时针90度
        7: QTransform().rotate(-90).scale(-1, 1),  # 逆时针90度+水平翻转
        8: QTransform().rotate(-90)  # 逆时针90度
    }
    
    @classmethod
    def get_orientation(cls, image_path):
        """获取图片的EXIF方向信息（优化版）"""
        try:
            with Image.open(image_path) as img:
                
                # 检查是否是支持EXIF的格式
                if img.format not in ('JPEG', 'TIFF','MPO'):
                    return 1
                
                # 获取EXIF数据（使用更可靠的获取方式）
                exif_data = img.info.get('exif')
                if not exif_data:
                    return 1
                
                # 使用piexif的字节加载方式
                exif_dict = piexif.load(exif_data)
                return exif_dict['0th'].get(piexif.ImageIFD.Orientation, 1)
                
        except (KeyError, AttributeError, ValueError):
            # 当EXIF数据不包含方向信息时
            return 1
        except Exception as e:
            print(f"读取EXIF方向信息失败: {str(e)}")
            return 1

    @classmethod
    def auto_rotate_image(cls, icon_path: str) -> QPixmap:
        try:
            # 获取EXIF方向信息
            orientation = cls.get_orientation(icon_path)
            
            # 创建QPixmap
            pixmap = QPixmap(icon_path)
            
            # 应用方向变换
            transform = cls._ORIENTATION_TRANSFORMS.get(orientation, QTransform())
            if not transform.isIdentity():  # 只在需要变换时执行
                pixmap = pixmap.transformed(transform, Qt.SmoothTransformation)
            
            return pixmap
            
        except Exception as e:
            print(f"处理图片失败 {icon_path}: {str(e)}")
            return QPixmap()


class CheckBoxListModel(QAbstractListModel):
    """自定义数据模型，用于存储文件夹名和复选框的状态。"""

    def __init__(self, items):
        super(CheckBoxListModel, self).__init__()
        self.items = ["全选"] + sorted(items)  # 第一个项作为"全选"，其余按字母排序
        self.checked_states = [False] * len(self.items)  # 初始化所有项为未选中状态
        self.checked_order = []  # 新增：记录选中顺序的列表

    def rowCount(self, parent=QModelIndex()):
        """返回模型中的行数（文件夹项数）。"""
        return len(self.items)

    def data(self, index, role=Qt.DisplayRole):
        """根据索引和角色返回相应的数据。"""
        if not index.isValid():
            return QVariant()
        if role == Qt.DisplayRole:
            return self.items[index.row()]  # 返回项目名称
        if role == Qt.UserRole:
            return self.checked_states[index.row()]  # 返回选中状态
        return QVariant()

    def setChecked(self, index):
        if not index.isValid():
            return

        row = index.row()
        if row == 0:  # 全选逻辑保持不变
            all_checked = not self.checked_states[0]
            self.checked_states = [all_checked] * len(self.items)
            self.checked_order = self.items[1:] if all_checked else []
            self.dataChanged.emit(self.index(0), self.index(len(self.items) - 1))
        else:
            # 更新选中状态
            self.checked_states[row] = not self.checked_states[row]
            
            # 更新选中顺序
            item = self.items[row]
            if self.checked_states[row]:
                if item not in self.checked_order:
                    self.checked_order.append(item)
            else:
                if item in self.checked_order:
                    self.checked_order.remove(item)
            
            self.updateSelectAllState()

        self.dataChanged.emit(index, index)

    def updateSelectAllState(self):
        """检查是否所有项目都被选中，并更新"全选"的状态。"""
        all_selected = all(self.checked_states[1:])
        self.checked_states[0] = all_selected
        self.dataChanged.emit(self.index(0), self.index(0))  # 更新"全选"选项的显示

    def getCheckedItems(self):
        """获取当前被选中的文件夹列表（按点击顺序）"""
        # 直接返回记录顺序的列表
        return self.checked_order.copy()


class CheckBoxDelegate(QStyledItemDelegate):
    """自定义委托，用于在 ComboBox 中绘制复选框。"""

    def paint(self, painter, option, index):
        """绘制复选框和文本。"""
        checked = index.data(Qt.UserRole)
        is_hovered = option.state & QStyle.State_MouseOver  # 检查鼠标是否悬停
        if is_hovered:
            # 设置鼠标悬停的颜色为加载的配置文件中的背景颜色，字体颜色为黑色
            background_color = load_color_settings()['background_color_default'] if load_color_settings()['background_color_default'] else "rgb(173, 216, 230)"
            background_color = rgb_str_to_qcolor(background_color) # 将字符串转换为QColor
            painter.fillRect(option.rect, background_color)  # 鼠标悬停时的颜色
            # painter.setPen(QPen(Qt.black))  # 设置字体颜色为黑色

        # 绘制复选框
        checkbox_style_option = QStyleOptionButton()
        checkbox_style_option.rect = option.rect.adjusted(0, 0, 0, 0)
        checkbox_style_option.state = QStyle.State_On if checked else QStyle.State_Off
        checkbox_style_option.state |= QStyle.State_Enabled

        QApplication.style().drawControl(QStyle.CE_CheckBox, checkbox_style_option, painter)
        # 绘制文本
        text_rect = option.rect
        text_rect.adjust(25, 0, 0, 0)  # 调整文本位置
        painter.drawText(text_rect, Qt.AlignVCenter, index.data(Qt.DisplayRole))

    def sizeHint(self, option, index):
        """返回项的大小，包括缩进和其他空间设置。"""
        size = super(CheckBoxDelegate, self).sizeHint(option, index)
        return QSize(size.width(), size.height())  # 设置项的大小
    

class SingleFileRenameDialog(QDialog):
    """单文件重命名对话框类"""
    def __init__(self, file_path, parent=None):
        super().__init__(parent)
        self.file_path = file_path
        self.file_name = os.path.basename(file_path)
        self.file_dir = os.path.dirname(file_path)
        self.name_without_ext, self.ext = os.path.splitext(self.file_name)
        self.new_file_path = None  # 添加新文件路径属性
        
        # 添加设置对象用于保存复选框状态
        self.settings = QtCore.QSettings('HiViewer', 'SingleFileRename')
        
        self.setup_ui()
        self.setup_connections()

    def setup_ui(self):
        """初始化UI"""
        self.setWindowTitle("重命名文件")
        self.setFixedSize(650, 180)
        
        # 主布局
        layout = QVBoxLayout()
        
        # 文件名显示
        self.file_label = QLabel(f"文件名：{self.file_name}")
        layout.addWidget(self.file_label)
        
        # 重命名输入区域
        name_layout = QHBoxLayout()
        name_label = QLabel("重命名为：")
        self.name_input = QLineEdit(self.name_without_ext)
        name_layout.addWidget(name_label)
        name_layout.addWidget(self.name_input)
        layout.addLayout(name_layout)
        
        # 显示扩展名选项 - 从设置中读取上次的状态
        show_ext_layout = QHBoxLayout()
        self.show_ext_checkbox = QCheckBox("显示扩展名")
        # 读取上次的选择,默认为False
        last_state = self.settings.value('show_extension', False, type=bool)
        self.show_ext_checkbox.setChecked(last_state)
        show_ext_layout.addWidget(self.show_ext_checkbox)
        show_ext_layout.addStretch()
        layout.addLayout(show_ext_layout)
        
        # 按钮区域
        button_layout = QHBoxLayout()
        button_layout.addStretch()
        self.ok_button = QPushButton("确定")
        self.cancel_button = QPushButton("取消")
        button_layout.addWidget(self.ok_button)
        button_layout.addWidget(self.cancel_button)
        layout.addLayout(button_layout)
        
        self.setLayout(layout)

        # 如果上次选择显示扩展名,则在输入框中显示完整文件名
        if last_state:
            self.name_input.setText(self.name_without_ext + self.ext)

    def setup_connections(self):
        """设置信号连接"""
        self.ok_button.clicked.connect(self.on_ok_clicked)
        self.cancel_button.clicked.connect(self.reject)
        self.show_ext_checkbox.stateChanged.connect(self.on_checkbox_changed)

    def on_checkbox_changed(self, state):
        """处理显示扩展名复选框状态改变"""
        # 保存当前选择到设置中
        self.settings.setValue('show_extension', state == Qt.Checked)
        
        if state == Qt.Checked:
            self.name_input.setText(self.name_without_ext + self.ext)
        else:
            self.name_input.setText(self.name_without_ext)

    def on_ok_clicked(self):
        """处理确定按钮点击"""
        new_name = self.name_input.text()
        if not new_name:
            show_message_box("文件名不能为空！", "错误", 500)
            return
            
        # 构建新文件路径
        if not self.show_ext_checkbox.isChecked():
            new_name = new_name + self.ext
        new_path = os.path.join(self.file_dir, new_name)
        
        # 检查文件是否已存在
        if os.path.exists(new_path) and new_path != self.file_path:
            show_message_box("文件已存在！", "错误", 500)
            return
            
        try:
            os.rename(self.file_path, new_path)
            self.new_file_path = new_path  # 更新新文件路径
            self.accept()
        except Exception as e:
            show_message_box(f"重命名失败: {str(e)}", "错误", 1000)

    def get_new_file_path(self):
        """返回新的文件路径"""
        return self.new_file_path
    
    


"""
设置独立封装类区域结束线
------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
"""


"""
设置主界面类区域开始线
------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
"""
class HiviewerMainwindow(QtWidgets.QMainWindow, Ui_MainWindow):
    def __init__(self, parent=None):
        super(HiviewerMainwindow, self).__init__(parent)

        # 1 设置版本信息
        default_version_path = os.path.join(os.path.dirname(__file__), "cache", "version.ini")
        # 读取本地配置文件中的版本信息
        self.version_info = version_init(default_version_path, VERSION='release-v2.3.2')
        # 获取github中发布的最新版本信息, self.new_version_info = pre_check_update()
        # 转移到函数self.create_splash_screen() --> self.update_splash_message --> self.pre_update() 中获取,先在此处初始化
        self.new_version_info = False 
        # 2 创建启动画面
        try:
            print("create_splash_screen()--创建启动画面")
            self.create_splash_screen()
        except Exception as e:
            print(f"create_splash_screen()--创建启动画面失败: {e}")

        # 3 设置主界面UI
        try:
            print("setupUi()--初始化主界面UI")
            # 默认设置是图片拖拽模式, self.setupUi(self) 中需要调用
            self.drag_flag = True
            self.setupUi(self)
        except Exception as e:
            print(f"setupUi()--初始化主界面UI失败: {e}")
        
        # 4 初始化其它所有组件
        try:
            print("initialize_components()--初始化其它组件")
            self.initialize_components()
        except Exception as e:
            print(f"initialize_components()--初始化其它组件失败: {e}")


    def initialize_components(self):
        """初始化所有组件"""

        # 设置图片&视频文件格式
        self.IMAGE_FORMATS = ('.png', '.jpg', '.jpeg', '.bmp', '.gif', '.tif', '.ico', '.webp')
        self.VIDEO_FORMATS = ('.mp4', '.avi', '.mov', '.wmv', '.mpeg', '.mpg', '.mkv')

        # 初始化属性
        self.files_list = []            # 文件名及基本信息列表
        self.paths_list = []            # 文件路径列表
        self.dirnames_list = []         # 选中的同级文件夹列表
        self.image_index_max = []       # 存储当前选中及复选框选中的，所有图片列有效行最大值
        self.preloading_file_name_paths = []  # 预加载图标前的文件路径列表
        self.compare_window = None            # 添加子窗口引用
        self.task_active = False              # 定时器任务变量
        self.last_key_press = False           # 记录第一次按下键盘空格键或B键
        self.selected_folders_history = False # 记录是否有效点击复选框，避免self.RT_QComboBox1的press事件出现重复连接信号的情况
        self.simple_mode = True          # 设置默认模式为简单模式，同EXIF信息功能
        self.current_theme = "默认主题"  # 设置初始主题为默认主题

        # 添加预加载相关的属性初始化
        self.current_preloader = None  # 当前预加载器引用
        self.preloading = False        # 预加载状态
        self.preload_queue = Queue()   # 预加载队列

        # 初始化线程池
        self.threadpool = QThreadPool()
        self.threadpool.setMaxThreadCount(max(4, os.cpu_count()))  

        # 初始化压缩工作线程,压缩包路径
        self.zip_path = None  
        self.compress_worker = None

        """加载颜色相关设置""" # 设置背景色和字体颜色，使用保存的设置或默认值
        self.color_settings = load_color_settings()
        self.background_color_default = self.color_settings.get("background_color_default", "rgb(173,216,230)")  # 深色背景色_好蓝
        self.background_color_table = self.color_settings.get("background_color_table", "rgb(127, 127, 127)")    # 表格背景色_18度灰
        self.font_color_default = self.color_settings.get("font_color_default", "rgb(0, 0, 0)")                  # 默认字体颜色_纯黑色
        self.font_color_exif = self.color_settings.get("font_color_exif", "rgb(255, 255, 255)")                  # Exif字体颜色_纯白色

        """加载字体相关设置""" # 初始化字体管理器,并获取字体，设置默认字体 self.custom_font
        font_paths = [
            os.path.join(os.path.dirname(os.path.abspath(__file__)), "fonts", "JetBrainsMapleMono_Regular.ttf"), # JetBrains Maple Mono
            os.path.join(os.path.dirname(os.path.abspath(__file__)), "fonts", "xialu_wenkai.ttf"),               # LXGW WenKai
            os.path.join(os.path.dirname(os.path.abspath(__file__)), "fonts", "MapleMonoNormal_Regular.ttf")     # Maple Mono Normal
        ]
        MultiFontManager.initialize(font_paths=font_paths)
        self.custom_font_jetbrains = MultiFontManager.get_font(font_family="JetBrains Maple Mono", size=12)
        self.custom_font_jetbrains_small = MultiFontManager.get_font(font_family="JetBrains Maple Mono", size=10)
        self.custom_font = self.custom_font_jetbrains
        if False: # 暂时移除，使用MultiFontManager.get_font()方法
            # 单个字体管理器，两种导入方式:
            # 第一种，直接使用字体管理器默认字体，只是恶
            self.custom_font = SingleFontManager.get_font(12)
            # 第二种，使用字体管理器初始化方法，传入字体路径    
            font_path_jetbrains = os.path.join(os.path.dirname(os.path.abspath(__file__)), "fonts", "JetBrainsMapleMono_Regular.ttf")
            self.custom_font = SingleFontManager.get_font(size=12, font_path=font_path_jetbrains)   
        

        # 设置主界面相关组件
        self.set_stylesheet()

        # 初始化主题，暂时移除，在load_settings() 中初始化
        # self.apply_theme()

        # 加载之前的设置    
        self.load_settings()  

        # 设置快捷键
        self.set_shortcut()

        # 设置右键菜单
        self.setup_context_menu()  

        # 模仿按下回车
        self.input_enter_action()  

        # 完成初始化后设置标志
        self.initialization_complete = True
        
        # 显示主窗口,暂时移除，在self.create_splash_screen()-->QTimer.singleShot(1000, self.show)函数中显示
        # self.show()


    """
    设置右键菜单函数区域开始线
    ---------------------------------------------------------------------------------------------------------------------------------------------
    """

    def setup_context_menu(self):
        """设置右键菜单"""
        self.context_menu = QMenu(self)
    
        # 设置菜单样式 modify by diamond_cz 20250217 优化右键菜单栏的显示
        self.context_menu.setStyleSheet(f"""
            QMenu {{
                /*background-color: #F0F0F0;   背景色 */

                font-family: "{self.custom_font_jetbrains_small.family()}";
                font-size: {self.custom_font_jetbrains_small.pointSize()}pt;    
            }}
            QMenu::item:selected {{
                background-color: {self.background_color_default};   /* 选中项背景色 */
                color: #000000;               /* 选中项字体颜色 */
            }}
        """)

        # 添加主菜单项并设置图标
        icon_path = os.path.join(os.path.dirname(__file__), "icons", "delete_ico_96x96.ico")
        delete_icon = QIcon(icon_path) 
        icon_path = os.path.join(os.path.dirname(__file__), "icons", "paste_ico_96x96.ico")
        paste_icon = QIcon(icon_path) 
        icon_path = os.path.join(os.path.dirname(__file__), "icons", "update_ico_96x96.ico")
        refresh_icon = QIcon(icon_path) 
        icon_path = os.path.join(os.path.dirname(__file__), "icons", "theme_ico_96x96.ico")
        theme_icon = QIcon(icon_path) 
        icon_path = os.path.join(os.path.dirname(__file__), "icons", "image_size_reduce_ico_96x96.ico")
        image_size_reduce_icon = QIcon(icon_path)
        icon_path = os.path.join(os.path.dirname(__file__), "icons", "ps_ico_96x96.ico")
        ps_icon = QIcon(icon_path) 
        icon_path = os.path.join(os.path.dirname(__file__), "icons", "cmd_ico_96x96.ico")
        command_icon = QIcon(icon_path)
        icon_path = os.path.join(os.path.dirname(__file__), "icons", "exif_ico_96x96.ico")
        exif_icon = QIcon(icon_path)
        icon_path = os.path.join(os.path.dirname(__file__), "icons", "raw_ico_96x96.ico")
        raw_icon = QIcon(icon_path)
        icon_path = os.path.join(os.path.dirname(__file__), "icons", "rename_ico_96x96.ico")
        rename_icon = QIcon(icon_path)
        icon_path = os.path.join(os.path.dirname(__file__), "icons", "about.ico")
        help_icon = QIcon(icon_path) 
        icon_path = os.path.join(os.path.dirname(__file__), "icons", "file_zip_ico_96x96.ico")
        zip_icon = QIcon(icon_path)
        icon_path = os.path.join(os.path.dirname(__file__), "icons", "TCP_ico_96x96.ico")
        tcp_icon = QIcon(icon_path)
        icon_path = os.path.join(os.path.dirname(__file__), "icons", "rorator_plus_ico_96x96.ico")
        rotator_icon = QIcon(icon_path)
        icon_path = os.path.join(os.path.dirname(__file__), "icons", "line_filtrate_ico_96x96.ico")
        filtrate_icon = QIcon(icon_path)
        icon_path = os.path.join(os.path.dirname(__file__), "icons", "win_folder_ico_96x96.ico")
        win_folder_icon = QIcon(icon_path)
        icon_path = os.path.join(os.path.dirname(__file__), "icons", "restart_ico_96x96.ico")
        restart_icon = QIcon(icon_path)


        # 创建二级菜单-删除选项
        sub_menu = QMenu("删除选项", self.context_menu) 
        sub_menu.setIcon(delete_icon)  
        sub_menu.addAction("从列表中删除(D)", self.delete_from_list)  
        sub_menu.addAction("从原文件删除(Ctrl+D)", self.delete_from_file)  

        # 创建二级菜单-复制选项
        sub_menu2 = QMenu("复制选项", self.context_menu)  
        sub_menu2.setIcon(paste_icon)  
        sub_menu2.addAction("复制文件路径(C)", self.copy_selected_file_path)  
        sub_menu2.addAction("复制文件(Ctrl+C)", self.copy_selected_files)  

        # 创建二级菜单-无损旋转
        sub_menu3 = QMenu("无损旋转", self.context_menu)  
        sub_menu3.setIcon(rotator_icon)  
        sub_menu3.addAction("逆时针旋转", lambda: self.jpg_lossless_rotator('l'))  
        sub_menu3.addAction("顺时针旋转", lambda: self.jpg_lossless_rotator('r'))  
        sub_menu3.addAction("旋转180度", lambda: self.jpg_lossless_rotator('u'))  
        sub_menu3.addAction("水平翻转", lambda: self.jpg_lossless_rotator('h'))  
        sub_menu3.addAction("垂直翻转", lambda: self.jpg_lossless_rotator('v'))  
        sub_menu3.addAction("自动校准EXIF旋转信息", lambda: self.jpg_lossless_rotator('auto'))  

        # 创建二级菜单-按行筛选
        sub_menu4 = QMenu("按行筛选", self.context_menu)  
        sub_menu4.setIcon(filtrate_icon)  
        sub_menu4.addAction("奇数行", lambda: self.show_filter_rows('odd'))  
        sub_menu4.addAction("偶数行", lambda: self.show_filter_rows('even'))  
        sub_menu4.addAction("3选1", lambda: self.show_filter_rows('three_1'))  
        sub_menu4.addAction("3选2", lambda: self.show_filter_rows('three_2'))  
        sub_menu4.addAction("5选1", lambda: self.show_filter_rows('five_1'))  

        # 将二级菜单添加到主菜单
        self.context_menu.addMenu(sub_menu)   
        self.context_menu.addMenu(sub_menu2)  
        self.context_menu.addMenu(sub_menu4)  
        self.context_menu.addMenu(sub_menu3)  
        
        # 设置右键菜单槽函数
        self.context_menu.addAction(exif_icon, "高通AEC10解析图片(I)", self.on_i_pressed)
        self.context_menu.addAction(zip_icon, "压缩文件(Z)", self.compress_selected_files)
        self.context_menu.addAction(theme_icon, "切换主题(P)", self.on_p_pressed)
        self.context_menu.addAction(image_size_reduce_icon, "图片瘦身(X)", self.jpgc_tool) 
        self.context_menu.addAction(ps_icon, "图片调整(L)", self.on_l_pressed)
        self.context_menu.addAction(tcp_icon, "截图功能(T)", self.tcp_tool)
        self.context_menu.addAction(command_icon, "批量执行命令工具(M)", self.execute_command)
        self.context_menu.addAction(rename_icon, "批量重命名工具(F4)", self.on_f4_pressed)
        self.context_menu.addAction(raw_icon, "RAW转JPG工具(F1)", self.on_f1_pressed)
        self.context_menu.addAction(win_folder_icon, "打开资源管理器(W)", self.reveal_in_explorer)
        self.context_menu.addAction(refresh_icon, "刷新(F5)", self.on_f5_pressed)
        self.context_menu.addAction(restart_icon, "重启程序", self.on_f12_pressed)
        self.context_menu.addAction(help_icon, "关于(Ctrl+H)", self.on_ctrl_h_pressed)

        # 连接右键菜单到表格
        self.RB_QTableWidget0.setContextMenuPolicy(Qt.CustomContextMenu)
        self.RB_QTableWidget0.customContextMenuRequested.connect(self.show_context_menu)

    def default(self):
        """删除"""
        show_message_box("暂未实现", "提示", 500)

    def show_context_menu(self, pos):
        """显示右键菜单"""
        self.context_menu.exec_(self.RB_QTableWidget0.mapToGlobal(pos))

    def create_splash_screen(self):
        """创建带渐入渐出效果的启动画面"""
        # 加载启动画面图片
        splash_path = os.path.join(os.path.dirname(__file__), "icons", "viewer_0.png")
        splash_pixmap = QPixmap(splash_path)
        
        if splash_pixmap.isNull():
            splash_pixmap = QPixmap(400, 200)
            splash_pixmap.fill(Qt.white)
            
        self.splash = QSplashScreen(splash_pixmap)
        
        # 获取当前屏幕并计算居中位置
        screen = QtWidgets.QApplication.desktop().screenNumber(QtWidgets.QApplication.desktop().cursor().pos())
        screen_geometry = QtWidgets.QApplication.desktop().screenGeometry(screen)
        x = screen_geometry.x() + (screen_geometry.width() - splash_pixmap.width()) // 2
        y = screen_geometry.y() + (screen_geometry.height() - splash_pixmap.height()) // 2
        self.splash.move(x, y)
        
        # 设置半透明效果
        self.splash.setWindowOpacity(0)
        
        # 创建渐入动画
        self.fade_anim = QtCore.QPropertyAnimation(self.splash, b"windowOpacity")
        self.fade_anim.setDuration(1000)  # 1000ms的渐入动画
        self.fade_anim.setStartValue(0)
        self.fade_anim.setEndValue(1)
        self.fade_anim.start()
        
        # 设置启动画面的样式
        self.splash.setStyleSheet("""
            QSplashScreen {
                background-color: rgba(0, 0, 0, 180);
                color: white;
                border-radius: 10px;
            }
        """)
        
        # 显示启动画面
        self.splash.show()
        
        # 启动进度更新定时器
        self.progress_timer = QTimer()
        self.progress_timer.timeout.connect(self.update_splash_message)
        self.dots_count = 0
        self.progress_timer.start(500)  # 每500ms更新一次

    def update_splash_message(self):
        """更新启动画面的加载消息"""
        # 更新进度点
        self.dots_count = (self.dots_count + 1) % 4
        dots = "." * self.dots_count
        
        # 使用HTML标签设置文字颜色为红色，并调整显示内容，文字颜色为配置文件（color_setting.json）中的背景颜色
        message = f'<div style="color: {self.background_color_default};">HiViewer_{self.version_info}</div>' \
                  f'<div style="color: {self.background_color_default};">正在启动...{dots}</div>'
        
        # 显示启动消息
        self.splash.showMessage(
            message, 
            Qt.AlignCenter | Qt.AlignBottom,
            Qt.white
        )
        
        # 检查是否完成初始化
        if hasattr(self, 'initialization_complete'):
            # 创建渐出动画
            self.fade_out = QtCore.QPropertyAnimation(self.splash, b"windowOpacity")
            self.fade_out.setDuration(1000)  # 1000ms的渐出动画
            self.fade_out.setStartValue(1)
            self.fade_out.setEndValue(0)
            self.fade_out.finished.connect(self.splash.close)
            self.fade_out.start()
            
            # 停止定时器
            self.progress_timer.stop()

            # 获取当前屏幕并计算居中位置，移动到该位置
            screen = QtWidgets.QApplication.desktop().screenNumber(QtWidgets.QApplication.desktop().cursor().pos())
            screen_geometry = QtWidgets.QApplication.desktop().screenGeometry(screen)
            x = screen_geometry.x() + (screen_geometry.width() - self.width()) // 2
            y = screen_geometry.y() + (screen_geometry.height() - self.height()) // 2
            self.move(x, y)

            # 预检查更新
            self.pre_update()

            # 显示主窗口
            self.show()

            # 延时显示主窗口,方便启动画面渐出  pre_update
            # QTimer.singleShot(1000, self.show)

            # 延时检查更新
            # QTimer.singleShot(3000, self.pre_update)



    
    """
    设置右键菜单函数区域结束线
    ---------------------------------------------------------------------------------------------------------------------------------------------
    """

    def set_stylesheet(self):
        """设置主界面图标以及标题"""
        print("set_stylesheet()--设置主界面相关组件")

        icon_path = os.path.join(os.path.dirname(__file__), "icons", "viewer_3.ico")
        self.setWindowIcon(QIcon(icon_path))
        self.setWindowTitle(f"HiViewer")

        # 设置窗口尺寸为分辨率的一半,改为固定比例
        # 主屏幕的几何信息
        # screen = QtWidgets.QApplication.desktop().screenGeometry() 
        # 根据鼠标的位置返回当前光标所在屏幕的几何信息
        screen = QtWidgets.QApplication.desktop().screenNumber(QtWidgets.QApplication.desktop().cursor().pos())
        screen_geometry = QtWidgets.QApplication.desktop().screenGeometry(screen)
        width = int(screen_geometry.width() * 0.65)
        height = int(screen_geometry.height() * 0.65)
        self.resize(width, height)

        # 启用拖放功能
        self.setAcceptDrops(True)

        """界面底部状态栏设置"""
        # self.statusbar --> self.statusbar_widget --> self.statusbar_QHBoxLayout --> self.statusbar_button1 self.statusbar_button2
        # 设置按钮无边框
        self.statusbar_button1.setFlat(True)
        self.statusbar_button2.setFlat(True)

        # 初始化版本更新按钮文本
        self.statusbar_button2.setText(f"🚀版本({self.version_info})")            

        # 初始化标签文本
        self.statusbar_label1.setText(f"🔉: 进度提示标签🍃")  # 根据需要设置标签的文本
        self.statusbar_label0.setText(f"📢:选中或筛选的文件夹中包含{self.image_index_max}张图")
        self.statusbar_label.setText(f"[0]已选择")

        
        """ 左侧组件
        设置左侧组件显示风格，背景颜色为淡蓝色，四角为圆形; 下面显示左侧组件name 
        self.Left_QTreeView | self.Left_QFrame
        self.L_radioButton1 | self.L_radioButton2 | self.L_pushButton1 | self.L_pushButton2
        modify by diamond_cz 20250403 移除按钮self.L_pushButton1 | self.L_pushButton2
        """  
        # 设置左侧按钮和复选框组件
        self.L_radioButton1.setChecked(True)  # 默认选择第一个单选按钮
        self.L_radioButton1.setText("隐藏文件")  # 设置按钮文本
        self.L_radioButton2.setText("显示文件")  # 设置按钮文本

        # self.Left_QTreeView
        self.file_system_model = QFileSystemModel(self)
        self.file_system_model.setRootPath('')  # 设置根路径为空，表示显示所有磁盘和文件夹
        self.Left_QTreeView.setModel(self.file_system_model)

        # 隐藏不需要的列，只显示名称列
        self.Left_QTreeView.header().hide()  # 隐藏列标题
        self.Left_QTreeView.setColumnWidth(0, 650)  # 设置名称列宽度，以显示横向滚动条
        self.Left_QTreeView.setColumnHidden(1, True)  # 隐藏大小列
        self.Left_QTreeView.setColumnHidden(2, True)  # 隐藏类型列
        self.Left_QTreeView.setColumnHidden(3, True)  # 隐藏修改日期列 

        # # 使用QDir的过滤器只显示文件夹
        self.file_system_model.setFilter(QDir.NoDot | QDir.NoDotDot | QDir.AllDirs)  



        """ 右侧组件
        设置右侧组件显示风格（列出了右侧第一行第二行第三行的组件名称）
        self.RT_QComboBox | self.RT_QPushButton2 | self.RT_QPushButton3
        self.RT_QComboBox0 | self.RT_QComboBox1 | self.RT_QComboBox2 | self.RT_QComboBox3 | self.RT_QPushButton5 | self.RT_QPushbutton6
        self.RB_QTableWidget0 | self.RB_Label
        """

        # 设置当前目录到地址栏，并将地址栏的文件夹定位到左侧文件浏览器中
        current_directory = os.path.dirname(os.path.abspath(__file__).capitalize())
        self.RT_QComboBox.addItem(current_directory)
        self.RT_QComboBox.lineEdit().setPlaceholderText("请在地址栏输入一个有效的路径")  # 设置提示文本
        
        # RB_QTableWidget0表格设置
        self.RB_QTableWidget0.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch) # 设置表格列宽自适应
   
        # RT_QComboBox0 添加下拉框选项
        self.RT_QComboBox0.addItem("显示图片文件")
        self.RT_QComboBox0.addItem("显示视频文件")
        self.RT_QComboBox0.addItem("显示所有文件")

        # RT_QComboBox2 添加下拉框选项
        self.RT_QComboBox2.addItem("按创建时间排序")
        self.RT_QComboBox2.addItem("按文件名称排序")
        self.RT_QComboBox2.addItem("按修改时间排序")
        self.RT_QComboBox2.addItem("按文件大小排序")
        self.RT_QComboBox2.addItem("按曝光时间排序")
        self.RT_QComboBox2.addItem("按ISO排序")
        self.RT_QComboBox2.addItem("按创建时间逆序排序")
        self.RT_QComboBox2.addItem("按文件名称逆序排序")
        self.RT_QComboBox2.addItem("按修改时间逆序排序")
        self.RT_QComboBox2.addItem("按文件大小逆序排序")
        self.RT_QComboBox2.addItem("按曝光时间逆序排序")
        self.RT_QComboBox2.addItem("按ISO逆序排序")

        # RT_QComboBox3 添加下拉框选项
        self.RT_QComboBox3.addItem("默认主题")
        self.RT_QComboBox3.addItem("暗黑主题")

        """RT_QComboBox1待完善功能: 在下拉框中多次选择复选框后再收起下拉框; modify by 2025-01-21, 在main_ui.py中使用自定义的 ComboBox已解决"""
        self.RT_QComboBox1.setEditable(True)  # 设置可编辑
        self.RT_QComboBox1.lineEdit().setReadOnly(True)  # 设置不可编辑
        self.RT_QComboBox1.lineEdit().setPlaceholderText("请选择")  # 设置提示文本
        

    def set_shortcut(self):
        """快捷键和槽函数连接事件"""

        """1.快捷键设置"""
        # 添加快捷键 切换主题
        self.p_shortcut = QShortcut(QKeySequence('p'), self)
        self.p_shortcut.activated.connect(self.on_p_pressed)
        # 添加快捷键，打开命令工具
        self.m_shortcut = QShortcut(QKeySequence('M'), self)
        self.m_shortcut.activated.connect(self.open_bat_tool)
        # 添加快捷键，切换下一组图片/视频
        self.space_shortcut = QShortcut(QKeySequence(Qt.Key_Space), self)
        self.space_shortcut.activated.connect(self.on_space_pressed)
        # 退出界面使用ALT+Q替换原来的ESC（Qt.Key_Escape），防误触
        self.esc_shortcut = QShortcut(QKeySequence(Qt.AltModifier + Qt.Key_Q), self)
        self.esc_shortcut.activated.connect(self.on_escape_pressed)
        # 拖拽模式使用ALT快捷键
        self.esc_shortcut = QShortcut(QKeySequence(Qt.AltModifier + Qt.Key_A), self)
        self.esc_shortcut.activated.connect(self.on_alt_pressed)
        # 极简模式和EXIF信息切换使用ALT+I快捷键
        self.esc_shortcut = QShortcut(QKeySequence(Qt.AltModifier + Qt.Key_I), self)
        self.esc_shortcut.activated.connect(self.show_exif)
        # 添加快捷键，切换上一组图片/视频
        self.b_shortcut = QShortcut(QKeySequence('b'), self)
        self.b_shortcut.activated.connect(self.on_b_pressed)
        # 添加快捷键 F1，打开MIPI RAW文件转换为JPG文件工具
        self.f1_shortcut = QShortcut(QKeySequence(Qt.Key_F1), self)
        self.f1_shortcut.activated.connect(self.on_f1_pressed)
        # 添加快捷键，打开批量执行命令工具
        self.f2_shortcut = QShortcut(QKeySequence(Qt.Key_F2), self)
        self.f2_shortcut.activated.connect(self.on_f2_pressed)
        # 添加快捷键，打开批量重命名工具
        self.f4_shortcut = QShortcut(QKeySequence(Qt.Key_F4), self)
        self.f4_shortcut.activated.connect(self.on_f4_pressed)
        # 添加快捷键 F5,刷新表格
        self.f5_shortcut = QShortcut(QKeySequence(Qt.Key_F5), self)
        self.f5_shortcut.activated.connect(self.on_f5_pressed)
        # 添加快捷键 i 切换极简模式
        self.p_shortcut = QShortcut(QKeySequence('i'), self)
        self.p_shortcut.activated.connect(self.on_i_pressed)
        # 添加快捷键 Ctrl+i 打开图片处理窗口
        self.i_shortcut = QShortcut(QKeySequence('l'), self)
        self.i_shortcut.activated.connect(self.on_l_pressed)
        # 添加快捷键 Ctrl+h 打开帮助信息显示
        self.h_shortcut = QShortcut(QKeySequence(Qt.ControlModifier + Qt.Key_H), self)
        self.h_shortcut.activated.connect(self.on_ctrl_h_pressed)
        # 添加快捷键 C,复制选中的文件路径
        self.c_shortcut = QShortcut(QKeySequence('c'), self)
        self.c_shortcut.activated.connect(self.copy_selected_file_path)
        # 添加快捷键 Ctrl+c 复制选中的文件
        self.c_shortcut = QShortcut(QKeySequence(Qt.ControlModifier + Qt.Key_C), self)
        self.c_shortcut.activated.connect(self.copy_selected_files)
        # 添加快捷键 D 从列表中删除选中的文件
        self.d_shortcut = QShortcut(QKeySequence('d'), self)
        self.d_shortcut.activated.connect(self.delete_from_list)
        # 添加快捷键 Ctrl+d 从原文件删除选中的文件
        self.d_shortcut = QShortcut(QKeySequence(Qt.ControlModifier + Qt.Key_D), self)
        self.d_shortcut.activated.connect(self.delete_from_file)
        # 添加快捷键 Z 压缩选中的文件
        self.z_shortcut = QShortcut(QKeySequence('z'), self)
        self.z_shortcut.activated.connect(self.compress_selected_files)
        # 添加快捷键 T 打开--局域网传输工具--，改为截图功能
        self.z_shortcut = QShortcut(QKeySequence('t'), self)
        self.z_shortcut.activated.connect(self.tcp_tool)
        # 添加快捷键 X 打开图片体积压缩工具
        self.z_shortcut = QShortcut(QKeySequence('x'), self)
        self.z_shortcut.activated.connect(self.jpgc_tool) 
        # 添加快捷键 W 打开资源管理器
        self.z_shortcut = QShortcut(QKeySequence('w'), self)
        self.z_shortcut.activated.connect(self.reveal_in_explorer) 

        """2. 槽函数连接事件"""
        # 连接左侧按钮槽函数
        self.Left_QTreeView.clicked.connect(self.update_combobox)        # 点击左侧文件浏览器时的连接事件
        self.L_radioButton1.toggled.connect(self.radio_button_file_off)  # 隐藏文件，只显示文件夹
        self.L_radioButton2.toggled.connect(self.radio_button_file_on)   # 显示所有文件
        
        # 连接右侧按钮槽函数
        self.RT_QComboBox.lineEdit().returnPressed.connect(self.input_enter_action) # 用户在地址栏输入文件路径后按下回车的动作反馈
        self.RT_QComboBox0.activated.connect(self.handleComboBox0Pressed)           # 点击（显示图片视频所有文件）下拉框选项时的处理事件
        self.RT_QComboBox1.view().pressed.connect(self.handleComboBoxPressed)       # 处理复选框选项被按下时的事件
        self.RT_QComboBox1.activated.connect(self.updateComboBox1Text)              # 更新显示文本
        self.RT_QComboBox2.activated.connect(self.handle_sort_option)               # 点击下拉框选项时，更新右侧表格
        self.RT_QComboBox3.activated.connect(self.handle_theme_selection)           # 点击下拉框选项时，更新主题
        self.RT_QPushButton3.clicked.connect(self.clear_combox)                     # 清除地址栏
        self.RT_QPushButton5.clicked.connect(self.compare)                          # 打开看图工具

        # 添加表格选择变化的信号连接 f"🎯[{count}]已选中"
        self.RB_QTableWidget0.itemSelectionChanged.connect(self.handle_table_selection)
        
        # 底部状态栏按钮连接函数
        self.statusbar_button1.clicked.connect(self.setting)   # 🔆设置按钮槽函数
        self.statusbar_button2.clicked.connect(self.update)    # 🚀版本按钮槽函数
        

    """
    左侧信号槽函数
    """
    def radio_button_file_off(self, checked):
        if checked:
            # 设置过滤器，只显示文件夹
            self.file_system_model.setFilter(QDir.NoDot | QDir.NoDotDot | QDir.AllDirs)  # 使用QDir的过滤器,隐藏文件,只显示文件夹
            print("L_radioButton1 被选中")
        else:
            print("L_radioButton1 未被选中")

    def radio_button_file_on(self, checked):
        if checked:
            print("L_radioButton2 被选中")
            self.file_system_model.setFilter(QDir.NoDot | QDir.NoDotDot |QDir.AllEntries)  # 显示所有文件和文件夹
        else:
            print("L_radioButton2 未被选中")

    def handle_single_image_editing(self):
        # 单个图片编辑功能
        print("单个图片编辑功能按钮被点击")
        # 打开图片编辑子界面, 同按下L键
        self.on_l_pressed()

    def batch_rename(self):
        print("批量重命名按钮被点击")
        # 获取当前选中的文件夹上一级文件夹路径
        current_folder = self.RT_QComboBox.currentText()
        current_folder = os.path.dirname(current_folder) 
        if current_folder:
            self.open_rename_tool(current_folder)
        else:
            # 弹出提示框, 0.5秒后自动关闭
            show_message_box("当前没有选中的文件夹", "提示", 500)

    """
    右侧信号槽函数
    """
    def input_enter_action(self):
        # 地址栏输入后按下回车的反馈
        print("input_enter_action()--在地址栏按下回车/拖拽了文件进来,开始在左侧文浏览器中定位")  # 打印输入内容
        self.locate_in_tree_view()
        # 初始化同级文件夹下拉框选项
        self.RT_QComboBox1_init()
        # 更新右侧表格
        self.update_RB_QTableWidget0()

    def clear_combox(self):
        print("clear_combox()--清除按钮被点击")
        # 清空地址栏
        self.RT_QComboBox.clear()
        # 刷新右侧表格
        self.update_RB_QTableWidget0()
        # 手动清除图标缓存
        IconCache.clear_cache() 
        # 释放内存
        self.cleanup() 
        
    
    def execute_command(self):
        print("execute_command()--命令按钮被点击")
        try:    
            self.open_bat_tool()
        except Exception as e:
            print(f"execute_command()-error--打开批量执行命令工具失败: {e}")
            return

    def compare(self):
        print("compare()-对比按钮被点击--调用on_space_pressed()")
        self.on_space_pressed()


    def setting(self):
        print("setting()-设置按钮被点击--setting()")
        # self.on_space_pressed()

    def update(self):
        print("setting()-版本按钮被点击--update()")
        check_update()

    def pre_update(self):
        print("pre_update()--预更新版本")

        # 获取self.new_version_info最新版本信息
        self.new_version_info = pre_check_update()
        if self.new_version_info:
            self.statusbar_button2.setToolTip(f"🚀有新版本可用: {self.version_info}-->{self.new_version_info}")
            self.apply_theme() # 更新样式表
        else:
            self.statusbar_button2.setToolTip("当前已是最新版本")
        


    def show_exif(self):
        """打开Exif信息显示，类似快捷键CTRL+P功能  """
        print("show_exif()--打开Exif信息显示")

        try:
            # 获取当前选中的文件类型
            selected_option = self.RT_QComboBox0.currentText()
            if selected_option == "显示所有文件":
                show_message_box("该功能只在显示图片文件时有效！", "提示", 500)
                return
            elif selected_option == "显示视频文件":
                show_message_box("该功能只在显示图片文件时有效！", "提示", 500)
                return
            elif selected_option == "显示图片文件":
                self.simple_mode = not self.simple_mode 

            if self.simple_mode:
                show_message_box("关闭Exif信息显示", "提示", 500)
            else:
                show_message_box("打开Exif信息显示", "提示", 500)
        except Exception as e:
            print(f"show_exif()-error--打开Exif信息显示失败: {e}")
        finally:
            # 更新 RB_QTableWidget0 中的内容    
            self.update_RB_QTableWidget0() 

    
    def show_filter_rows(self, row_type):
        """显示筛选行"""
        print(f"show_filter_rows()--显示筛选行")
        try:
            # 按照传入的行类型，筛选行，显示需要的行
            if row_type == 'odd': # 传入奇数行，需要先选中偶数行，然后从列表中删除偶数行，最后显示奇数行
                self.filter_rows('even')
                self.delete_from_list()
            elif row_type == 'even': # 传入偶数行，需要先选中奇数行，然后从列表中删除奇数行，最后显示偶数行
                self.filter_rows('odd')
                self.delete_from_list()
            elif row_type == 'three_1': # 传入3选1，需要先选中3选2，然后从列表中删除3选2，最后显示3选1
                self.filter_rows('three_2')
                self.delete_from_list()
            elif row_type == 'three_2': # 传入3选2，需要先选中3选1，然后从列表中删除3选1，最后显示3选2
                self.filter_rows('three_1')
                self.delete_from_list()
            elif row_type == 'five_1': # 传入5选1，需要先选中5选4，然后从列表中删除5选4，最后显示5选1
                self.filter_rows('five_4')
                self.delete_from_list()
            else:
                show_message_box(f"未知筛选模式: {row_type}", "错误", 1000)
        except Exception as e:
            print(f"show_filter_rows()-error--显示筛选行失败: {e}")
            return

    def filter_rows(self, row_type):
        """批量选中指定模式行（使用类switch结构优化）"""
        
        # 清空选中状态
        self.RB_QTableWidget0.clearSelection()
        # 获取总行数
        total_rows = self.RB_QTableWidget0.rowCount()
        # 获取选中状态
        selection = self.RB_QTableWidget0.selectionModel()
        # 定义选择范围
        selection_range = QItemSelection()

        # 定义条件映射字典（实际行号从1开始计算）
        condition_map = {
            'odd': lambda r: (r + 1) % 2 == 1,  # 奇数行（1,3,5...）
            'even': lambda r: (r + 1) % 2 == 0,  # 偶数行（2,4,6...）
            'three_1': lambda r: (r + 1) % 3 == 1,  # 3选1（1,4,7...）
            'three_2': lambda r: (r + 1) % 3 != 0,  # 3选2（1,2,4,5...）
            'five_1': lambda r: (r + 1) % 5 == 1,  # 5选1（1,6,11...）
            'five_4': lambda r: (r + 1) % 5 != 1  # 5选4（2,3,4,5...）
        }

        # 获取判断条件
        condition = condition_map.get(row_type)
        if not condition:
            show_message_box(f"未知筛选模式: {row_type}", "错误", 1000)
            return

        try:
            # 批量选择符合条件的行
            for row in range(total_rows):
                if condition(row):
                    row_selection = QItemSelection(
                        self.RB_QTableWidget0.model().index(row, 0),
                        self.RB_QTableWidget0.model().index(row, self.RB_QTableWidget0.columnCount()-1)
                    )
                    selection_range.merge(row_selection, QItemSelectionModel.Select)

            # 应用选择并滚动定位
            if not selection_range.isEmpty():
                selection.select(selection_range, QItemSelectionModel.Select | QItemSelectionModel.Rows)
                first_row = selection_range[0].top()
                self.RB_QTableWidget0.scrollTo(
                    self.RB_QTableWidget0.model().index(first_row, 0),
                    QAbstractItemView.PositionAtTop
                )

        except Exception as e:
            print(f"filter_rows()-error--批量选中指定模式行失败: {e}")
            return

    def jpg_lossless_rotator(self, para=''):
        """无损旋转图片"""
        print(f"jpg_lossless_rotator()-启动无损旋转图片任务:")
        try:
            # 取消当前的预加载任务
            self.cancel_preloading()

            # 构建jpegoptim的完整路径
            jpegr_path = os.path.join(os.path.dirname(__file__), 'tools', 'jpegr_lossless_rotator', 'jpegr.exe')
            if not os.path.exists(jpegr_path):
                show_message_box(f"jpegr.exe 不存在，请检查/tools/jpegr_lossless_rotator/", "提示", 1500)
                return
            
            # 获取选中的单元格中的路径
            files = self.copy_selected_file_path(0)
            # 获取选中的文件夹
            target_dir_paths = {os.path.dirname(file) for file in files}
            
            # 创建进度条
            if para == 'auto':
                progress_dialog = QProgressDialog("正在无损旋转图片...", "取消", 0, len(target_dir_paths), self)
            else:
                progress_dialog = QProgressDialog("正在无损旋转图片...", "取消", 0, len(files), self)
            progress_dialog.setWindowTitle("无损旋转进度")
            progress_dialog.setModal(True)
            progress_dialog.setFixedSize(450, 150)
            progress_dialog.setStyleSheet("QProgressDialog { border: none; }")
            progress_dialog.setVisible(False)

            if para == 'auto' and target_dir_paths:
                # 显示进度条,及时响应
                progress_dialog.setVisible(True)
                progress_dialog.setValue(0)
                QApplication.processEvents()
                
                for index_, dir_path in enumerate(target_dir_paths):

                    # 拼接参数命令字符串
                    command = f"{jpegr_path} -{para} -s \"{dir_path}\""

                    # 调用jpegoptim命令并捕获返回值
                    result = subprocess.run(command, check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)

                    # 检查返回码
                    if result.returncode == 0:
                        progress_dialog.setValue(index_ + 1)
                        if progress_dialog.wasCanceled():
                            show_message_box(f"用户手动自动校准EXIF旋转信息操作, \n已自动校准前{index_+1}个文件夹,共{len(target_dir_paths)}张", "提示", 3000)
                            break  # 如果用户取消了操作，则退出循环
                    else:
                        print("自动校准EXIF旋转信息命令执行失败, 返回码:", result.returncode)
                        print("错误信息:", result.stderr)
                        return
                    
                # 添加进度条完成后的销毁逻辑
                progress_dialog.finished.connect(progress_dialog.deleteLater)  # 进度条完成时销毁    

                show_message_box("自动校准EXIF旋转信息成功!", "提示", 1500) 

                # 清图标缓存，刷新表格
                IconCache.clear_cache()

                # 更新表格
                self.update_RB_QTableWidget0() 

                # 退出当前函数
                return
                    
            # 进行无损旋转相关的调用
            if files:
                # 显示进度条,及时响应
                progress_dialog.setVisible(True)
                progress_dialog.setValue(0)
                QApplication.processEvents()

                for index, file in enumerate(files):
                    if not file.lower().endswith(self.IMAGE_FORMATS):
                        # show_message_box("文件格式错误，仅支持对图片文件进行无损旋转", "提示", 500)
                        # progress_dialog.setVisible(False)
                        print(f"函数jpg_lossless_rotator:{os.path.basename(file)}文件格式错误，仅支持对图片文件进行无损旋转")
                        progress_dialog.setValue(index + 1)
                        continue                    

                    # 拼接参数命令字符串
                    command = f"{jpegr_path} -{para} -s \"{file}\""

                    # 调用jpegoptim命令并捕获返回值
                    result = subprocess.run(command, check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)

                    # 检查返回码
                    if result.returncode == 0:
                        # 更新进度条
                        progress_dialog.setValue(index + 1)
                        if progress_dialog.wasCanceled():
                            show_message_box(f"用户手动取消无损旋转操作，\n已无损旋转前{index+1}张图,共{len(files)}张", "提示", 3000)
                            break  # 如果用户取消了操作，则退出循环
                    else:
                        print("命令执行失败，返回码:", result.returncode)
                        print("错误信息:", result.stderr)
                        return
                
                # 添加进度条完成后的销毁逻辑
                progress_dialog.finished.connect(progress_dialog.deleteLater)  # 进度条完成时销毁                

                # 提示信息
                show_message_box(f"选中的{len(files)}张图片已完成无损旋转", "提示", 1000)

                # 清图标缓存，刷新表格
                IconCache.clear_cache()

                # 更新表格
                self.update_RB_QTableWidget0() 

        except subprocess.CalledProcessError as e:
            print(f"jpg_lossless_rotator()-error--无损旋转图片失败: {e}")
            return


    def copy_selected_file_path(self,flag=1):
        """复制所有选中的单元格的文件路径到系统粘贴板"""
        selected_items = self.RB_QTableWidget0.selectedItems()  # 获取选中的项
        if not selected_items:
            show_message_box("没有选中的项！", "提示", 500)
            return
        
        # 用于存储所有选中的文件路径
        file_paths = []  
        try:
            for item in selected_items:
                row = item.row()
                col = item.column()

                # 构建文件完整路径
                file_name = self.RB_QTableWidget0.item(row, col).text().split('\n')[0]  # 获取文件名
                column_name = self.RB_QTableWidget0.horizontalHeaderItem(col).text()  # 获取列名
                current_directory = self.RT_QComboBox.currentText()  # 获取当前选中的目录
                # 移除传统构建路径方法
                # full_path = os.path.join(os.path.dirname(current_directory), column_name, file_name)
                # 使用 Path 构建路径，自动处理跨平台的路径问题
                full_path = str(Path(current_directory).parent / column_name / file_name)

                if os.path.isfile(full_path):
                    file_paths.append(full_path)  # 添加有效文件路径到列表

            if file_paths:
                # 将文件路径复制到剪贴板，使用换行符分隔
                clipboard_text = "\n".join(file_paths)
                clipboard = QApplication.clipboard()
                clipboard.setText(clipboard_text)

                if flag:
                    show_message_box(f"{len(file_paths)} 个文件的路径已复制到剪贴板", "提示", 2000)
                else:
                    return file_paths
            else:
                show_message_box("没有有效的文件路径", "提示", 2000)

        except Exception as e:
            print(f"copy_selected_file_path()-error--复制文件路径失败: {e}")
            return


    def copy_selected_files(self):
        """复制选中的单元格对应的所有文件到系统剪贴板"""
        selected_items = self.RB_QTableWidget0.selectedItems()  # 获取选中的项
        if not selected_items:
            show_message_box("没有选中的项！", "提示", 500)
            return

        # 用于存储所有选中的文件路径
        file_paths = []  
        try:
            for item in selected_items:
                row = item.row()
                col = item.column()

                # 构建文件完整路径
                file_name = self.RB_QTableWidget0.item(row, col).text().split('\n')[0]  # 获取文件名
                column_name = self.RB_QTableWidget0.horizontalHeaderItem(col).text()  # 获取列名
                current_directory = self.RT_QComboBox.currentText()  # 获取当前选中的目录
                full_path = str(Path(current_directory).parent / column_name / file_name)

                if os.path.isfile(full_path):
                    file_paths.append(full_path)  # 添加有效文件路径到列表

            if file_paths:
                # 创建QMimeData对象
                mime_data = QtCore.QMimeData()
                mime_data.setUrls([QtCore.QUrl.fromLocalFile(path) for path in file_paths])  # 设置文件路径

                # 将QMimeData放入剪贴板
                clipboard = QApplication.clipboard()
                clipboard.setMimeData(mime_data)

                show_message_box(f"{len(file_paths)} 个文件已复制到剪贴板", "提示", 2000)
            else:
                show_message_box("没有有效的文件路径", "提示", 2000)

        except Exception as e:
            print(f"copy_selected_files()-error--复制文件失败: {e}")
            return


    def delete_from_list(self):
        """从列表中删除选中的单元格"""
        print(f"delete_from_list()-从列表中删除选中的单元格")

        selected_items = self.RB_QTableWidget0.selectedItems()
        if not selected_items:
            show_message_box("没有选中的项！", "提示", 500)
            return
        
        # 收集要删除的项目信息
        items_to_delete = []
        try:
            for item in selected_items:
                col = item.column()
                row = item.row()
                file_name = self.RB_QTableWidget0.item(row, col).text().split('\n')[0].strip()
                
                # 获取对应列的文件夹名称
                column_name = self.RB_QTableWidget0.horizontalHeaderItem(col).text()
                
                # 在paths_list中查找对应的索引
                col_idx = self.dirnames_list.index(column_name) if column_name in self.dirnames_list else -1
                
                if col_idx != -1 and row < len(self.paths_list[col_idx]):
                    # 验证文件名是否完全匹配
                    path_file_name = os.path.basename(self.paths_list[col_idx][row])
                    if file_name == path_file_name:
                        items_to_delete.append((col_idx, row))
            
            # 按列和行的逆序排序，确保删除时不会影响其他项的索引
            items_to_delete.sort(reverse=True)
            
            # 执行删除操作
            for col_idx, row in items_to_delete:
                if col_idx < len(self.files_list) and row < len(self.files_list[col_idx]):
                    del self.files_list[col_idx][row]
                    del self.paths_list[col_idx][row]
            
            # 更新表格显示
            self.update_RB_QTableWidget0_from_list(self.files_list, self.paths_list, self.dirnames_list)
    
        except Exception as e:
            print(f"delete_from_list()-error--删除失败: {e}")
            return

    def delete_from_file(self):
        """从源文件删除选中的单元格并删除原文件"""
        print(f"delete_from_file()-从原文件删除选中的单元格并删除原文件")

        selected_items = self.RB_QTableWidget0.selectedItems()  # 获取选中的项
        if not selected_items:
            show_message_box("没有选中的项！", "提示", 500)
            return
        # 收集要删除的文件路径
        file_paths_to_delete = []
        try:
            for item in selected_items:
                row = item.row()
                col = item.column()
                file_name = self.RB_QTableWidget0.item(row, col).text().split('\n')[0]  # 获取文件名
                column_name = self.RB_QTableWidget0.horizontalHeaderItem(col).text()  # 获取列名
                current_directory = self.RT_QComboBox.currentText()  # 获取当前选中的目录
                full_path = str(Path(current_directory).parent / column_name / file_name)

                if os.path.isfile(full_path):
                    file_paths_to_delete.append(full_path)  # 添加有效文件路径到列表

            # 删除文件
            for file_path in file_paths_to_delete:
                try:
                    os.remove(file_path)  # 删除文件
                except Exception as e:
                    show_message_box(f"删除文件失败: {file_path}, 错误: {e}", "提示", 500)

            # 删除表格中的行，可以直接更新表格
            self.update_RB_QTableWidget0()
            show_message_box(f"{len(file_paths_to_delete)} 个文件已从列表中删除并删除原文件", "提示", 1000)

        except Exception as e:
            print(f"delete_from_file()-error--删除失败: {e}")
            return


    def compress_selected_files(self):
        """压缩选中的文件并复制压缩包文件到剪贴板"""
        print("compress_selected_files()-启动压缩文件任务")
        try:
            selected_items = self.RB_QTableWidget0.selectedItems()
            if not selected_items:
                show_message_box("没有选中的项！", "提示", 500)
                return

            # 获取压缩包名称
            zip_name, ok = QtWidgets.QInputDialog.getText(self, "输入压缩包名称", "请输入压缩包名称（不带扩展名）:")
            if not ok or not zip_name:
                show_message_box("未输入有效的名称！", "提示", 500)
                return

            # 准备要压缩的文件列表
            files_to_compress = []
            current_directory = self.RT_QComboBox.currentText()
        
            for item in selected_items:
                row = item.row()
                col = item.column()
                file_name = self.RB_QTableWidget0.item(row, col).text().split('\n')[0]
                column_name = self.RB_QTableWidget0.horizontalHeaderItem(col).text()
                full_path = str(Path(current_directory).parent / column_name / file_name)
                
                if os.path.isfile(full_path):
                    files_to_compress.append((full_path, file_name))

            if not files_to_compress:
                show_message_box("没有有效的文件可压缩", "提示", 500)
                return

            # 设置压缩包路径
            cache_dir = os.path.join(os.path.dirname(__file__), "cache")
            os.makedirs(cache_dir, exist_ok=True)
            self.zip_path = os.path.join(cache_dir, f"{zip_name}.zip")

            # 创建并启动压缩工作线程
            self.compress_worker = CompressWorker(files_to_compress, self.zip_path)
            
            # 连接信号
            self.compress_worker.signals.progress.connect(self.on_compress_progress)
            self.compress_worker.signals.finished.connect(self.on_compress_finished)
            self.compress_worker.signals.error.connect(self.on_compress_error)
            self.compress_worker.signals.cancel.connect(self.cancel_compression)

            # 显示进度窗口
            self.progress_dialog = ProgressDialog(self)
            self.progress_dialog.show()

            # 启动压缩任务
            self.threadpool.start(self.compress_worker)

        except Exception as e:
            print(f"compress_selected_files()-error--压缩失败: {e}")
            return  

    def tcp_tool(self):
        """打开TCP工具,移除tcp功能,替换为截图功能"""
        try:
            
            # 调用截图工具
            WScreenshot.run()

            if False:
                tcp_path = os.path.join(os.path.join(os.path.dirname(__file__), "tools"), "tcp.exe")
                if not os.path.isfile(tcp_path):
                    show_message_box(f"未找到TCP工具: {tcp_path}", "错误", 1500)
                    return
                # 使用startfile保持窗口可见（适用于GUI程序）
                # 该方法只适用于window系统，其余系统（mac,linux）需要通过subprocess实现
                os.startfile(tcp_path)
            
        except Exception as e:
            # show_message_box(f"启动TCP工具失败: {str(e)}", "错误", 2000)
            show_message_box(f"启动截图功能失败: {str(e)}", "错误", 2000)

    def jpgc_tool(self):
        """打开图片体积压缩工具_升级版"""
        try:
            tools_dir = os.path.join(os.path.dirname(__file__), "tools")
            tcp_path = os.path.join(tools_dir, "JPGC.exe")
            
            if not os.path.isfile(tcp_path):
                show_message_box(f"未找到JPGC工具: {tcp_path}", "错误", 1500)
                return
                
            # 使用startfile保持窗口可见（适用于GUI程序）
            # 该方法只适用于window系统，其余系统（mac,linux）需要通过subprocess实现
            os.startfile(tcp_path)
            
        except Exception as e:
            show_message_box(f"启动JPGC工具失败: {str(e)}", "错误", 2000)


    def reveal_in_explorer(self):
        """在资源管理器中高亮定位选中的文件（跨平台优化版）"""
        try:
            # 获取首个选中项（优化性能，避免处理多选）
            if not (selected := self.RB_QTableWidget0.selectedItems()):
                show_message_box("请先选择要定位的文件", "提示", 1000)
                return

            # 缓存路径对象避免重复计算
            current_dir = Path(self.RT_QComboBox.currentText()).resolve()
            item = selected[0]
            
            # 直接获取列名（避免多次调用horizontalHeaderItem）
            if not (col_name := self.RB_QTableWidget0.horizontalHeaderItem(item.column()).text()):
                raise ValueError("无效的列名")
            col_name = self.RB_QTableWidget0.horizontalHeaderItem(item.column()).text()

            # 强化路径处理
            file_name = item.text().split('\n', 1)[0].strip()  # 移除前后空格
            full_path = (current_dir.parent / col_name / file_name).resolve()

            if not full_path.exists():
                show_message_box(f"文件不存在: {full_path.name}", "错误", 1500)
                return

            # 跨平台处理优化
            if sys.platform == 'win32':
                # 转换为Windows风格路径并处理特殊字符
                win_path = str(full_path).replace('/', '\\')
                if ' ' in win_path:  # 自动添加双引号
                    win_path = f'"{win_path}"'
                # 使用start命令更可靠
                command = f'start explorer /select,{win_path}'
                # 移除check=True参数避免误报
                subprocess.run(command, shell=True, creationflags=subprocess.CREATE_NO_WINDOW)

            elif sys.platform == 'darwin':
                # 使用open命令直接定位文件
                subprocess.run(['open', '-R', str(full_path)], check=True)

            else:  # Linux/Unix
                subprocess.run(['xdg-open', str(full_path.parent)], check=True)

        except subprocess.CalledProcessError as e:
            show_message_box(f"定位命令执行失败: {str(e)}", "错误", 2000)
        except FileNotFoundError:
            show_message_box("找不到系统命令，请检查系统环境", "错误", 2000)
        except Exception as e:
            show_message_box(f"定位文件失败: {str(e)}", "错误", 2000)


    def on_compress_progress(self, current, total):
        """处理压缩进度"""
        progress_value = int((current / total) * 100)  # 计算进度百分比
        self.progress_dialog.update_progress(progress_value)
        self.progress_dialog.set_message(f"显示详情：正在压缩文件... {current}/{total}")

    def cancel_compression(self):
        """取消压缩任务"""
        if self.compress_worker:
            self.compress_worker.cancel()  # 假设CompressWorker有一个cancel方法
        self.progress_dialog.close()  # 关闭进度窗口
        show_message_box("压缩已取消", "提示", 500)

        # 若是压缩取消，则删除缓存文件中的zip文件
        cache_dir = os.path.join(os.path.dirname(__file__), "cache")
        if os.path.exists(cache_dir):
            force_delete_folder(cache_dir, '.zip')

    def on_compress_finished(self, zip_path):
        """处理压缩完成"""
        self.progress_dialog.close()  # 关闭进度窗口
        # 将压缩包复制到剪贴板
        mime_data = QtCore.QMimeData()
        url = QtCore.QUrl.fromLocalFile(zip_path)
        mime_data.setUrls([url])
        QApplication.clipboard().setMimeData(mime_data)
        # 更新状态栏信息显示
        self.statusbar_label1.setText(f"🔉: 压缩完成🍃")
        show_message_box(f"文件已压缩为: {zip_path} 并复制到剪贴板", "提示", 500)

    def on_compress_error(self, error_msg):
        """处理压缩错误"""
        self.progress_dialog.close()  # 关闭进度窗口
        # 更新状态栏信息显示
        self.statusbar_label1.setText(f"🔉: 压缩出错🍃")
        show_message_box(error_msg, "错误", 2000)


    """
    自定义功能函数区域：
    拖拽功能函数 self.dragEnterEvent(), self.dropEvent()
    左侧文件浏览器与地址栏联动功能函数 self.locate_in_tree_view, selfupdate_combobox
    右侧表格显示功能函数 self.update_RB_QTableWidget0()
    """


    def dragEnterEvent(self, event):
        # 如果拖入的是文件夹，则接受拖拽
        if event.mimeData().hasUrls():

            event.accept()

    def dropEvent(self, event):
        # 获取拖放的文件夹路径,并插入在首行，方便地查看最近添加的文件夹路径
        for url in event.mimeData().urls():
            folder_path = url.toLocalFile()
            if os.path.isdir(folder_path):  
                self.RT_QComboBox.insertItem(0, folder_path)
                self.RT_QComboBox.setCurrentText(folder_path)
                # 定位到左侧文件浏览器中
                self.locate_in_tree_view()
                # 将同级文件夹添加到 RT_QComboBox1 中
                self.RT_QComboBox1_init() 
                # 更新右侧RB_QTableWidget0表格
                self.update_RB_QTableWidget0() 
                break  
        
    # 点击左侧文件浏览器时的功能函数
    def update_combobox(self, index):
        """左侧文件浏览器点击定位更新右侧combobox函数"""
        print("update_combobox函数: ")

        # 清空历史的已选择
        self.statusbar_label.setText(f"[0]已选择")
        
        # 获取左侧文件浏览器中当前点击的文件夹路径，并显示在地址栏
        current_path = self.file_system_model.filePath(index)
        if os.path.isdir(current_path):
            self.RT_QComboBox.setCurrentText(current_path)
            if self.RT_QComboBox.findText(current_path) == -1:
                self.RT_QComboBox.insertItem(0, current_path)
            print(f"1. 点击了左侧文件，该文件夹已更新到地址栏中: {current_path}")

        print("2. 将地址栏文件夹的同级文件夹更新到下拉复选框")
        # 将同级文件夹添加到 RT_QComboBox1 中
        self.RT_QComboBox1_init()      
        # 更新右侧RB_QTableWidget0表格
        self.update_RB_QTableWidget0() 
        
    # 在左侧文件浏览器中定位地址栏(RT_QComboBox)中当前显示的目录
    def locate_in_tree_view(self):
        """左侧文件浏览器点击定位函数"""
        print("locate_in_tree_view()--在左侧文件浏览器中定位地址栏路径")
        try:
            current_directory = self.RT_QComboBox.currentText()
            # 检查路径是否有效
            if not os.path.exists(current_directory): 
                print("locate_in_tree_view()--地址栏路径不存在")
                return  
            # 获取当前目录的索引
            index = self.file_system_model.index(current_directory)  
            # 检查索引是否有效
            if index.isValid():
                # 设置当前索引
                self.Left_QTreeView.setCurrentIndex(index)    
                # 展开该目录
                self.Left_QTreeView.setExpanded(index, True)  
                # 滚动到该项，确保垂直方向居中
                self.Left_QTreeView.scrollTo(index, QAbstractItemView.PositionAtCenter)
                
                # 手动设置水平方向进度条
                self.Left_QTreeView.horizontalScrollBar().setValue(0)
            
                print(f"locate_in_tree_view()--定位成功")
            else:
                print("locate_in_tree_view()--索引无效-无法定位")

        except Exception as e:
            print(f"locate_in_tree_view()--定位失败: {e}")
            return


    def update_RB_QTableWidget0_from_list(self, file_infos_list, file_paths, dir_name_list):
        """从当前列表中更新表格，适配从当前列表删除文件功能"""
        print(f"update_RB_QTableWidget0_from_list()--从当前列表中更新表格")
        try:    
            # 取消当前的预加载任务
            self.cancel_preloading()
            # 清空表格和缓存
            self.RB_QTableWidget0.clear()
            self.RB_QTableWidget0.setRowCount(0)
            self.RB_QTableWidget0.setColumnCount(0)
            self.image_index_max = [] # 清空图片列有效行最大值 

            # 先初始化表格结构和内容，不加载图标,并获取图片列有效行最大值
            self.image_index_max = self.init_table_structure(file_infos_list, dir_name_list)

            # 对file_paths进行转置,实现加载图标按行加载
            file_name_paths = list(chain.from_iterable(zip_longest(*file_paths, fillvalue=None)))

            if file_name_paths:  # 确保有文件路径才开始预加载
                self.start_image_preloading(file_name_paths)

        except Exception as e:
            print(f"update_RB_QTableWidget0_from_list()-error--从当前列表中更新表格任务失败: {e}")


    def update_RB_QTableWidget0(self):
        """更新右侧表格功能函数"""
        print(f"update_RB_QTableWidget0()--更新右侧表格内容:")
        try:
            try:    
                # 取消当前的预加载任务
                self.cancel_preloading()
            except Exception as e:
                print(f"取消预加载任务失败: {e}")
            
            try:    
                # 清空表格和缓存
                self.RB_QTableWidget0.clear()
                self.RB_QTableWidget0.setRowCount(0)
                self.RB_QTableWidget0.setColumnCount(0)
                self.image_index_max = [] # 清空图片列有效行最大值  
            except Exception as e:
                print(f"清空表格和缓存失败: {e}")
            
            try:
                # 收集文件名基本信息以及文件路径，并将相关信息初始化为类中全局变量
                file_infos_list, file_paths, dir_name_list = self.collect_file_paths()
                # 初始化文件名及基本信息列表
                self.files_list = file_infos_list      
                # 初始化文件路径列表
                self.paths_list = file_paths          
                # 初始化选中的同级文件夹列表
                self.dirnames_list = dir_name_list    

            except Exception as e:
                print(f"收集文件路径失败: {e}")
            
            try:
                # 先初始化表格结构和内容，不加载图标,并获取图片列有效行最大值
                self.image_index_max = self.init_table_structure(file_infos_list, dir_name_list)
            except Exception as e:
                print(f"初始化表格结构和内容失败: {e}")

            try:
                """对file_paths进行转置,实现加载图标按行加载"""
                file_name_paths = list(chain.from_iterable(zip_longest(*file_paths, fillvalue=None)))

                # 初始化预加载图标线程前的问价排列列表
                self.preloading_file_name_paths = file_name_paths     

            except Exception as e:
                print(f"处理文件路径失败: {e}")

            try:    
                # 开始预加载图标    
                if file_name_paths:  # 确保有文件路径才开始预加载
                    self.start_image_preloading(file_name_paths)
            except Exception as e:
                print(f"开始预加载图标失败: {e}")

        except Exception as e:
            # 返回错误信息
            print(f"update_RB_QTableWidget0()-error--更新右侧表格失败！\n错误信息: {e}")
            return


    def init_table_structure(self, file_name_list, dir_name_list):
        """初始化表格结构和内容，不包含图标"""

        # 设置表格的列数
        self.RB_QTableWidget0.setColumnCount(len(file_name_list))
        # 设置列标题为当前选中的文件夹名，设置列名为文件夹名
        self.RB_QTableWidget0.setHorizontalHeaderLabels(dir_name_list)  

        # 判断是否存在文件
        if not file_name_list or not file_name_list[0]:
            return []  
        
        # 设置表格的行数
        max_cols = max(len(row) for row in file_name_list) 
        self.RB_QTableWidget0.setRowCount(max_cols)  
        self.RB_QTableWidget0.setIconSize(QSize(48, 48))  

        pic_num_list = [] # 用于记录每列的图片数量
        flag_ = 0 # 用于记录是否需要设置固定行高
        # 填充 QTableWidget,先填充文件名称，后填充图标(用多线程的方式后加载图标)
        for col_index, row in enumerate(file_name_list):
            pic_num_list.append(len(row))
            for row_index, value in enumerate(row):
                if value[4][0] is None and value[4][1] is None:
                    resolution = " "
                else:
                    resolution = f"{value[4][0]}x{value[4][1]}"
                if value[5] is None:
                    exposure_time = " "
                else:
                    exposure_time = value[5]
                if value[6] is None:
                    iso = " "
                else: 
                    iso = value[6]
                # 文件名称、分辨率、曝光时间、ISO
                if resolution == " " and exposure_time == " " and iso == " ": 
                    item_text = value[0]
                    flag_ = 0 
                else:
                    item_text = value[0] + "\n" + f"{resolution} {exposure_time} {iso}"
                    flag_ = 1 # 设置flag_为1，设置行高
                item = QTableWidgetItem(item_text)
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)  # 禁止编辑
                self.RB_QTableWidget0.setItem(row_index, col_index, item)  # 设置单元格项
            ###############################    列  ,     行   ，内容    ######################

        # 设置单元格行高固定为60,如果flag_为0，则不设置行高
        if flag_:
            for row in range(self.RB_QTableWidget0.rowCount()):
                self.RB_QTableWidget0.setRowHeight(row, 60)
        else:
            for row in range(self.RB_QTableWidget0.rowCount()):
                self.RB_QTableWidget0.setRowHeight(row, 52)


        # # 更新标签显示  
        self.statusbar_label0.setText(f"📢:当前选中的文件夹中包含 {pic_num_list} 张图")  

        return pic_num_list

        
    def collect_file_paths(self):
        """收集需要显示的文件路径"""
        # 初始化列表
        file_infos = []  # 文件名列表
        file_paths = []  # 文件路径列表
        dir_name_list = [] # 文件夹名列表

        try:
            # 获取复选框中选择的文件夹路径列表
            selected_folders = self.model.getCheckedItems()  # 获取选中的文件夹
            current_directory = self.RT_QComboBox.currentText() # 当前选中的文件夹目录 
            parent_directory = os.path.dirname(current_directory)  # 获取父目录
            
            # 构建所有需要显示的文件夹路径
            selected_folders_path = [os.path.join(parent_directory, path) for path in selected_folders]
            selected_folders_path.insert(0, current_directory)  # 将当前选中的文件夹路径插入到列表的最前面
            
            # 获取文件夹名列表
            dir_name_list = [os.path.basename(dir_name) for dir_name in selected_folders_path]
            
            # 处理每个文件夹
            for folder in selected_folders_path:
                if not os.path.exists(folder):
                    continue
                    
                file_name_list, file_path_list = self.filter_files(folder)
                if file_name_list:  # 只添加非空列表
                    file_infos.append(file_name_list)
                    file_paths.append(file_path_list)
                
            return file_infos, file_paths, dir_name_list
            
        except Exception as e:
            print(f"collect_file_paths函数_收集文件路径失败: {e}")
            return [], [], []
        
    def filter_files(self, folder):
        """根据选项过滤文件"""
        files_and_dirs_with_mtime = [] 
        selected_option = self.RT_QComboBox0.currentText()
        sort_option = self.RT_QComboBox2.currentText()

        # 使用 os.scandir() 获取文件夹中的条目
        with os.scandir(folder) as entries:
            # 使用列表推导式和 DirEntry 对象的 stat() 方法获取文件元组，比os.listdir()更高效,性能更高
            for entry in entries:
                if entry.is_file():
                    if selected_option == "显示图片文件":
                        if entry.name.lower().endswith(self.IMAGE_FORMATS):

                            if self.simple_mode:
                                width = None   # 宽度  
                                height = None  # 高度
                                exposure_time = None  # 曝光时间
                                iso = None  # ISO
                            else:       
                                # 获取图片的元数据
                                try:
                                    with Image.open(entry.path) as img:
                                        width, height = img.size  # 获取分辨率
                                        exif_data = img._getexif()  # 获取EXIF数据
                                        if exif_data is None:
                                            # 设置默认值
                                            exposure_time = None
                                            iso = None
                                        else:
                                            # 读取ISO
                                            iso = exif_data.get(34855)  # ISO
                                            if iso is None:
                                                iso = None
                                            # 读取EXP
                                            exposure_time_ = exif_data.get(33434)  # 曝光时间
                                            if exposure_time_ is None:
                                                exposure_time = None
                                            elif isinstance(exposure_time_, tuple) and len(exposure_time_) == 2 and exposure_time_[1] != 0:
                                                exposure_time = f"{exposure_time_[0]}/{exposure_time_[1]}"
                                            elif isinstance(exposure_time_, (int, float)):
                                                try:    
                                                    fraction = Fraction(exposure_time_)
                                                    exposure_time = f"{fraction.numerator}/{fraction.denominator}"
                                                except Exception:
                                                    exposure_time = str(exposure_time_)
                                            elif hasattr(exposure_time_, 'numerator') and hasattr(exposure_time_, 'denominator'):
                                                try:    
                                                    fraction = Fraction(exposure_time_.numerator, exposure_time_.denominator)
                                                    exposure_time = f"{fraction.numerator}/{fraction.denominator}"
                                                except Exception:
                                                    exposure_time = None
                                            elif isinstance(exposure_time_, str):
                                                try:
                                                    fraction = Fraction(exposure_time_)
                                                    exposure_time = f"{fraction.numerator}/{fraction.denominator}"
                                                except Exception:
                                                    exposure_time = exposure_time_  

                                            # 处理曝光时间，确保分母为1
                                            if exposure_time:        
                                                if exposure_time.split('/')[0] == '1':
                                                    pass
                                                else:
                                                    if exposure_time.split('/')[0] != '0':
                                                        t_ = 1
                                                        b_ = int(exposure_time.split('/')[1]) // int(exposure_time.split('/')[0])   
                                                        exposure_time = f"{t_}/{b_}"

                                except Exception as e:
                                    print(f"读取图片元数据失败: {entry.path}, 错误: {e}")
                                    # 设置默认值
                                    exposure_time = None
                                    iso = None
                            
                            # 文件名称、创建时间、修改时间、文件大小、分辨率、曝光时间、ISO、文件路径
                            files_and_dirs_with_mtime.append((entry.name, entry.stat().st_ctime, entry.stat().st_mtime, 
                                                          entry.stat().st_size, (width, height), 
                                                          exposure_time, iso, entry.path))
                        else:
                            continue
                    elif selected_option == "显示视频文件":
                        if entry.name.lower().endswith(self.VIDEO_FORMATS):     
                            # 文件名称、创建时间、修改时间、文件大小、分辨率、曝光时间、ISO、文件路径
                            files_and_dirs_with_mtime.append((entry.name, entry.stat().st_ctime, entry.stat().st_mtime, 
                                                          entry.stat().st_size, (None, None), 
                                                          None, None, entry.path))
                        else:
                            continue
                    elif selected_option == "显示所有文件":
                            # 文件名称、创建时间、修改时间、文件大小、分辨率、曝光时间、ISO、文件路径
                            files_and_dirs_with_mtime.append((entry.name, entry.stat().st_ctime, entry.stat().st_mtime, 
                                                          entry.stat().st_size, (None, None), 
                                                          None, None, entry.path))
                    else: # 没有选择任何选项就跳过
                        print("filter_files函数:selected_option没有选择任何选项,跳过")
                        continue

        # 排序
        if sort_option == "按创建时间排序":  # 按创建时间排序，reverse=False 表示升序，即最小的在前面
            files_and_dirs_with_mtime.sort(key=lambda x: x[1], reverse=False)
        elif sort_option == "按修改时间排序":  # 按修改时间排序，reverse=False 表示升序，即最小的在前面
            files_and_dirs_with_mtime.sort(key=lambda x: x[2], reverse=False)
        elif sort_option == "按文件大小排序":  # 按文件大小排序，reverse=False 表示升序，即最小的在前面
            files_and_dirs_with_mtime.sort(key=lambda x: x[3], reverse=False)
        elif sort_option == "按文件名称排序":  # 按文件名称排序，reverse=False 表示升序，即最小的在前面
            # files_and_dirs_with_mtime.sort(key=lambda x: x[0], reverse=False)
            # modify by diamond_cz 修改按文件名排序功能，实现类似window支持数字排序
            files_and_dirs_with_mtime.sort(key=lambda x: natural_sort_key(x[0]), reverse=False)
        elif sort_option == "按创建时间逆序排序":  # 按创建时间排序，reverse=True 表示降序，即最大的在前面
            files_and_dirs_with_mtime.sort(key=lambda x: x[1], reverse=True)
        elif sort_option == "按修改时间逆序排序":  # 按修改时间排序，reverse=True 表示降序，即最大的在前面
            files_and_dirs_with_mtime.sort(key=lambda x: x[2], reverse=True)
        elif sort_option == "按文件大小逆序排序":  # 按文件大小排序，reverse=True 表示降序，即最大的在前面
            files_and_dirs_with_mtime.sort(key=lambda x: x[3], reverse=True)
        elif sort_option == "按文件名称逆序排序":  # 按文件名称排序，reverse=True 表示降序，即最大的在前面
            # files_and_dirs_with_mtime.sort(key=lambda x: x[0], reverse=True)
            files_and_dirs_with_mtime.sort(key=lambda x: natural_sort_key(x[0]), reverse=True)
        # 极简模式下不使能曝光、ISO排序选项
        elif not self.simple_mode and sort_option == "按曝光时间排序" and selected_option == "显示图片文件":  # 按曝光时间排序，reverse=False 表示升序，即最小的在前面
            # 排序中若存在None,则提供默认值0  
            files_and_dirs_with_mtime.sort(key=lambda x: int(x[5].split('/')[1]) if x[5] is not None else 0, reverse=False)
        elif not self.simple_mode and sort_option == "按ISO排序" and selected_option == "显示图片文件":  # 按ISO排序，reverse=False 表示升序，即最小的在前面
            # 排序中若存在None,则提供默认值0  
            files_and_dirs_with_mtime.sort(key=lambda x: int(x[5].split('/')[1]) if x[5] is not None else 0, reverse=False)
        elif not self.simple_mode and sort_option == "按曝光时间逆序排序" and selected_option == "显示图片文件":  # 按曝光时间排序，reverse=True 表示降序，即最大的在前面
            # 排序中若存在None,则提供默认值0  
            files_and_dirs_with_mtime.sort(key=lambda x: int(x[5].split('/')[1]) if x[5] is not None else 0, reverse=False)
        elif not self.simple_mode and sort_option == "按ISO逆序排序" and selected_option == "显示图片文件":  # 按ISO排序，reverse=True 表示降序，即最大的在前面
            # 排序中若存在None,则提供默认值0  
            files_and_dirs_with_mtime.sort(key=lambda x: int(x[5].split('/')[1]) if x[5] is not None else 0, reverse=False) 
        else:  # 默认按创建时间排序，reverse=False 表示升序，即最小的在前面
            files_and_dirs_with_mtime.sort(key=lambda x: x[1], reverse=False)

        # 获取文件路径列表，files_and_dirs_with_mtime的最后一列
        file_paths = [item[-1] for item in files_and_dirs_with_mtime]

        return files_and_dirs_with_mtime, file_paths

        
    def start_image_preloading(self, file_paths):
        """开始预加载图片"""
        if self.preloading:
            return
        
        # 设置预加载状态
        self.preloading = True
        
        
        # 创建新的预加载器
        self.current_preloader = ImagePreloader(file_paths)
        self.current_preloader.signals.progress.connect(self.update_preload_progress)
        self.current_preloader.signals.batch_loaded.connect(self.on_batch_loaded)
        self.current_preloader.signals.finished.connect(self.on_preload_finished)
        self.current_preloader.signals.error.connect(self.on_preload_error)
        
        # 启动预加载
        self.threadpool.start(self.current_preloader)

        print("start_image_preloading函数: 开始预加载图标, 启动预加载线程")
        
    def cancel_preloading(self):
        """取消当前预加载任务"""
        try:
            if self.current_preloader and self.preloading:
                self.current_preloader._stop = True  # 使用 _stop 属性而不是 stop() 方法
                self.preloading = False
                self.current_preloader = None
                
        except Exception as e:
            print(f"取消预加载任务失败: {e}")

    def on_batch_loaded(self, batch):
        """处理批量加载完成的图标"""
        for path, icon in batch:
            # 更新表格中对应的图标
            self.update_table_icon(path, icon)
            
    def update_table_icon(self, file_path, icon):
        """更新表格中的指定图标
        通过先查找行来优化图标更新效率
        """
        filename = os.path.basename(file_path)
        folder = os.path.basename(os.path.dirname(file_path))
        
        # 先在每一行中查找文件名
        for row in range(self.RB_QTableWidget0.rowCount()):
            # 遍历每一列查找匹配的文件夹
            for col in range(self.RB_QTableWidget0.columnCount()):
                header = self.RB_QTableWidget0.horizontalHeaderItem(col)
                item = self.RB_QTableWidget0.item(row, col)
                
                if (header and header.text() == folder and 
                    item and item.text().split('\n')[0] == filename):
                    item.setIcon(icon)
                    return  # 找到并更新后直接返回

    def update_preload_progress(self, current, total):
        """处理预加载进度"""
        # 更新状态栏信息显示
        self.statusbar_label1.setText(f"🔉: 图标加载进度...{current}/{total}🍃")
        
    def on_preload_finished(self):
        """处理预加载完成"""
        print("on_preload_finished()--图标预加载完成")
        # 更新状态栏信息显示
        self.statusbar_label1.setText(f"🔉: 图标已全部加载🍃")
        gc.collect()
        
    def on_preload_error(self, error):
        """处理预加载错误"""
        print(f"on_preload_error()--图标预加载错误: {error}")

    def RT_QComboBox1_init(self):
        """自定义RT_QComboBox1, 添加复选框选项"""
        print("RT_QComboBox1_init()--开始添加地址栏文件夹的同级文件夹到下拉复选框中")
        try:
            # 获取地址栏当前路径    
            current_directory = self.RT_QComboBox.currentText()
            # 检查路径是否有效
            if not os.path.exists(current_directory): 
                print("RT_QComboBox1_init()--地址栏路径不存在")
                return  
            # 获取父目录中的文件夹列表
            sibling_folders = self.getSiblingFolders(current_directory)  
            # 使用文件夹列表和父目录初始化模型
            self.model = CheckBoxListModel(sibling_folders)  
            # 绑定模型到 QComboBox
            self.RT_QComboBox1.setModel(self.model)  
            # 设置自定义委托
            self.RT_QComboBox1.setItemDelegate(CheckBoxDelegate())  
            # 禁用右键菜单
            self.RT_QComboBox1.setContextMenuPolicy(Qt.NoContextMenu)  
        except Exception as e:
            print(f"RT_QComboBox1_init()--初始化失败: {e}")

    def handleComboBoxPressed(self, index):
        """处理复选框选项被按下时的事件。"""
        print("handleComboBoxPressed()--更新复选框状态")
        try:
            if not index.isValid():
                print("handleComboBoxPressed()--下拉复选框点击无效")
                return
            self.model.setChecked(index)  # 更新复选框的状态
        except Exception as e:
            print(f"handleComboBoxPressed()--更新复选框状态失败: {e}")

    def handleComboBox0Pressed(self):
        """处理（显示图片视频所有文件）下拉框选项被按下时的事件。"""
        print("handleComboBox0Pressed()--更新（显示图片视频所有文件）下拉框状态")
        # self.locate_in_tree_view() # 定位到左侧文件浏览器中
        # self.RT_QComboBox1_init() # 将同级文件夹添加到 RT_QComboBox1 中
        self.update_RB_QTableWidget0() # 更新右侧RB_QTableWidget0表格

    def updateComboBox1Text(self):
        """更新 RT_QComboBox1 的显示文本。"""    
        print("updateComboBox1Text()--下拉框中的复选框被选中了，更新显示文本")
        try:
            selected_folders = self.model.getCheckedItems()  # 获取选中的文件夹
            current_text = '; '.join(selected_folders) if selected_folders else "(请选择)"
            self.RT_QComboBox1.setCurrentText(current_text)  # 更新 ComboBox 中的内容
            # 更新表格内容
            self.update_RB_QTableWidget0()  
        except Exception as e:
            print(f"updateComboBox1Text()--更新显示文本失败: {e}")

    def getSiblingFolders(self, folder_path):
        """获取指定文件夹的同级文件夹列表。"""
        print(f"getSiblingFolders()--获取{folder_path}的同级文件夹列表")
        try:
            parent_folder = os.path.dirname(folder_path)  # 获取父文件夹路径
            return [
                name for name in os.listdir(parent_folder)
                    if os.path.isdir(os.path.join(parent_folder, name)) and name != os.path.basename(folder_path)  # 过滤出同级文件夹，不包括当前选择的文件夹
                ]
        except Exception as e:
            print(f"getSiblingFolders()--获取同级文件夹列表失败: {e}")
            return []

    def handle_table_selection(self):
        """处理表格选中事件"""
        selected = self.RB_QTableWidget0.selectedItems()
        if selected:
            # 示例：更新状态栏显示选中数量
            # count = len({item.row() for item in selected})  # 获取不重复的行数
            self.statusbar_label.setText(f"[{len(selected)}]已选择")
            
            # 可以在这里添加更多选中后的处理逻辑
            # 例如：显示选中文件的预览、获取文件详细信息等

    def handle_sort_option(self):
        """处理排序选项"""
        print("handle_sort_option()--处理排序选项")
        try:
            sort_option = self.RT_QComboBox2.currentText()
            if self.simple_mode:
                if sort_option == "按曝光时间排序":
                    # 弹出提示框
                    show_message_box("极简模式下不使能曝光时间排序", "提示", 500)
                    return
                elif sort_option == "按ISO排序":
                    # 弹出提示框    
                    show_message_box("极简模式下不使能ISO排序", "提示", 500)
                    return

            self.update_RB_QTableWidget0()  # 更新右侧表格
        except Exception as e:
            print(f"handle_sort_option()--处理排序选项失败: {e}")

    def handle_theme_selection(self):
        """处理下拉框选择"""
        # 获取下拉框的当前选择
        print("handle_theme_selection()--处理下拉框选择")
        try:
            selected_theme = self.RT_QComboBox3.currentText()
            if selected_theme == "默认主题":
                self.current_theme = "默认主题"
            elif selected_theme == "暗黑主题":  # 修改为 "暗黑主题"
                self.current_theme = "暗黑主题"
            
            # 更新主题
            self.apply_theme()
        except Exception as e:
            print(f"handle_theme_selection()--处理下拉框选择失败: {e}")

    def toggle_theme(self):
        """切换主题"""
        print("toggle_theme()--切换主题")
        try:
            if self.current_theme == "默认主题":
                self.current_theme = "暗黑主题"
            else:
                self.current_theme = "默认主题"

            # 更新主题
            self.apply_theme()
        except Exception as e:
            print(f"toggle_theme()--切换主题失败: {e}")

    def apply_theme(self):
        """初始化主题"""
        print("apply_theme()--更新当前主题")
        try:
            if self.current_theme == "暗黑主题":
                self.setStyleSheet(self.dark_style())     # 暗黑主题
            else:
                self.setStyleSheet(self.default_style())  # 默认主题
        except Exception as e:
            print(f"apply_theme()--应用当前主题失败: {e}")


    def default_style(self):
        """返回默认模式的样式表"""

        # 定义通用颜色变量
        BACKCOLOR = self.background_color_default  # 浅蓝色背景
        FONTCOLOR = self.font_color_default        # 默认字体颜色
        GRAY = "rgb(127, 127, 127)"                # 灰色
        WHITE = "rgb(238,238,238)"                 # 白色
        QCOMBox_BACKCOLOR = "rgb(255,242,223)"     # 下拉框背景色

        
        table_style = f"""
            QTableWidget#RB_QTableWidget0 {{
                /* 表格整体样式 */
                background-color: {GRAY};
                color: {FONTCOLOR};
            }}
            
            QTableWidget#RB_QTableWidget0::item {{
                /* 单元格样式 */
                background-color: {GRAY};
                color: {FONTCOLOR};
            }}
            
            QTableWidget#RB_QTableWidget0::item:selected {{
                /* 选中单元格样式 */
                background-color: {BACKCOLOR};
                color: {FONTCOLOR};
            }}
            
            /* 添加表头样式 */
            QHeaderView::section {{
                background-color: {BACKCOLOR};
                color: {FONTCOLOR};
                text-align: center;
                padding: 3px;
                margin: 1px;
                font-family: "{self.custom_font.family()}";
                font-size: {self.custom_font.pointSize()}pt;
            }}
            
            /* 修改左上角区域样式 */
            QTableWidget#RB_QTableWidget0::corner {{
                background-color: {BACKCOLOR};  /* 设置左上角背景色 */
                color: {FONTCOLOR};
            }}
        """
        left_qframe_style = f"""
            QFrame#Left_QFrame {{ 
                background-color: {GRAY};
                color: {FONTCOLOR};
                border-radius: 10px;
                border: 1px solid {GRAY};
            }}
        """

        # 按钮组件和复选框组件样式
        button_style = f"""
            QPushButton {{
                background-color: {WHITE};
                color: {FONTCOLOR};
                text-align: center;
                font-family: "{self.custom_font.family()}";
                font-size: {self.custom_font.pointSize()}pt;
            }}
            QPushButton:hover {{
                border: 1px solid {BACKCOLOR};
                background-color: {BACKCOLOR};
                color: {FONTCOLOR};
            }}
        """

        # 设置单选按钮样式
        radio_button_style = f"""   
            QRadioButton {{
                text-align: left;
                color: {FONTCOLOR};
                font-family: "{self.custom_font.family()}";
                font-size: {self.custom_font.pointSize()}pt;
            }}
        """

        # 左侧文件浏览区域样式
        left_area_style = f"""
            QTreeView#Left_QTreeView {{
                background-color: {BACKCOLOR};
                color: {FONTCOLOR};
                border-radius: 10px;
            }}
        """
        
        # 下拉框通用样式模板
        combobox_style = f"""
            QComboBox {{
                /* 下拉框本体样式 */
                background-color: {QCOMBox_BACKCOLOR};
                color: {FONTCOLOR};
                selection-background-color: {BACKCOLOR};
                selection-color: {FONTCOLOR};
                min-height: 30px;
                font-family: "{self.custom_font.family()}";
                font-size: {self.custom_font.pointSize()}pt;
            }}
            
            QComboBox QAbstractItemView {{
                /* 下拉列表样式 */
                background-color: {QCOMBox_BACKCOLOR};
                color: {FONTCOLOR};
                selection-background-color: {BACKCOLOR};
                selection-color: {FONTCOLOR};
                font-family: "{self.custom_font.family()}";
                font-size: {self.custom_font.pointSize()}pt;
            }}
            
            QComboBox QAbstractItemView::item {{
                /* 下拉项样式 */
                min-height: 25px;
                padding: 5px;
                font-family: "{self.custom_font.family()}";
                font-size: {self.custom_font.pointSize()}pt;
            }}

            QComboBox::hover {{
                background-color: {BACKCOLOR};
                color: {FONTCOLOR};
            }}  

        """

        # 下拉框通用样式模板2
        combobox_style2 = f"""
            QComboBox {{
                /* 下拉框本体样式 */
                background-color: {QCOMBox_BACKCOLOR};
                color: {FONTCOLOR};
                selection-background-color: {BACKCOLOR};
                selection-color: {FONTCOLOR};
                min-height: 30px;
                font-family: "{self.custom_font.family()}";
                font-size: {self.custom_font.pointSize()}pt;
            }}
            
            QComboBox QAbstractItemView {{
                /* 下拉列表样式 */
                background-color: {QCOMBox_BACKCOLOR};
                color: {FONTCOLOR};
                selection-background-color: {BACKCOLOR};
                selection-color: {FONTCOLOR};
                font-family: "{self.custom_font.family()}";
                font-size: {self.custom_font.pointSize()}pt;
            }}
            
            QComboBox QAbstractItemView::item {{
                /* 下拉项样式 */
                min-height: 25px;
                padding: 5px;
                font-family: "{self.custom_font.family()}";
                font-size: {self.custom_font.pointSize()}pt;
            }}

        """
        # background-color: {WHITE}; border-radius: 10px;
        statusbar_label_style = f"""
            border: none;
            color: {"rgb(255,255,255)"};
            font-family: {self.custom_font_jetbrains_small.family()};
            font-size: {self.custom_font_jetbrains_small.pointSize()}pt;
        
        """

        statusbar_button_style = f"""
            QPushButton {{
                background-color: {WHITE};
                color: {FONTCOLOR};
                text-align: center;
                font-family: "{self.custom_font_jetbrains_small.family()}";
                font-size: {self.custom_font_jetbrains_small.pointSize()}pt;
                border-radius: 10px;
            }}
            QPushButton:hover {{
                border: 1px solid {BACKCOLOR};
                background-color: {BACKCOLOR};
                color: {FONTCOLOR};
            }}
        """

        statusbar_button_style_version = f"""
            QPushButton {{
                background-color: {"rgb(245,108,108)"};
                color: {FONTCOLOR};
                text-align: center;
                font-family: "{self.custom_font_jetbrains_small.family()}";
                font-size: {self.custom_font_jetbrains_small.pointSize()}pt;
                border-radius: 10px;
            }}
            QPushButton:hover {{
                border: 1px solid {BACKCOLOR};
                background-color: {"rgb(245,108,108)"};
                color: {FONTCOLOR};
            }}
        """        

        # self.custom_font_jetbrains_small   "rgb(234,118, 32)"
        statusbar_style = f"""
            border: none;
            background-color: {GRAY};
            color: {FONTCOLOR};
            font-family: {self.custom_font_jetbrains_small.family()};
            font-size: {self.custom_font_jetbrains_small.pointSize()}pt;
            
        """

        # 设置左上侧文件浏览区域样式
        self.Left_QTreeView.setStyleSheet(left_area_style)

        # 设置左下角侧框架样式
        self.Left_QFrame.setStyleSheet(left_qframe_style)
        self.L_radioButton1.setStyleSheet(radio_button_style)  # 设置左对齐
        self.L_radioButton2.setStyleSheet(radio_button_style)  # 设置左对齐


        # 设置右侧顶部按钮下拉框样式
        self.RT_QPushButton3.setStyleSheet(button_style)
        self.RT_QPushButton5.setStyleSheet(button_style)

        self.RT_QComboBox.setStyleSheet(combobox_style2)
        self.RT_QComboBox1.setStyleSheet(combobox_style2)

        self.RT_QComboBox0.setStyleSheet(combobox_style)
        self.RT_QComboBox2.setStyleSheet(combobox_style)
        self.RT_QComboBox3.setStyleSheet(combobox_style)

        # 设置右侧中间表格区域样式
        self.RB_QTableWidget0.setStyleSheet(table_style)

        # 设置底部状态栏区域样式 self.statusbar --> self.statusbar_widget --> self.statusbar_QHBoxLayout --> self.statusbar_button1 self.statusbar_button2
        self.statusbar.setStyleSheet(statusbar_style)
        self.statusbar_button1.setStyleSheet(statusbar_button_style)
        # 设置版本按钮更新样式
        if self.new_version_info:
            self.statusbar_button2.setStyleSheet(statusbar_button_style_version)
        else:
            self.statusbar_button2.setStyleSheet(statusbar_button_style)
        self.statusbar_label.setStyleSheet(statusbar_label_style)
        self.statusbar_label0.setStyleSheet(statusbar_label_style)
        self.statusbar_label1.setStyleSheet(statusbar_label_style)


        # 返回主窗口样式
        return f""" 
                /* 浅色模式 */
            """

    def dark_style(self):
            """返回暗黑模式的样式表"""

            BACKCOLOR_ = self.background_color_default  # 配置中的背景色
            # 定义通用颜色变量
            BACKCOLOR = "rgb( 15, 17, 30)"   # 浅蓝色背景
            GRAY = "rgb(127, 127, 127)"      # 灰色
            WHITE = "rgb(238,238,238)"       # 白色
            BLACK = "rgb( 34, 40, 49)"       # 黑色

            
            table_style = f"""
                QTableWidget#RB_QTableWidget0 {{
                    /* 表格整体样式 */
                    background-color: {BLACK};
                    color: {WHITE};
                }}
                
                QTableWidget#RB_QTableWidget0::item {{
                    /* 单元格样式 */
                    background-color: {GRAY};
                    color: {BLACK};
                }}
                
                QTableWidget#RB_QTableWidget0::item:selected {{
                    /* 选中单元格样式 */
                    background-color: {BLACK};
                    color: {WHITE};
                }}
                
                /* 添加表头样式 */
                QHeaderView::section {{
                    background-color: {BLACK};
                    color: {WHITE};
                    text-align: center;
                    padding: 3px;
                    margin: 1px;
                    font-family: "{self.custom_font.family()}";
                    font-size: {self.custom_font.pointSize()}pt;
                }}
              
                /* 设置空列头的背景色 */
                QTableWidget::verticalHeader {{
                    background-color: {BACKCOLOR}; /* 空列头背景色 */
                }}                
                
                /* 修改滚动条样式 */
                QScrollBar:vertical {{
                    background: {BLACK}; /* 滚动条背景 */
                    width: 10px; /* 滚动条宽度 */
                    margin: 22px 0 22px 0; /* 上下边距 */
                }}
                QScrollBar::handle:vertical {{
                    background: {GRAY}; /* 滚动条滑块颜色 */
                    min-height: 20px; /* 滚动条滑块最小高度 */
                }}
                QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{
                    background: none; /* 隐藏上下箭头 */
                }}
                QScrollBar::up-arrow:vertical, QScrollBar::down-arrow:vertical {{
                    background: none; /* 隐藏箭头 */
                }}
                QScrollBar:horizontal {{
                    background: {BLACK}; /* 滚动条背景 */
                    height: 10px; /* 滚动条高度 */
                    margin: 0 22px 0 22px; /* 左右边距 */
                }}
                QScrollBar::handle:horizontal {{
                    background: {GRAY}; /* 滚动条滑块颜色 */
                    min-width: 20px; /* 滚动条滑块最小宽度 */
                }}
                QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {{
                    background: none; /* 隐藏左右箭头 */
                }}
                QScrollBar::left-arrow:horizontal, QScrollBar::right-arrow:horizontal {{
                    background: none; /* 隐藏箭头 */
                }}
                
            """
            left_qframe_style = f"""
                QFrame#Left_QFrame {{ 
                    background-color: {BLACK};
                    color: {WHITE};
                    border-radius: 10px;
                    border: 1px solid {GRAY};
                }}
            """

            # 按钮组件和复选框组件样式
            button_style = f"""
                QPushButton {{
                    background-color: rgb( 58, 71, 80);
                    color: {WHITE};
                    text-align: center;
                    font-family: "{self.custom_font.family()}";
                    font-size: {self.custom_font.pointSize()}pt;
                }}
                QPushButton:hover {{
                    border: 1px solid {BACKCOLOR};
                    background-color: {BACKCOLOR};
                }}
            """

            # 设置单选按钮样式
            radio_button_style = f"""   
                QRadioButton {{
                    text-align: left;
                    color: {WHITE};
                    font-family: "{self.custom_font.family()}";
                    font-size: {self.custom_font.pointSize()}pt;
                }}
            """


            # 左侧文件浏览区域样式
            left_area_style = f"""
                QTreeView#Left_QTreeView {{
                    background-color: {BLACK};
                    color: {WHITE};
                    border-radius: 10px;
                }}

                /* 修改滚动条样式 */
                QScrollBar:vertical {{
                    background: {BLACK}; /* 滚动条背景 */
                    width: 10px; /* 滚动条宽度 */
                    margin: 22px 0 22px 0; /* 上下边距 */
                }}
                QScrollBar::handle:vertical {{
                    background: {GRAY}; /* 滚动条滑块颜色 */
                    min-height: 20px; /* 滚动条滑块最小高度 */
                }}
                QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{
                    background: none; /* 隐藏上下箭头 */
                }}
                QScrollBar::up-arrow:vertical, QScrollBar::down-arrow:vertical {{
                    background: none; /* 隐藏箭头 */
                }}
                QScrollBar:horizontal {{
                    background: {BLACK}; /* 滚动条背景 */
                    height: 10px; /* 滚动条高度 */
                    margin: 0 22px 0 22px; /* 左右边距 */
                }}
                QScrollBar::handle:horizontal {{
                    background: {GRAY}; /* 滚动条滑块颜色 */
                    min-width: 20px; /* 滚动条滑块最小宽度 */
                }}
                QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {{
                    background: none; /* 隐藏左右箭头 */
                }}
                QScrollBar::left-arrow:horizontal, QScrollBar::right-arrow:horizontal {{
                    background: none; /* 隐藏箭头 */
                }}

            """
            
            # 下拉框通用样式模板
            combobox_style = f"""
                QComboBox {{
                    /* 下拉框本体样式 */
                    background-color: {BLACK};
                    color: {WHITE};
                    selection-background-color: {BACKCOLOR};
                    selection-color: {WHITE};
                    min-height: 30px;
                    font-family: "{self.custom_font.family()}";
                    font-size: {self.custom_font.pointSize()}pt;
                }}
                
                QComboBox QAbstractItemView {{
                    /* 下拉列表样式 */
                    background-color: {BLACK};
                    color: {WHITE};
                    selection-background-color: {WHITE};
                    selection-color: {BLACK};
                    font-family: "{self.custom_font.family()}";
                    font-size: {self.custom_font.pointSize()}pt;
                }}
                
                QComboBox QAbstractItemView::item {{
                    /* 下拉项样式 */
                    min-height: 25px;
                    padding: 5px;
                    font-family: "{self.custom_font.family()}";
                    font-size: {self.custom_font.pointSize()}pt;
                }}

                QComboBox::hover {{
                    background-color: {BACKCOLOR};
                    color: {WHITE};
                }}  

            """

            # 下拉框通用样式模板2
            combobox_style2 = f"""
                QComboBox {{
                    /* 下拉框本体样式 */
                    background-color: {BLACK};
                    color: {WHITE};
                    selection-background-color: {BACKCOLOR};
                    selection-color: {WHITE};
                    min-height: 30px;
                    font-family: "{self.custom_font.family()}";
                    font-size: {self.custom_font.pointSize()}pt;
                }}
                
                QComboBox QAbstractItemView {{
                    /* 下拉列表样式 */
                    background-color: {WHITE};
                    color: {BLACK};
                    selection-background-color: {BACKCOLOR_};
                    selection-color: {WHITE};
                    font-family: "{self.custom_font.family()}";
                    font-size: {self.custom_font.pointSize()}pt;
                }}
                
                QComboBox QAbstractItemView::item {{
                    /* 下拉项样式 */
                    min-height: 25px;
                    padding: 5px;
                    font-family: "{self.custom_font.family()}";
                    font-size: {self.custom_font.pointSize()}pt;
                }}

            """
    
    
            statusbar_label_style = f"""
                border: none;
                color: {WHITE};
                font-family: {self.custom_font_jetbrains_small.family()};
                font-size: {self.custom_font_jetbrains_small.pointSize()}pt;
                
            """

            statusbar_button_style = f"""
                QPushButton {{
                    background-color: {BLACK};
                    color: {WHITE};
                    text-align: center;
                    font-family: "{self.custom_font_jetbrains_small.family()}";
                    font-size: {self.custom_font_jetbrains_small.pointSize()}pt;
                }}
                QPushButton:hover {{
                    border: 1px solid {BACKCOLOR};
                    background-color: {BACKCOLOR};
                    color: {WHITE};
                }}
            """

            statusbar_button_style_version = f"""
                QPushButton {{
                    background-color: {"rgb(245,108,108)"};
                    color: {WHITE};
                    text-align: center;
                    font-family: "{self.custom_font_jetbrains_small.family()}";
                    font-size: {self.custom_font_jetbrains_small.pointSize()}pt;
                }}
                QPushButton:hover {{
                    border: 1px solid {BACKCOLOR};
                    background-color: {"rgb(245,108,108)"};
                    color: {WHITE};
                }}
            """  

            statusbar_style = f"""
                border: none;
                background-color: {BLACK};
                color: {WHITE};
            """


            # 设置左上侧文件浏览区域样式
            self.Left_QTreeView.setStyleSheet(left_area_style)

            # 设置左下角侧框架样式
            self.Left_QFrame.setStyleSheet(left_qframe_style)
            self.L_radioButton1.setStyleSheet(radio_button_style)  # 设置左对齐
            self.L_radioButton2.setStyleSheet(radio_button_style)  # 设置左对齐


            # 设置右侧顶部按钮下拉框样式
            self.RT_QPushButton3.setStyleSheet(button_style)
            self.RT_QPushButton5.setStyleSheet(button_style)

            self.RT_QComboBox.setStyleSheet(combobox_style2)
            self.RT_QComboBox1.setStyleSheet(combobox_style2)

            self.RT_QComboBox0.setStyleSheet(combobox_style)
            self.RT_QComboBox2.setStyleSheet(combobox_style)
            self.RT_QComboBox3.setStyleSheet(combobox_style)

            # 设置右侧中间表格区域样式
            self.RB_QTableWidget0.setStyleSheet(table_style)

            # 设置底部状态栏区域样式 self.statusbar --> self.statusbar_widget --> self.statusbar_QHBoxLayout --> self.statusbar_button1 self.statusbar_button2
            self.statusbar.setStyleSheet(statusbar_style)
            self.statusbar_button1.setStyleSheet(statusbar_button_style)
            # self.statusbar_button2.setStyleSheet(statusbar_button_style)
            # 设置版本按钮更新样式
            if self.new_version_info:
                self.statusbar_button2.setStyleSheet(statusbar_button_style_version)
            else:
                self.statusbar_button2.setStyleSheet(statusbar_button_style)
            self.statusbar_label.setStyleSheet(statusbar_label_style)
            self.statusbar_label0.setStyleSheet(statusbar_label_style)
            self.statusbar_label1.setStyleSheet(statusbar_label_style)

            # 返回主窗口样式
            return f"""
                QWidget#main_body {{ /* 主窗口背景色 */
                    background-color: black;
                    color: white;
                }}

                QSplitter {{ /* 分割器背景色 */
                    background-color: black;
                    color: white;
                }}
                QSplitter::handle {{ /* 分割器手柄背景色 */
                    background-color: black;
                    color: white;
                }}
                QSplitter::handle:hover {{ /* 分割器手柄悬停背景色 */
                    background-color: black;
                    color: white;
                }}

                QGroupBox#Left_QGroupBox {{ /* 左侧组框1_背景色 */
                    background-color: black;
                    color: white;
                }}
                QGroupBox#Left_QGroupBox::title {{ /* 左侧组框1_标题背景色 */
                    background-color: black;
                    color: white;
                }}
                QGroupBox#Left_QGroupBox::title:hover {{ /* 左侧组框1_标题悬停背景色 */
                    background-color: black;
                    color: white;
                }}

                QGroupBox#Right_Top_QGroupBox {{ /* 右侧组框2_背景色 */
                    background-color: black;
                    color: white;
                }}   
                QGroupBox#Right_Top_QGroupBox::title {{ /* 右侧组框2_标题背景色 */
                    background-color: black;
                    color: white;
                }}
                QGroupBox#Right_Top_QGroupBox::title:hover {{ /* 右侧组框2_标题悬停背景色 */
                    background-color: black;
                    color: white;
                }}

                QGroupBox#Right_Bottom_QGroupBox {{ /* 右侧组框3_背景色 */
                    background-color: black;
                    color: white;
                }}   
                QGroupBox#Right_Bottom_QGroupBox::title {{ /* 右侧组框3_标题背景色 */
                    background-color: black;
                    color: white;
                }}
                QGroupBox#Right_Bottom_QGroupBox::title:hover {{ /* 右侧组框3_标题悬停背景色 */
                    background-color: black;
                    color: white;
                }}
                
                /* 表格样式 */
                
            """

    def cleanup(self):
        """清理资源"""
        self.cancel_preloading()
        if self.compare_window:
            self.compare_window.deleteLater()
            self.compare_window = None
            
        self.threadpool.clear()
        self.threadpool.waitForDone()
        
        gc.collect()

    """缓存文件路径列表，避免重复加载"""
    def load_settings(self):
        """从JSON文件加载设置"""
        print("load_settings()--从JSON文件加载之前的设置")
        try:
            settings_path = os.path.join(os.path.dirname(__file__), "cache", "basic_settings.json")
            if os.path.exists(settings_path):
                with open(settings_path, "r", encoding='utf-8', errors='ignore') as f:
                    settings = json.load(f)

                    # 恢复地址栏历史记录和当前目录
                    combobox_history = settings.get("combobox_history", [])
                    self.RT_QComboBox.clear()
                    self.RT_QComboBox.addItems(combobox_history)
                    current_directory = settings.get("current_directory", "")
                    if current_directory and os.path.exists(current_directory):
                        self.RT_QComboBox.setCurrentText(current_directory)

                    # 恢复文件类型选择
                    selected_option = settings.get("file_type_option", "显示图片文件")
                    index = self.RT_QComboBox0.findText(selected_option)
                    if index >= 0:
                        self.RT_QComboBox0.setCurrentIndex(index)

                    # 恢复排序方式
                    sort_option = settings.get("sort_option", "按创建时间排序")
                    index = self.RT_QComboBox2.findText(sort_option)
                    if index >= 0:
                        self.RT_QComboBox2.setCurrentIndex(index)

                    # 恢复主题设置
                    theme_option = settings.get("theme_option", "默认主题")
                    index = self.RT_QComboBox3.findText(theme_option)
                    if index >= 0:
                        self.RT_QComboBox3.setCurrentIndex(index)
                        self.current_theme = settings.get("current_theme", "默认主题")
                        self.apply_theme()

                    # 恢复文件夹选择状态
                    all_items = settings.get("combobox1_all_items", [])
                    checked_items = settings.get("combobox1_checked_items", [])
                    
                    if all_items:
                        self.model = CheckBoxListModel(all_items)
                        self.RT_QComboBox1.setModel(self.model)
                        self.RT_QComboBox1.setItemDelegate(CheckBoxDelegate())
                        self.RT_QComboBox1.setContextMenuPolicy(Qt.NoContextMenu)

                        # 恢复选中状态
                        for i, item in enumerate(self.model.items):
                            if item in checked_items:
                                self.model.setChecked(self.model.index(i))
                        # 更新同级文件夹下拉框选项
                        self.updateComboBox1Text()
                    else:
                        # 初始化同级文件夹下拉框选项
                        self.RT_QComboBox1_init()

                    # 恢复极简模式状态,默认开启
                    self.simple_mode = settings.get("simple_mode", True)

                    # 恢复拖拽模式状态,默认开启
                    self.drag_flag = settings.get("drag_flag", True)
            else:
                # 若没有cache/设置，则在此初始化主题设置--默认主题
                self.apply_theme()

        except Exception as e:
            print(f"加载设置时出错: {e}")
            return

    def save_settings(self):
        """保存当前设置到JSON文件"""
        try:
            settings_path = os.path.join(os.path.dirname(__file__), "cache", "basic_settings.json")
            
            # 确保cache目录存在
            cache_dir = os.path.dirname(settings_path)
            if not os.path.exists(cache_dir):
                os.makedirs(cache_dir)

            # 收集所有需要保存的设置
            settings = {
                # 地址栏历史记录和当前目录
                "combobox_history": [self.RT_QComboBox.itemText(i) for i in range(self.RT_QComboBox.count())],
                "current_directory": self.RT_QComboBox.currentText(),
                
                # 文件类型选择
                "file_type_option": self.RT_QComboBox0.currentText(),
                
                # 文件夹选择状态
                "combobox1_checked_items": self.model.getCheckedItems() if hasattr(self, 'model') and self.model else [],
                "combobox1_all_items": self.model.items[1:] if hasattr(self, 'model') and self.model else [],
                
                # 排序方式
                "sort_option": self.RT_QComboBox2.currentText(),
                
                # 主题设置
                "theme_option": self.RT_QComboBox3.currentText(),
                "current_theme": self.current_theme,
                
                # 极简模式状态
                "simple_mode": self.simple_mode,

                # 拖拽模式状态
                "drag_flag": self.drag_flag

            }

            # 保存设置到JSON文件
            with open(settings_path, "w", encoding='utf-8', errors='ignore') as f:
                json.dump(settings, f, ensure_ascii=False, indent=4)

        except Exception as e:
            print(f"保存设置时出错: {e}")


    def press_space_and_b_get_selected_file_paths(self, key_type):
        """返回右侧表格选中的文件的路径列表"""
        try:
            selected_items = self.RB_QTableWidget0.selectedItems()  # 获取选中的项
            if not selected_items:
                print("没有选中的项！")
                # 弹出提示框
                show_message_box("没有选中的项！", "提示", 500)
                return [], []
            
            # 清除所有选中的项
            self.RB_QTableWidget0.clearSelection() 
            # 获取最大最小的行索引
            row_max = self.RB_QTableWidget0.rowCount() - 1 
            row_min = 0
            # 用于存储文件路径的列表
            file_paths = []  
            # 用于存储当前选中图片张数
            current_image_index = []    
            
            # 判断是否是首次按键
            if not self.last_key_press:
                step_row = 0  # 首次按键不移动
                # 第二次进入设置为True
                self.last_key_press = True
            else:
                # 统计行索引需要移动step
                if len(set([item.column() for item in selected_items])) == len(selected_items):
                    # 如果选中项的个数和图片列数相等，则表示是单选，行索引移动step_row = 1
                    step_row = 1
                else:# 如果选中项的个数和图片列数不相等，则表示是多选，行索引移动step_row = 选中项的行索引去重后长度
                     step_row = len(set([item.row() for item in selected_items]))   
            
            # 遍历选中的项
            for item in selected_items:
                # 获取当前项的列索引行索引
                col_index = item.column()
                row_index = item.row()
                # 判断按下space和b来控制选中的单元格上移和下移
                if key_type == 'space':    # 空格键获取下一组图片
                    row_index += step_row
                elif key_type == 'b':      # B键获取上一组图片
                    row_index -= step_row
                else:
                    print("没有按下space和b键")

                if row_index > row_max or row_index < row_min:  # 修正边界检查
                    self.RB_QTableWidget0.clearSelection()      # 清除所有选中的项
                    print(f"已超出表格范围: {row_index}")
                    return [], []
                else:
                    item = self.RB_QTableWidget0.item(row_index, col_index)
                    if item and item.text():
                        item.setSelected(True)  # 选中当前单元格
                        # 构建图片完整路径
                        file_name = item.text().split('\n')[0]  # 获取文件名，修改获取方式(第一行为需要的文件名)
                        column_name = self.RB_QTableWidget0.horizontalHeaderItem(col_index).text()
                        current_directory = self.RT_QComboBox.currentText()  # 获取当前选中的目录
                        full_path = str(Path(current_directory).parent / column_name / file_name)
                        
                        if os.path.isfile(full_path):
                            file_paths.append(full_path)  # 只有在是有效文件时才添加到列表中
                        else:
                            print(f"无效文件路径: {full_path}")  # 输出无效文件路径的提示   
                    else:
                        print(f"item is None or item.text() is None")

                # 如果选中项的列数和图片列数相等，则打印当前处理图片张数
                if not self.image_index_max: # 如果image_index_max为空，则初始化为当前表格的最大行数
                    print("image_index_max is None")
                    self.image_index_max = [self.RB_QTableWidget0.rowCount()] * self.RB_QTableWidget0.columnCount()
                if row_index+1 > self.image_index_max[col_index]:
                    pass
                else:
                    current_image_index.append(f"{row_index+1}/{self.image_index_max[col_index]}")

            # 将选中的单元格滚动到视图中间位置
            self.RB_QTableWidget0.scrollToItem(item, QtWidgets.QAbstractItemView.PositionAtCenter)
                                  
            print(f"当前选中图片张数：{current_image_index}")
            
            return file_paths, current_image_index  # 返回文件路径列表
        except Exception as e:
            print(f"处理键盘按下事件时发生错误: {e}")
            return [], []
    
    def on_f1_pressed(self):
        """处理F1键按下事件"""
        try:
            self.open_mipi2raw_tool()
        except Exception as e:
            print(f"on_f1_pressed()-error--处理F1键按下事件失败: {e}")
            return


    """键盘按下事件处理""" 
    def on_f2_pressed(self):
        """处理F2键按下事件"""
        selected_items = self.RB_QTableWidget0.selectedItems()  # 获取选中的项
        if not selected_items:
            show_message_box("没有选中的项！", "提示", 500)
            return
            
        current_folder, _ = self.press_space_and_b_get_selected_file_paths('test')
        if not current_folder:
            show_message_box("没有选中的项！", "提示", 500)
            return

        try:    
            if len(selected_items) == 1:
                # 单文件重命名
                dialog = SingleFileRenameDialog(current_folder[0], self)
                if dialog.exec_() == QDialog.Accepted:
                    
                    # 获取新的文件路径
                    new_file_path = dialog.get_new_file_path()
                    
                    if new_file_path:
                        # 获取新的文件名
                        new_file_name = os.path.basename(new_file_path)
                        # 获取选中的单元格
                        item = selected_items[0]
                        row = item.row()
                        col = item.column() 

                        # 更新单元格内容
                        current_text = item.text()
                        if '\n' in current_text:  # 如果有多行文本
                            # 保持原有的其他信息，只更新文件名
                            lines = current_text.split('\n')
                            lines[0] = new_file_name  # 更新第一行的文件名
                            new_text = '\n'.join(lines)
                        else:
                            new_text = new_file_name
                            
                        # 设置新的单元格文本
                        self.RB_QTableWidget0.item(row, col).setText(new_text)
            else:
                # 多文件重命名
                self.open_rename_tool(current_folder)

        except Exception as e:
            print(f"on_f2_pressed()-error--处理F2键按下事件失败: {e}")
            return



    def on_f4_pressed(self):
        """处理F4键按下事件"""
        current_folder = self.RT_QComboBox.currentText()
        current_folder = os.path.dirname(current_folder) # 获取当前选中的文件夹上一级文件夹路径
        if current_folder:
            try:
                self.open_rename_tool(current_folder)
            except Exception as e:
                print(f"on_f4_pressed()-error--处理F4键按下事件失败: {e}")
                return
        else:
            # 弹出提示框    
            show_message_box("当前没有选中的文件夹", "提示", 500)

    def on_f5_pressed(self):
        """处理F5键按下事件"""

        try:    
            # 刷新表格
            show_message_box("刷新表格&清除缓存-", "提示", 500)

            # 删除缓存文件中的zip文件
            cache_dir = os.path.join(os.path.dirname(__file__), "cache")
            if os.path.exists(cache_dir):
                force_delete_folder(cache_dir, '.zip')

            # 清除图标缓存
            IconCache.clear_cache()
            
            # 更新表格
            self.update_RB_QTableWidget0()

        except Exception as e:
            print(f"on_f5_pressed()-error--刷新表格&清除缓存失败: {e}")
            return

    def on_f12_pressed(self):
        """处理F12键按下事件,重启程序"""
        self.close()
        try:
            program_path = os.path.join(os.path.dirname(__file__), "hiviewer.exe")
            if os.path.exists(program_path):
                
                # 使用os.startfile启动程序
                os.startfile(program_path)
                
                # 等待5秒确保程序启动
                time.sleep(3)  
                print(f"已启动程序: hiviewer.exe")
                
                return True
            else:
                print(f"程序文件不存在: {program_path}")
                return False
        except Exception as e:
            print(f"启动程序失败: {e}")
            return False

    def on_escape_pressed(self):
        print("escape被按下了")
        self.close()  # 关闭主界面
        self.save_settings()

    def on_alt_pressed(self):
        self.drag_flag = not self.drag_flag
        if self.drag_flag:
            show_message_box("切换到拖拽模式", "提示", 500)
        else:
            show_message_box("关闭拖拽模式", "提示", 500)
        

    def on_p_pressed(self):
        """处理P键按下事件"""
        print("on_p_pressed()-切换主题--P键已按下, 更新下拉框选项")
        try:
            if self.current_theme == "默认主题":
                self.RT_QComboBox3.setCurrentIndex(self.RT_QComboBox3.findText("暗黑主题"))
            else:
                self.RT_QComboBox3.setCurrentIndex(self.RT_QComboBox3.findText("默认主题"))

            # 更新主题
            self.toggle_theme()
        except Exception as e:
            print(f"on_p_pressed()--切换主题失败: {e}")
                

    def on_i_pressed(self):
        """处理i键按下事件,调用高通工具后台解析图片的exif信息"""
        # 获取当前选中的文件类型
        selected_option = self.RT_QComboBox.currentText()
        try:

            # 创建并显示自定义对话框,传入图片列表
            dialog = Qualcom_Dialog(selected_option)

            # 显示对话框
            if dialog.exec_() == QDialog.Accepted:

                # 执行命名
                dict_info = dialog.get_data()
                # print(f"用户加载的路径信息: {dict_info}")

                qualcom_path = dict_info.get("Qualcom工具路径","")
                images_path = dict_info.get("Image文件夹路径","")
                metadata_path = os.path.join(os.path.dirname(__file__), "tools", "metadata.exe")

                # 拼接参数命令字符串
                if qualcom_path and images_path and os.path.exists(metadata_path) and os.path.exists(images_path) and os.path.exists(qualcom_path):
                    command = f"{metadata_path} --chromatix \"{qualcom_path}\" --folder \"{images_path}\""

                    """
                    # 添加检查 图片文件夹目录下是否已存在xml文件，不存在则启动线程解析图片
                    # xml_exists = [f for f in os.listdir(images_path) if f.endswith('_new.xml')]

                    针对上面的代码，优化了检查'_new.xml'文件的逻辑:
                    1. os.listdir(images_path) 列出文件夹中的所有文件
                    2. os.path.exists(os.path.join(images_path, f)) 检查文件是否存在
                    3. any() 函数会在找到第一个符合条件的文件时立即返回 True, 避免不必要的遍历
                    """
                    # 检查图片文件夹目录下是否存在xml文件，不存在则启动线程解析图片
                    xml_exists = any(f for f in os.listdir(images_path) if f.endswith('_new.xml'))

                    # 创建线程，必须在主线程中连接信号
                    self.command_thread = CommandThread(command, images_path)
                    self.command_thread.finished.connect(self.on_command_finished)  # 连接信号
                    # self.command_thread.finished.connect(self.cleanup_thread)  # 连接清理槽

                    if not xml_exists:
                        self.command_thread.start()  # 启动线程
                        show_message_box("正在使用高通工具后台解析图片Exif信息...", "提示", 1000)
                    else:
                        show_message_box("已有xml文件, 无须解析图片", "提示", 1000)

                        # 解析xml文件将其保存到excel中去
                        save_excel_data(images_path)

            # 无论对话框是接受还是取消，都手动销毁对话框
            dialog.deleteLater()
            dialog = None

        except Exception as e:
            print(f"on_i_pressed()-error--处理i键按下事件失败: {e}")
            return


    def on_command_finished(self, success, error_message, images_path=None):
        """处理命令执行完成的信号"""
        try:
            if success and images_path:
                # 解析xml文件将其保存到excel中去
                save_excel_data(images_path)
                # 提示
                show_message_box("后台解析图片成功！", "提示", 1000)
                print(f"高通工具后台解析图片成功！")
            else:
                show_message_box(f"高通工具后台解析图片失败: {error_message}", "提示", 2000)
                print(f"高通工具后台解析图片失败: {error_message}")

        except Exception as e:
            show_message_box(f"高通工具后台解析图片失败: {error_message}", "提示", 2000)
            print(f"on_command_finished()-error--高通工具后台解析图片成功失败: {e}")
            return


    def on_l_pressed(self):
        """处理L键打开图片处理工具"""
        try:
            # 获取选中项并验证
            selected_items = self.RB_QTableWidget0.selectedItems()
            if not selected_items or len(selected_items) != 1:
                show_message_box("请选择单个图片文件", "提示", 500)
                return

            # 构建完整文件路径
            current_dir = self.RT_QComboBox.currentText()
            if not current_dir:
                show_message_box("当前目录无效", "提示", 500)
                return

            # 获取文件信息
            item = selected_items[0]
            column_name = self.RB_QTableWidget0.horizontalHeaderItem(item.column()).text()
            file_name = item.text().split('\n')[0]
            full_path = str(Path(current_dir).parent / column_name / file_name)

            # 验证文件有效性
            if not full_path.lower().endswith(self.IMAGE_FORMATS):
                show_message_box(f"不支持的文件格式: {os.path.splitext(full_path)[1]}", "提示", 500)
                return
                
            if not os.path.isfile(full_path):
                show_message_box(f"文件不存在: {os.path.basename(full_path)}", "提示", 500)
                return

            # 打开处理窗口
            self.open_image_process_window(full_path)

        except Exception as e:
            error_msg = f"打开图片失败: {str(e)}"
            show_message_box(error_msg, "错误", 1000)

    def on_ctrl_h_pressed(self):
        """处理Ctrl+h键按下事件, 打开帮助信息"""
        try:
            # 单例模式管理帮助窗口
            if not hasattr(self, 'help_dialog'):
                # 构建文档路径,使用说明文档+版本更新文档
                doc_dir = os.path.join(os.path.dirname(__file__), "docs")
                User_path = os.path.join(doc_dir, "User_Manual.md")
                Version_path = os.path.join(doc_dir, "Version_Updates.md")
                
                # 验证文档文件存在性
                if not os.path.isfile(User_path) or not os.path.isfile(Version_path):
                    show_message_box(f"帮助文档未找到:\n{User_path}or{Version_path}", "配置错误", 2000)
                    return
                
                # 初始化对话框
                self.help_dialog = AboutDialog(User_path,Version_path)

            # 激活现有窗口
            self.help_dialog.show()
            self.help_dialog.raise_()
            self.help_dialog.activateWindow()
            # 链接关闭事件
            self.help_dialog.finished.connect(self.close_helpinfo)
            
        except Exception as e:
            error_msg = f"无法打开帮助文档:\n{str(e)}\n请检查程序是否包含文件: ./docs/update_main_logs.md"
            show_message_box(error_msg, "严重错误", 3000)

    def close_helpinfo(self):
        """关闭对话框事件"""
        # 手动销毁对话框
        if hasattr(self, 'help_dialog'):
            # 强制删除
            del self.help_dialog
            print("成功销毁对话框")


    def on_b_pressed(self):
        """处理B键按下事件，用于查看上一组图片/视频"""
        try:
            # 按键防抖机制，防止快速多次按下导致错误，设置0.5秒内不重复触发
            current_time = time.time()
            if hasattr(self, 'last_space_press_time') and current_time - self.last_space_press_time < 0.5:  
                return
            self.last_space_press_time = current_time

            # 获取选中单元格的文件路径和索引
            selected_file_paths, image_indexs = self.press_space_and_b_get_selected_file_paths('b')
            if not selected_file_paths:
                return
            
            # 限制最多选中8个文件
            if len(selected_file_paths) > 8:
                show_message_box("最多只能同时选中8个文件", "提示", 1000)
                # 恢复第一次按下键盘空格键或B键
                self.last_key_press = False 
                return

            # 获取所有文件的扩展名并去重
            file_extensions = {os.path.splitext(path)[1].lower() for path in selected_file_paths}
            
            # 检查是否存在多种文件类型
            if len(file_extensions) > 1:
                flag_video = 0
                flag_image = 0
                flag_other = 0
                # 检查文件类型的合法性
                for ext in file_extensions:
                    if ext in self.VIDEO_FORMATS:
                        flag_video = 1
                        # show_message_box("视频播放功能暂不支持", "提示", 500)
                    elif ext in self.IMAGE_FORMATS:
                        flag_image = 1
                        # show_message_box(f"不支持多选{ext}格式文件", "提示", 500)
                    else:
                        flag_other = 1
                        # show_message_box("不支持的文件格式", "提示", 500)
                
                if flag_video and flag_image and flag_other:
                    show_message_box("不支持同时选中多种文件格式", "提示", 500)
                    return
                
                if flag_video and flag_image and not flag_other:
                    show_message_box("不支持同时选中视频和图片文件", "提示", 500)
                    return

                if flag_video and not flag_image and flag_other:
                    show_message_box("不支持同时选中视频和其它文件", "提示", 500)
                    return
                
                if not flag_image and flag_video and flag_other:
                    show_message_box("不支持同时选中图片和其它文件", "提示", 500)
                    return

            # 获取统一的文件类型
            file_ext = file_extensions.pop()  # 只有一个元素，直接获取

            # 根据文件类型处理
            if file_ext in self.VIDEO_FORMATS:
                
                # 限制视频文件的数量
                if len(selected_file_paths) > 5:
                    show_message_box("最多支持同时比较5个视频文件", "提示", 1000)
                    # 恢复第一次按下键盘空格键或B键
                    self.last_key_press = False 
                    return
                
                self.create_video_player(selected_file_paths, image_indexs)
            elif file_ext in self.IMAGE_FORMATS:
                self.create_compare_window(selected_file_paths, image_indexs)
            else:
                show_message_box("不支持的文件格式", "提示", 1000)

        except Exception as e:
            print(f"处理B键时发生错误: {e}")

    def on_space_pressed(self):
        """处理空格键按下事件"""
        try:
            # 按键防抖机制，防止快速多次按下导致错误，设置0.5秒内不重复触发
            current_time = time.time()
            if hasattr(self, 'last_space_press_time') and current_time - self.last_space_press_time < 0.5:  
                return
            self.last_space_press_time = current_time

            # 获取选中单元格的文件路径和索引
            selected_file_paths, image_indexs = self.press_space_and_b_get_selected_file_paths('space')
            if not selected_file_paths:
                return
            
            # 限制最多选中8个文件
            if len(selected_file_paths) > 8:
                show_message_box("最多只能同时选中8个文件", "提示", 1000)
                # 恢复第一次按下键盘空格键或B键
                self.last_key_press = False 
                return

            # 获取所有文件的扩展名并去重
            file_extensions = {os.path.splitext(path)[1].lower() for path in selected_file_paths}
            
            # 检查是否存在多种文件类型
            if len(file_extensions) > 1:
                flag_video = 0
                flag_image = 0
                flag_other = 0
                # 检查文件类型的合法性
                for ext in file_extensions:
                    if ext in self.VIDEO_FORMATS:
                        flag_video = 1
                        # show_message_box("视频播放功能暂不支持", "提示", 500)
                    elif ext in self.IMAGE_FORMATS:
                        flag_image = 1
                        # show_message_box(f"不支持多选{ext}格式文件", "提示", 500)
                    else:
                        flag_other = 1
                        # show_message_box("不支持的文件格式", "提示", 500)
                
                if flag_video and flag_image and flag_other:
                    show_message_box("不支持同时选中多种文件格式", "提示", 500)
                    return
                
                if flag_video and flag_image and not flag_other:
                    show_message_box("不支持同时选中视频和图片文件", "提示", 500)
                    return

                if flag_video and not flag_image and flag_other:
                    show_message_box("不支持同时选中视频和其它文件", "提示", 500)
                    return
                
                if not flag_image and flag_video and flag_other:
                    show_message_box("不支持同时选中图片和其它文件", "提示", 500)
                    return

            # 获取统一的文件类型
            file_ext = file_extensions.pop()  # 只有一个元素，直接获取

            # 根据文件类型处理
            if file_ext in self.VIDEO_FORMATS:
                
                # 限制视频文件的数量
                if len(selected_file_paths) > 5:
                    show_message_box("最多支持同时比较5个视频文件", "提示", 1000)
                    return
                
                # 打开视频对比界面
                self.create_video_player(selected_file_paths, image_indexs)

            elif file_ext in self.IMAGE_FORMATS:
                
                # 打开看图对比界面
                self.create_compare_window(selected_file_paths, image_indexs)
            
            else:
                show_message_box("不支持的文件格式", "提示", 1000)

        except Exception as e:
            print(f"on_space_pressed()-error--处理空格键时发生错误: {e}")
            return

    def create_compare_window(self, selected_file_paths, image_indexs):
        """创建看图子窗口的统一方法"""
        
        # 暂停预加载
        # self.pause_preloading() # modify by diamond_cz 20250217 不暂停预加载，看图时默认后台加载图标
        
        # 在主界面加载并显示进度条
        self.set_progress_bar(len(selected_file_paths))

        if not self.compare_window:
            self.compare_window = SubMainWindow(selected_file_paths, image_indexs, self)
        else:  
            self.compare_window.set_images(selected_file_paths, image_indexs)
            print("看图子界面已存在进入窗口！")

        # 延时100ms后关闭进度条显示
        QTimer.singleShot(100, self.on_progress_complete)

        # 设置看图界面标题
        self.compare_window.setWindowTitle("图片对比界面")
        # self.compare_window.setWindowFlags(Qt.Window)
        # 设置窗口图标
        icon_path = os.path.join(os.path.dirname(__file__), "icons", "viewer.ico")
        self.compare_window.setWindowIcon(QIcon(icon_path))
        self.compare_window.closed.connect(self.on_compare_window_closed)
        self.compare_window.show()

        # self.hide()  # modify by diamond_cz 20250217 不隐藏主界面
        

    def on_compare_window_closed(self):
        """处理子窗口关闭事件"""

        # self.show() # self.hide()  # modify by diamond_cz 20250217 不隐藏主界面
        if self.compare_window:
            print("主界面触发子窗口关闭事件,设置隐藏")
            # 删除引用以释放资源
            # self.compare_window.should_close = True
            self.compare_window.deleteLater()
            self.compare_window = None
            # self.compare_window.close()
            
        # else:
        #     print("子窗口不存在")

        # self.show() # 显示主窗口
        # self.resume_preloading() # modify by diamond_cz 20250217 已取消暂定预加载逻辑，不用恢复预加载
        # 恢复第一次按下键盘空格键或B键
        self.last_key_press = False  

    def pause_preloading(self):
        """暂停预加载"""
        if self.current_preloader and self.preloading:
            self.current_preloader.pause()
            print("预加载已暂停")

    def resume_preloading(self):
        """恢复预加载"""
        if self.current_preloader and self.preloading:
            self.current_preloader.resume()
            print("预加载已恢复")

    def create_video_player(self, selected_file_paths, image_indexs):
        """创建视频播放器的统一方法"""
        self.video_player = VideoWall(selected_file_paths) #, image_indexs
        self.video_player.setWindowTitle("多视频播放程序")
        self.video_player.setWindowFlags(Qt.Window) 
        # 设置窗口图标
        icon_path = os.path.join(os.path.dirname(__file__), "icons", "video_icon.ico")
        self.video_player.setWindowIcon(QIcon(icon_path))
        self.video_player.closed.connect(self.on_video_player_closed)
        self.video_player.show()
        self.hide()  # 隐藏主窗口

    def open_rename_tool(self, current_folder):
        """创建批量重命名的统一方法"""
        self.rename_tool = FileOrganizer()
        self.rename_tool.select_folder(current_folder)  # 传递当前文件夹路径
        self.rename_tool.setWindowTitle("批量重命名")
        # 设置窗口最大化
        # self.rename_tool.showMaximized()
        # 设置窗口图标
        icon_path = os.path.join(os.path.dirname(__file__), "icons", "rename_ico_96x96.ico")
        self.rename_tool.setWindowIcon(QIcon(icon_path))
        # 链接关闭事件
        self.rename_tool.closed.connect(self.on_rename_tool_closed) 
        self.rename_tool.show()
        self.hide()

    def open_image_process_window(self, image_path):
        """创建图片处理子窗口的统一方法"""
        self.image_process_window = SubCompare(image_path)
        self.image_process_window.setWindowTitle("图片调整")
        self.image_process_window.setWindowFlags(Qt.Window)
        # 设置窗口最大化
        # self.image_process_window.showMaximized()
        # 设置窗口图标
        icon_path = os.path.join(os.path.dirname(__file__), "icons", "ps_ico_96x96.ico")
        self.image_process_window.setWindowIcon(QIcon(icon_path))
        # 链接关闭事件
        self.image_process_window.closed.connect(self.on_image_process_window_closed) 
        self.image_process_window.show()
        self.hide()

    def open_bat_tool(self):
        """创建批量执行命令的统一方法"""
        self.bat_tool = LogVerboseMaskApp()
        self.bat_tool.setWindowTitle("批量执行命令")
        # 设置窗口图标
        icon_path = os.path.join(os.path.dirname(__file__), "icons", "cmd_ico_96x96.ico")
        self.bat_tool.setWindowIcon(QIcon(icon_path))
        # 设置窗口最大化
        # self.bat_tool.showMaximized()
        # 链接关闭事件 未添加
        self.bat_tool.closed.connect(self.on_bat_tool_closed)
        self.bat_tool.show()
        self.hide()

    def open_mipi2raw_tool(self):
        """打开MIPI RAW文件转换为JPG文件工具"""
        self.mipi2raw_tool = Mipi2RawConverterApp()
        self.mipi2raw_tool.setWindowTitle("MIPI RAW文件转换为JPG文件")
        
        # 设置窗口图标
        icon_path = os.path.join(os.path.dirname(__file__), "icons", "raw_ico_96x96.ico")
        self.mipi2raw_tool.setWindowIcon(QIcon(icon_path))

        # 添加链接关闭事件
        self.mipi2raw_tool.closed.connect(self.on_mipi2ram_tool_closed)
        self.mipi2raw_tool.show()
        

    def on_video_player_closed(self):
        """处理视频播放器关闭事件"""
        if self.video_player: # 删除引用以释放资源
            self.video_player.deleteLater()
            self.video_player = None
        self.show() # 显示主窗口

        # 恢复第一次按下键盘空格键或B键
        self.last_key_press = False 

    def on_rename_tool_closed(self):
        """处理重命名工具关闭事件"""
        if self.rename_tool:
            self.rename_tool.deleteLater()
            self.rename_tool = None
        self.show()
        self.update_RB_QTableWidget0() # 更新右侧RB_QTableWidget0表格 

    def on_image_process_window_closed(self):
        """处理图片处理子窗口关闭事件"""
        if self.image_process_window:
            self.image_process_window.deleteLater()
            self.image_process_window = None
        self.show() 

    def on_bat_tool_closed(self):
        """处理批量执行命令工具关闭事件"""
        if self.bat_tool:
            self.bat_tool.deleteLater()
            self.bat_tool = None
        self.show()

    def on_mipi2ram_tool_closed(self):
        """处理MIPI RAW文件转换为JPG文件工具关闭事件"""
        if self.mipi2raw_tool:
            self.mipi2raw_tool.deleteLater()
            self.mipi2raw_tool = None
        self.show()


    """
    设置定时器进度条区域结束线
    ------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
    """
    def set_progress_bar(self, num_all):
        """设置进度条"""
        # 检查是否已经存在进度条
        if hasattr(self, 'progress_bar') and self.progress_bar is not None:
            return  # 如果已经有进度条在显示，则直接返回
        
        # 添加进度条并设置样式
        self.progress_bar = QProgressBar(self)
        # 初始化时设置进度条位置
        self.update_progress_bar_position()

        # 设置进度条样式
        self.progress_bar.setStyleSheet("""
            QProgressBar {
                text-align: center;
            }
            QProgressBar::chunk {
                background-color: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                stop:0 #36D1DC, stop:1 #5B86E5);
            }
        """)

        self.progress_bar.setAlignment(Qt.AlignCenter)  # 设置文字居中

        # 启动进度条显示、
        self.progress_bar.setMaximum(num_all)
        self.progress_bar.setValue(0)
        self.progress_bar.setVisible(True)


        # 添加定时器以实现滚动效果
        self.progress_timer = QTimer(self)
        self.progress_timer.timeout.connect(self.update_progress)
        self.progress_timer.start(50)  # 每50ms更新一次进度

        # 设置默认不显示
        # self.progress_bar.setVisible(False) 

    def update_progress_bar_position(self):
        """更新进度条位置，确保其始终在窗口中心"""
        self.progress_bar.setGeometry(
            (self.width() - self.progress_bar.width()) // 2,
            (self.height() - self.progress_bar.height()) // 2,
            400, 40
        )

    def update_progress(self):
        """更新进度条值"""
        if self.progress_bar is None:  # 检查 progress_bar 是否存在
            self.progress_timer.stop()  # 如果不存在，停止定时器
            return
        
        current_value = self.progress_bar.value()
        if current_value < self.progress_bar.maximum():
            self.progress_bar.setValue(current_value + 1)
        else:
            self.progress_timer.stop()  # 达到最大值后停止定时器


    def on_progress_complete(self):
        """进度条完成后的回调函数"""
        if self.progress_bar is not None:  # 检查 progress_bar 是否存在
            self.progress_bar.setValue(self.progress_bar.maximum())
            QApplication.processEvents()
            self.progress_bar.setVisible(False)  # 隐藏进度条
            self.progress_timer.stop()  # 达到最大值后停止定时器
            self.progress_bar.deleteLater()  # 销毁进度条
            self.progress_bar = None  # 将 progress_bar 设置为 None

    """
    设置定时器进度条区域结束线
    ------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
    """

    def closeEvent(self, event):
        """重写关闭事件以保存设置和清理资源"""
        print("主界面关闭事件")
        self.cleanup()
        self.save_settings()
        event.accept()

"""
设置主界面类区域结束线
------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
"""

"""
设置日志区域开始线
------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
需要导入下面两个python内置库:
import logging
from logging.handlers import RotatingFileHandler

相关使用方法：

1. **DEBUG**（调试信息）：
    logging.debug("正在尝试连接数据库...")
    # 适用场景：
    # - 记录程序执行流程
    # - 关键变量值跟踪
    # - 方法进入/退出日志
    # 生产环境应关闭DEBUG级别


2. **INFO**（运行状态信息）：
    logging.info(f"成功加载用户配置：{user_id}")
    # 适用场景：
    # - 重要业务操作记录
    # - 系统状态变更
    # - 成功执行的正常流程
    

3. **WARNING**（预期内异常）：
    logging.warning("缓存未命中，回退到默认配置")
    # 适用场景：
    # - 可恢复的异常情况
    # - 非关键路径的失败操作
    # - 降级处理情况

4. ERROR（严重错误）：
    try:
        # 可能出错的代码
    except Exception as e:
        logging.error("数据库连接失败", exc_info=True)
    # 适用场景：
    # - 关键操作失败
    # - 不可恢复的异常
    # - 影响核心功能的错误

最佳实践建议：


1. **性能监控**：
    start = time.time()
    # 业务操作
    logging.info(f"操作完成，耗时：{time.time()-start:.2f}s")
    
# 好的日志：
logging.info(f"文件处理成功 [大小：{size}MB] [类型：{file_type}]")

# 通过配置文件动态调整
logging.getLogger().setLevel(logging.DEBUG if DEBUG else logging.INFO)


"""

def setup_logging():
    # 创建日志目录
    log_dir = os.path.join(os.path.dirname(__file__), "logs")
    os.makedirs(log_dir, exist_ok=True)
    
    # 基础配置
    log_format = "%(asctime)s - %(levelname)s - %(filename)s:%(lineno)d - %(message)s"
    date_format = "%Y-%m-%d %H:%M:%S"
    formatter = logging.Formatter(fmt=log_format, datefmt=date_format)

    # 控制台处理器（开发环境使用）
    console_handler = logging.StreamHandler()
    console_handler.setFormatter(formatter)
    console_handler.setLevel(logging.DEBUG)  # 开发时设为DEBUG，生产环境可改为INFO

    # 文件处理器（带轮转功能）
    file_handler = RotatingFileHandler(
        filename=os.path.join(log_dir, "hiviewer.log"),
        maxBytes=10*1024*1024,  # 10MB
        backupCount=5,
        encoding="utf-8"
    )
    file_handler.setFormatter(formatter)
    file_handler.setLevel(logging.INFO)

    # 主日志器配置
    main_logger = logging.getLogger()
    main_logger.setLevel(logging.DEBUG)
    main_logger.addHandler(console_handler)
    main_logger.addHandler(file_handler)

    # 第三方库日志级别调整
    logging.getLogger("PIL").setLevel(logging.WARNING)
    logging.getLogger("urllib3").setLevel(logging.WARNING)
    logging.getLogger("cv2").setLevel(logging.WARNING)

"""
设置日志区域结束线
------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
"""

if __name__ == '__main__':
    print("main()--主界面程序启动")

    # 初始化日志文件
    # setup_logging()  

    # 设置主程序app
    app = QtWidgets.QApplication(sys.argv)
    app_icon = QIcon(os.path.join(os.path.dirname(__file__), "icons", "viewer_3.ico"))
    app.setWindowIcon(app_icon)

    # 设置主界面
    window = HiviewerMainwindow()

    sys.exit(app.exec_())
